"""
routers/store_dispatch.py
=========================
Phase 2 of the Store + Dispatch module.

Scope (locked in chat 2026-05-12):
  - Material + Customer master via Excel import / export (ERP later)
  - Production-linked flow: Store issues raw material to a line, line
    produces FG cycles, Dispatch packs into LOTS of a configured size,
    Loads carry one-or-many lots to a customer.
  - Lot size is per-material (configurable on Material Master row).
  - Same shift timings as production (no separate Store schedule).
  - No barcode scanners on Store/Dispatch yet → manual qty entry.

Tables
------
  mes_materials          master (RM / FG / PKG / CONS)
  mes_customers          customer master (FG buyer)
  mes_store_grn          incoming material (supplier → store)
  mes_store_issues       outgoing material (store → line)
  mes_dispatch_lots      FG packs ready to dispatch
  mes_dispatch_loads     truck loads (1 customer, N lots)

Stock balance is COMPUTED on-the-fly via SUM(grn) − SUM(issue), no
denormalised balance table — keeps the schema honest and lets the
admin post historical corrections without breaking running totals.

Excel I/O
---------
GET  /api/store/materials/template      blank template
POST /api/store/materials/import        bulk upsert from .xlsx
GET  /api/store/materials/export        all materials → .xlsx
(same trio for customers)

CSV alternative supported on import for sites that don't have Excel.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, date
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook

from database import get_conn, dict_cursor
from auth import require_admin, get_current_user

router = APIRouter(prefix="/api/store", tags=["store"])
dispatch_router = APIRouter(prefix="/api/dispatch", tags=["dispatch"])


# ════════════════════════════════════════════════════════════════════
#  Schema
# ════════════════════════════════════════════════════════════════════
def _ensure_tables() -> None:
    """Idempotent.  Called on every endpoint hit; trivial cost."""
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_materials (
                id           SERIAL PRIMARY KEY,
                code         VARCHAR(64) UNIQUE NOT NULL,
                name         VARCHAR(200) NOT NULL,
                mat_type     VARCHAR(10)  NOT NULL DEFAULT 'RM'
                              CHECK (mat_type IN ('RM','FG','PKG','CONS')),
                uom          VARCHAR(16)  NOT NULL DEFAULT 'PCS',
                min_stock    NUMERIC(14,3) NOT NULL DEFAULT 0,
                max_stock    NUMERIC(14,3) NOT NULL DEFAULT 0,
                lot_size     NUMERIC(14,3) NOT NULL DEFAULT 0,
                line_id      INTEGER,
                supplier     VARCHAR(120),
                is_active    BOOLEAN NOT NULL DEFAULT TRUE,
                created_at   TIMESTAMP DEFAULT NOW(),
                updated_at   TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_customers (
                id           SERIAL PRIMARY KEY,
                code         VARCHAR(64) UNIQUE NOT NULL,
                name         VARCHAR(200) NOT NULL,
                address      TEXT,
                contact      VARCHAR(120),
                phone        VARCHAR(40),
                email        VARCHAR(120),
                is_active    BOOLEAN NOT NULL DEFAULT TRUE,
                created_at   TIMESTAMP DEFAULT NOW(),
                updated_at   TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_store_grn (
                id            SERIAL PRIMARY KEY,
                grn_no        VARCHAR(64),
                material_id   INTEGER NOT NULL REFERENCES mes_materials(id),
                qty           NUMERIC(14,3) NOT NULL,
                supplier      VARCHAR(120),
                received_at   TIMESTAMP NOT NULL DEFAULT NOW(),
                received_by   VARCHAR(120),
                shift_date    DATE,
                shift_name    VARCHAR(10),
                remarks       TEXT
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_grn_mat ON mes_store_grn (material_id, received_at)")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_store_issues (
                id            SERIAL PRIMARY KEY,
                material_id   INTEGER NOT NULL REFERENCES mes_materials(id),
                line_id       INTEGER NOT NULL,
                qty           NUMERIC(14,3) NOT NULL,
                issued_at     TIMESTAMP NOT NULL DEFAULT NOW(),
                issued_by     VARCHAR(120),
                shift_date    DATE,
                shift_name    VARCHAR(10),
                remarks       TEXT
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_iss_mat ON mes_store_issues (material_id, issued_at)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_iss_line ON mes_store_issues (line_id, shift_date)")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_dispatch_lots (
                id            SERIAL PRIMARY KEY,
                lot_no        VARCHAR(64) UNIQUE NOT NULL,
                line_id       INTEGER NOT NULL,
                material_id   INTEGER NOT NULL REFERENCES mes_materials(id),
                lot_size      NUMERIC(14,3) NOT NULL,
                qty_packed    NUMERIC(14,3) NOT NULL,
                shift_date    DATE NOT NULL,
                shift_name    VARCHAR(10),
                status        VARCHAR(20) NOT NULL DEFAULT 'READY'
                              CHECK (status IN ('READY','LOADED','DISPATCHED','CANCELLED')),
                packed_at     TIMESTAMP NOT NULL DEFAULT NOW(),
                packed_by     VARCHAR(120),
                load_id       INTEGER,
                remarks       TEXT
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_lot_status ON mes_dispatch_lots (status, line_id, shift_date)")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_dispatch_loads (
                id            SERIAL PRIMARY KEY,
                load_no       VARCHAR(64) UNIQUE NOT NULL,
                customer_id   INTEGER NOT NULL REFERENCES mes_customers(id),
                vehicle_no    VARCHAR(40),
                driver_name   VARCHAR(120),
                driver_phone  VARCHAR(40),
                status        VARCHAR(20) NOT NULL DEFAULT 'PLANNED'
                              CHECK (status IN ('PLANNED','DISPATCHED','CANCELLED')),
                planned_at    TIMESTAMP NOT NULL DEFAULT NOW(),
                planned_by    VARCHAR(120),
                dispatched_at TIMESTAMP,
                dispatched_by VARCHAR(120),
                remarks       TEXT
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_load_status ON mes_dispatch_loads (status, planned_at)")
        conn.commit()


# ════════════════════════════════════════════════════════════════════
#  Helpers
# ════════════════════════════════════════════════════════════════════
def _to_float(v, default=0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _gen_no(prefix: str) -> str:
    """Generate a human-friendly auto-number like 'GRN-26051200001'."""
    return f"{prefix}-{datetime.now().strftime('%y%m%d%H%M%S')}"


def _username(user) -> str:
    if isinstance(user, dict):
        return user.get("username") or "system"
    return getattr(user, "username", "system") or "system"


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ════════════════════════════════════════════════════════════════════
#  Material Master
# ════════════════════════════════════════════════════════════════════
class MaterialUpsert(BaseModel):
    code:       str
    name:       str
    mat_type:   str = "RM"
    uom:        str = "PCS"
    min_stock:  float = 0
    max_stock:  float = 0
    lot_size:   float = 0
    line_id:    Optional[int] = None
    supplier:   Optional[str] = None
    is_active:  bool = True


@router.get("/materials")
def list_materials(mat_type: Optional[str] = None,
                   line_id:  Optional[int] = None,
                   user=Depends(get_current_user)):
    _ensure_tables()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        sql = """SELECT m.*, l.line_name
                   FROM mes_materials m
              LEFT JOIN mes_lines l ON l.id = m.line_id
                  WHERE m.is_active = TRUE"""
        params: list = []
        if mat_type:
            sql += " AND m.mat_type = %s"
            params.append(mat_type.upper())
        if line_id is not None:
            sql += " AND (m.line_id = %s OR m.line_id IS NULL)"
            params.append(line_id)
        sql += " ORDER BY m.mat_type, m.code"
        cur.execute(sql, tuple(params))
        return cur.fetchall()


@router.post("/materials", status_code=201)
def upsert_material(body: MaterialUpsert, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_materials
                (code, name, mat_type, uom, min_stock, max_stock,
                 lot_size, line_id, supplier, is_active, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON CONFLICT (code) DO UPDATE
                SET name      = EXCLUDED.name,
                    mat_type  = EXCLUDED.mat_type,
                    uom       = EXCLUDED.uom,
                    min_stock = EXCLUDED.min_stock,
                    max_stock = EXCLUDED.max_stock,
                    lot_size  = EXCLUDED.lot_size,
                    line_id   = EXCLUDED.line_id,
                    supplier  = EXCLUDED.supplier,
                    is_active = EXCLUDED.is_active,
                    updated_at= NOW()
            RETURNING id
        """, (body.code.strip(), body.name.strip(),
              body.mat_type.upper(), body.uom.upper(),
              body.min_stock, body.max_stock, body.lot_size,
              body.line_id, body.supplier, body.is_active))
        new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id, "ok": True}


@router.delete("/materials/{material_id}")
def delete_material(material_id: int, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE mes_materials SET is_active=FALSE WHERE id=%s", (material_id,))
        conn.commit()
    return {"ok": True}


@router.get("/materials/template")
def materials_template(admin=Depends(require_admin)):
    """Blank .xlsx with the right column headers + 1 sample row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"
    headers = ["code","name","mat_type","uom","min_stock","max_stock",
               "lot_size","line_name","supplier","is_active"]
    ws.append(headers)
    ws.append(["RM-001","Steel Strip 1.2mm","RM","KG",500,2000,0,"","Tata Steel",True])
    ws.append(["FG-YNC-001","Upper Rail Assy","FG","PCS",0,0,50,"YNC-SS","",True])
    # Column widths
    widths = [16,40,10,8,12,12,10,18,20,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    return _xlsx_response(wb, "materials_template.xlsx")


@router.get("/materials/export")
def materials_export(admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT m.code, m.name, m.mat_type, m.uom,
                              m.min_stock, m.max_stock, m.lot_size,
                              l.line_name, m.supplier, m.is_active
                         FROM mes_materials m
                    LEFT JOIN mes_lines l ON l.id = m.line_id
                        ORDER BY m.mat_type, m.code""")
        rows = cur.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"
    headers = ["code","name","mat_type","uom","min_stock","max_stock",
               "lot_size","line_name","supplier","is_active"]
    ws.append(headers)
    for r in rows:
        ws.append([r["code"], r["name"], r["mat_type"], r["uom"],
                   float(r["min_stock"] or 0), float(r["max_stock"] or 0),
                   float(r["lot_size"] or 0),
                   r.get("line_name") or "",
                   r.get("supplier") or "",
                   bool(r["is_active"])])
    widths = [16,40,10,8,12,12,10,18,20,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    return _xlsx_response(wb, f"materials_{date.today().isoformat()}.xlsx")


@router.post("/materials/import")
async def materials_import(file: UploadFile = File(...), admin=Depends(require_admin)):
    """Bulk upsert from .xlsx (columns must match template).
    line_name (text) → line_id resolution is fuzzy: exact match on
    mes_lines.line_name first, then case-insensitive prefix."""
    _ensure_tables()
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not read workbook: {exc}")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty workbook")
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    required = ["code", "name"]
    for col in required:
        if col not in header:
            raise HTTPException(400, f"Missing required column: {col}")
    idx = {h: i for i, h in enumerate(header)}

    inserted = updated = skipped = 0
    errors: List[str] = []
    with get_conn() as conn:
        # Pre-load line_name → id map for fuzzy resolution
        cur = dict_cursor(conn)
        cur.execute("SELECT id, line_name FROM mes_lines")
        lines = cur.fetchall()
        line_by_name = {(l["line_name"] or "").strip().lower(): l["id"] for l in lines}

        cur2 = conn.cursor()
        for ridx, row in enumerate(rows[1:], start=2):
            try:
                code = (row[idx["code"]] or "").strip() if idx.get("code") is not None else ""
                name = (row[idx["name"]] or "").strip() if idx.get("name") is not None else ""
                if not code or not name:
                    skipped += 1
                    continue
                mat_type = (str(row[idx["mat_type"]]).strip().upper()
                            if "mat_type" in idx and row[idx["mat_type"]] else "RM")
                if mat_type not in ("RM","FG","PKG","CONS"):
                    mat_type = "RM"
                uom = (str(row[idx["uom"]]).strip().upper()
                       if "uom" in idx and row[idx["uom"]] else "PCS")
                min_stock = _to_float(row[idx["min_stock"]]) if "min_stock" in idx else 0.0
                max_stock = _to_float(row[idx["max_stock"]]) if "max_stock" in idx else 0.0
                lot_size  = _to_float(row[idx["lot_size"]])  if "lot_size"  in idx else 0.0

                # line_name → line_id
                line_id = None
                if "line_name" in idx and row[idx["line_name"]]:
                    ln = str(row[idx["line_name"]]).strip().lower()
                    if ln in line_by_name:
                        line_id = line_by_name[ln]
                    else:
                        # fuzzy prefix
                        for k, v in line_by_name.items():
                            if k.startswith(ln) or ln.startswith(k):
                                line_id = v; break

                supplier = (str(row[idx["supplier"]]).strip()
                            if "supplier" in idx and row[idx["supplier"]] else None)
                is_active = True
                if "is_active" in idx and row[idx["is_active"]] is not None:
                    v = row[idx["is_active"]]
                    is_active = str(v).strip().lower() not in ("0","false","no","n")

                # Check existing for inserted vs updated count
                cur2.execute("SELECT 1 FROM mes_materials WHERE code=%s", (code,))
                exists = cur2.fetchone() is not None

                cur2.execute("""
                    INSERT INTO mes_materials
                        (code, name, mat_type, uom, min_stock, max_stock,
                         lot_size, line_id, supplier, is_active, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                    ON CONFLICT (code) DO UPDATE
                        SET name=EXCLUDED.name, mat_type=EXCLUDED.mat_type,
                            uom=EXCLUDED.uom, min_stock=EXCLUDED.min_stock,
                            max_stock=EXCLUDED.max_stock, lot_size=EXCLUDED.lot_size,
                            line_id=EXCLUDED.line_id, supplier=EXCLUDED.supplier,
                            is_active=EXCLUDED.is_active, updated_at=NOW()
                """, (code, name, mat_type, uom, min_stock, max_stock,
                      lot_size, line_id, supplier, is_active))
                if exists: updated += 1
                else:      inserted += 1
            except Exception as exc:
                errors.append(f"Row {ridx}: {exc}")
                skipped += 1
        conn.commit()
    return {"ok": True, "inserted": inserted, "updated": updated,
            "skipped": skipped, "errors": errors[:20]}


# ════════════════════════════════════════════════════════════════════
#  Customer Master
# ════════════════════════════════════════════════════════════════════
class CustomerUpsert(BaseModel):
    code:      str
    name:      str
    address:   Optional[str] = None
    contact:   Optional[str] = None
    phone:     Optional[str] = None
    email:     Optional[str] = None
    is_active: bool = True


@dispatch_router.get("/customers")
def list_customers(user=Depends(get_current_user)):
    _ensure_tables()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT * FROM mes_customers WHERE is_active=TRUE ORDER BY name")
        return cur.fetchall()


@dispatch_router.post("/customers", status_code=201)
def upsert_customer(body: CustomerUpsert, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_customers
                (code, name, address, contact, phone, email, is_active, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,NOW())
            ON CONFLICT (code) DO UPDATE
                SET name=EXCLUDED.name, address=EXCLUDED.address,
                    contact=EXCLUDED.contact, phone=EXCLUDED.phone,
                    email=EXCLUDED.email, is_active=EXCLUDED.is_active,
                    updated_at=NOW()
            RETURNING id
        """, (body.code.strip(), body.name.strip(),
              body.address, body.contact, body.phone, body.email,
              body.is_active))
        new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id, "ok": True}


@dispatch_router.delete("/customers/{customer_id}")
def delete_customer(customer_id: int, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE mes_customers SET is_active=FALSE WHERE id=%s", (customer_id,))
        conn.commit()
    return {"ok": True}


@dispatch_router.get("/customers/template")
def customers_template(admin=Depends(require_admin)):
    wb = Workbook(); ws = wb.active; ws.title = "Customers"
    headers = ["code","name","address","contact","phone","email","is_active"]
    ws.append(headers)
    ws.append(["CUST-001","Maruti Suzuki India Ltd",
               "Plot 1, Sector 18, Gurugram", "Mr Sharma",
               "+91-9876543210", "purchase@maruti.example", True])
    widths = [12,40,50,18,18,28,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    return _xlsx_response(wb, "customers_template.xlsx")


@dispatch_router.get("/customers/export")
def customers_export(admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT code, name, address, contact, phone, email, is_active
                         FROM mes_customers ORDER BY name""")
        rows = cur.fetchall()
    wb = Workbook(); ws = wb.active; ws.title = "Customers"
    headers = ["code","name","address","contact","phone","email","is_active"]
    ws.append(headers)
    for r in rows:
        ws.append([r["code"], r["name"], r.get("address") or "",
                   r.get("contact") or "", r.get("phone") or "",
                   r.get("email") or "", bool(r["is_active"])])
    widths = [12,40,50,18,18,28,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    return _xlsx_response(wb, f"customers_{date.today().isoformat()}.xlsx")


@dispatch_router.post("/customers/import")
async def customers_import(file: UploadFile = File(...), admin=Depends(require_admin)):
    _ensure_tables()
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not read workbook: {exc}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty workbook")
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "code" not in header or "name" not in header:
        raise HTTPException(400, "Required columns: code, name")
    idx = {h: i for i, h in enumerate(header)}

    inserted = updated = skipped = 0
    errors: List[str] = []
    with get_conn() as conn:
        cur = conn.cursor()
        for ridx, row in enumerate(rows[1:], start=2):
            try:
                code = (row[idx["code"]] or "").strip() if idx.get("code") is not None else ""
                name = (row[idx["name"]] or "").strip() if idx.get("name") is not None else ""
                if not code or not name:
                    skipped += 1; continue
                address = (str(row[idx["address"]]).strip()
                           if "address" in idx and row[idx["address"]] else None)
                contact = (str(row[idx["contact"]]).strip()
                           if "contact" in idx and row[idx["contact"]] else None)
                phone   = (str(row[idx["phone"]]).strip()
                           if "phone" in idx and row[idx["phone"]] else None)
                email   = (str(row[idx["email"]]).strip()
                           if "email" in idx and row[idx["email"]] else None)
                is_active = True
                if "is_active" in idx and row[idx["is_active"]] is not None:
                    v = row[idx["is_active"]]
                    is_active = str(v).strip().lower() not in ("0","false","no","n")

                cur.execute("SELECT 1 FROM mes_customers WHERE code=%s", (code,))
                exists = cur.fetchone() is not None
                cur.execute("""
                    INSERT INTO mes_customers
                        (code, name, address, contact, phone, email, is_active, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,NOW())
                    ON CONFLICT (code) DO UPDATE
                        SET name=EXCLUDED.name, address=EXCLUDED.address,
                            contact=EXCLUDED.contact, phone=EXCLUDED.phone,
                            email=EXCLUDED.email, is_active=EXCLUDED.is_active,
                            updated_at=NOW()
                """, (code, name, address, contact, phone, email, is_active))
                if exists: updated += 1
                else:      inserted += 1
            except Exception as exc:
                errors.append(f"Row {ridx}: {exc}")
                skipped += 1
        conn.commit()
    return {"ok": True, "inserted": inserted, "updated": updated,
            "skipped": skipped, "errors": errors[:20]}


# ════════════════════════════════════════════════════════════════════
#  Store: GRN (inbound) + Issue (outbound)
# ════════════════════════════════════════════════════════════════════
class GRNBody(BaseModel):
    grn_no:      Optional[str] = None
    material_id: int
    qty:         float
    supplier:    Optional[str] = None
    received_at: Optional[str] = None     # ISO; default NOW
    shift_date:  Optional[str] = None
    shift_name:  Optional[str] = None
    remarks:     Optional[str] = None


@router.post("/grn", status_code=201)
def add_grn(body: GRNBody, user=Depends(get_current_user)):
    _ensure_tables()
    if body.qty <= 0:
        raise HTTPException(400, "qty must be > 0")
    rec_at = (datetime.fromisoformat(body.received_at)
              if body.received_at else datetime.now())
    sh_date = (datetime.strptime(body.shift_date, "%Y-%m-%d").date()
               if body.shift_date else rec_at.date())
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_store_grn
                (grn_no, material_id, qty, supplier,
                 received_at, received_by, shift_date, shift_name, remarks)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (body.grn_no or _gen_no("GRN"),
              body.material_id, body.qty, body.supplier,
              rec_at, _username(user), sh_date,
              body.shift_name, body.remarks))
        new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id, "ok": True}


@router.get("/grn")
def list_grn(material_id: Optional[int] = None,
             date_from:    Optional[str] = None,
             date_to:      Optional[str] = None,
             user=Depends(get_current_user)):
    _ensure_tables()
    sql = """SELECT g.*, m.code AS material_code, m.name AS material_name, m.uom
               FROM mes_store_grn g
               JOIN mes_materials m ON m.id = g.material_id
              WHERE 1=1"""
    params: list = []
    if material_id is not None:
        sql += " AND g.material_id = %s"; params.append(material_id)
    if date_from:
        sql += " AND g.received_at >= %s"; params.append(date_from)
    if date_to:
        sql += " AND g.received_at <= %s"; params.append(date_to + " 23:59:59")
    sql += " ORDER BY g.received_at DESC LIMIT 500"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


class IssueBody(BaseModel):
    material_id: int
    line_id:     int
    qty:         float
    issued_at:   Optional[str] = None
    shift_date:  Optional[str] = None
    shift_name:  Optional[str] = None
    remarks:     Optional[str] = None


@router.post("/issues", status_code=201)
def add_issue(body: IssueBody, user=Depends(get_current_user)):
    _ensure_tables()
    if body.qty <= 0:
        raise HTTPException(400, "qty must be > 0")
    iss_at = (datetime.fromisoformat(body.issued_at)
              if body.issued_at else datetime.now())
    sh_date = (datetime.strptime(body.shift_date, "%Y-%m-%d").date()
               if body.shift_date else iss_at.date())
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_store_issues
                (material_id, line_id, qty, issued_at, issued_by,
                 shift_date, shift_name, remarks)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (body.material_id, body.line_id, body.qty,
              iss_at, _username(user), sh_date,
              body.shift_name, body.remarks))
        new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id, "ok": True}


@router.get("/issues")
def list_issues(material_id: Optional[int] = None,
                line_id:     Optional[int] = None,
                date_from:   Optional[str] = None,
                date_to:     Optional[str] = None,
                user=Depends(get_current_user)):
    _ensure_tables()
    sql = """SELECT i.*, m.code AS material_code, m.name AS material_name, m.uom,
                    l.line_name
               FROM mes_store_issues i
               JOIN mes_materials m ON m.id = i.material_id
          LEFT JOIN mes_lines l    ON l.id = i.line_id
              WHERE 1=1"""
    params: list = []
    if material_id is not None:
        sql += " AND i.material_id = %s"; params.append(material_id)
    if line_id is not None:
        sql += " AND i.line_id = %s"; params.append(line_id)
    if date_from:
        sql += " AND i.issued_at >= %s"; params.append(date_from)
    if date_to:
        sql += " AND i.issued_at <= %s"; params.append(date_to + " 23:59:59")
    sql += " ORDER BY i.issued_at DESC LIMIT 500"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


# ── Stock balance (computed: SUM(grn) − SUM(issue)) ──────────────────
@router.get("/stock")
def stock_balance(mat_type: Optional[str] = None,
                  user=Depends(get_current_user)):
    """Current on-hand for every material — computed from GRN − Issues."""
    _ensure_tables()
    sql = """
        SELECT m.id, m.code, m.name, m.mat_type, m.uom,
               m.min_stock, m.max_stock, m.lot_size,
               m.supplier, l.line_name,
               COALESCE(g.in_qty, 0)::float  AS in_qty,
               COALESCE(i.out_qty, 0)::float AS out_qty,
               (COALESCE(g.in_qty, 0) - COALESCE(i.out_qty, 0))::float AS balance,
               g.last_in_at, i.last_out_at
          FROM mes_materials m
     LEFT JOIN mes_lines l ON l.id = m.line_id
     LEFT JOIN (SELECT material_id, SUM(qty) AS in_qty, MAX(received_at) AS last_in_at
                  FROM mes_store_grn GROUP BY material_id) g
            ON g.material_id = m.id
     LEFT JOIN (SELECT material_id, SUM(qty) AS out_qty, MAX(issued_at) AS last_out_at
                  FROM mes_store_issues GROUP BY material_id) i
            ON i.material_id = m.id
         WHERE m.is_active = TRUE
    """
    params: list = []
    if mat_type:
        sql += " AND m.mat_type = %s"
        params.append(mat_type.upper())
    sql += " ORDER BY m.mat_type, m.code"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()
    # Annotate status
    for r in rows:
        bal = float(r["balance"] or 0)
        mn  = float(r["min_stock"] or 0)
        mx  = float(r["max_stock"] or 0)
        if mn > 0 and bal <= mn:        r["status"] = "LOW"
        elif mx > 0 and bal >= mx:      r["status"] = "HIGH"
        elif bal <= 0:                  r["status"] = "OUT"
        else:                           r["status"] = "OK"
    return rows


# ════════════════════════════════════════════════════════════════════
#  Dispatch: Lots + Loads
# ════════════════════════════════════════════════════════════════════

class LotBody(BaseModel):
    line_id:     int
    material_id: int
    lot_size:    Optional[float] = None  # if None, takes material.lot_size
    qty_packed:  float                   # actual qty in this box
    shift_date:  Optional[str] = None
    shift_name:  Optional[str] = None
    remarks:     Optional[str] = None


@dispatch_router.post("/lots", status_code=201)
def add_lot(body: LotBody, user=Depends(get_current_user)):
    _ensure_tables()
    if body.qty_packed <= 0:
        raise HTTPException(400, "qty_packed must be > 0")
    sh_date = (datetime.strptime(body.shift_date, "%Y-%m-%d").date()
               if body.shift_date else date.today())
    # Default lot_size from material if not provided
    lot_size = body.lot_size
    if lot_size is None:
        with get_conn() as conn:
            cur = dict_cursor(conn)
            cur.execute("SELECT lot_size FROM mes_materials WHERE id=%s", (body.material_id,))
            r = cur.fetchone()
            lot_size = float(r["lot_size"]) if r and r["lot_size"] else body.qty_packed
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_dispatch_lots
                (lot_no, line_id, material_id, lot_size, qty_packed,
                 shift_date, shift_name, status, packed_by, remarks)
            VALUES (%s,%s,%s,%s,%s,%s,%s,'READY',%s,%s)
            RETURNING id
        """, (_gen_no("LOT"), body.line_id, body.material_id,
              lot_size, body.qty_packed, sh_date,
              body.shift_name, _username(user), body.remarks))
        new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id, "ok": True}


@dispatch_router.get("/lots")
def list_lots(status:     Optional[str] = None,
              line_id:    Optional[int] = None,
              shift_date: Optional[str] = None,
              user=Depends(get_current_user)):
    _ensure_tables()
    sql = """SELECT lt.*, m.code AS material_code, m.name AS material_name, m.uom,
                    l.line_name,
                    ld.load_no, ld.status AS load_status
               FROM mes_dispatch_lots lt
               JOIN mes_materials m ON m.id = lt.material_id
          LEFT JOIN mes_lines l    ON l.id = lt.line_id
          LEFT JOIN mes_dispatch_loads ld ON ld.id = lt.load_id
              WHERE 1=1"""
    params: list = []
    if status:
        sql += " AND lt.status = %s"; params.append(status.upper())
    if line_id is not None:
        sql += " AND lt.line_id = %s"; params.append(line_id)
    if shift_date:
        sql += " AND lt.shift_date = %s"; params.append(shift_date)
    sql += " ORDER BY lt.packed_at DESC LIMIT 500"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


@dispatch_router.post("/lots/{lot_id}/cancel")
def cancel_lot(lot_id: int, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""UPDATE mes_dispatch_lots
                          SET status='CANCELLED'
                        WHERE id=%s AND status IN ('READY','LOADED')""",
                    (lot_id,))
        conn.commit()
    return {"ok": True}


# ── Loads ─────────────────────────────────────────────────────────────
class LoadCreateBody(BaseModel):
    customer_id:  int
    vehicle_no:   Optional[str] = None
    driver_name:  Optional[str] = None
    driver_phone: Optional[str] = None
    lot_ids:      List[int]    = []
    remarks:      Optional[str] = None


@dispatch_router.post("/loads", status_code=201)
def create_load(body: LoadCreateBody, user=Depends(get_current_user)):
    """Create a planned load and attach any number of READY lots to it.
    Attached lots flip to LOADED."""
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_dispatch_loads
                (load_no, customer_id, vehicle_no, driver_name,
                 driver_phone, status, planned_by, remarks)
            VALUES (%s,%s,%s,%s,%s,'PLANNED',%s,%s)
            RETURNING id
        """, (_gen_no("LOAD"), body.customer_id, body.vehicle_no,
              body.driver_name, body.driver_phone,
              _username(user), body.remarks))
        load_id = cur.fetchone()[0]
        # Attach lots
        if body.lot_ids:
            cur.execute("""
                UPDATE mes_dispatch_lots
                   SET load_id=%s, status='LOADED'
                 WHERE id = ANY(%s) AND status='READY'
            """, (load_id, body.lot_ids))
        conn.commit()
    return {"id": load_id, "ok": True}


@dispatch_router.post("/loads/{load_id}/dispatch")
def dispatch_load(load_id: int, user=Depends(get_current_user)):
    """Mark a load as DISPATCHED — gate-pass equivalent.
    Cascades to its attached lots."""
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""UPDATE mes_dispatch_loads
                          SET status='DISPATCHED',
                              dispatched_at=NOW(),
                              dispatched_by=%s
                        WHERE id=%s AND status='PLANNED'""",
                    (_username(user), load_id))
        cur.execute("""UPDATE mes_dispatch_lots
                          SET status='DISPATCHED'
                        WHERE load_id=%s AND status='LOADED'""",
                    (load_id,))
        conn.commit()
    return {"ok": True}


@dispatch_router.post("/loads/{load_id}/cancel")
def cancel_load(load_id: int, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        # Free attached lots back to READY
        cur.execute("""UPDATE mes_dispatch_lots
                          SET load_id=NULL, status='READY'
                        WHERE load_id=%s AND status='LOADED'""",
                    (load_id,))
        cur.execute("""UPDATE mes_dispatch_loads
                          SET status='CANCELLED'
                        WHERE id=%s AND status='PLANNED'""",
                    (load_id,))
        conn.commit()
    return {"ok": True}


@dispatch_router.get("/loads")
def list_loads(status:    Optional[str] = None,
               date_from: Optional[str] = None,
               date_to:   Optional[str] = None,
               user=Depends(get_current_user)):
    _ensure_tables()
    sql = """SELECT ld.*, c.code AS customer_code, c.name AS customer_name,
                    (SELECT COUNT(*) FROM mes_dispatch_lots WHERE load_id=ld.id) AS lot_count,
                    (SELECT COALESCE(SUM(qty_packed),0)
                       FROM mes_dispatch_lots WHERE load_id=ld.id)::float AS total_qty
               FROM mes_dispatch_loads ld
               JOIN mes_customers c ON c.id = ld.customer_id
              WHERE 1=1"""
    params: list = []
    if status:
        sql += " AND ld.status = %s"; params.append(status.upper())
    if date_from:
        sql += " AND ld.planned_at >= %s"; params.append(date_from)
    if date_to:
        sql += " AND ld.planned_at <= %s"; params.append(date_to + " 23:59:59")
    sql += " ORDER BY ld.planned_at DESC LIMIT 500"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


@dispatch_router.get("/loads/{load_id}")
def load_detail(load_id: int, user=Depends(get_current_user)):
    _ensure_tables()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT ld.*, c.code AS customer_code, c.name AS customer_name,
                              c.address AS customer_address
                         FROM mes_dispatch_loads ld
                         JOIN mes_customers c ON c.id = ld.customer_id
                        WHERE ld.id = %s""", (load_id,))
        load = cur.fetchone()
        if not load:
            raise HTTPException(404, "load not found")
        cur.execute("""SELECT lt.*, m.code AS material_code, m.name AS material_name,
                              m.uom, l.line_name
                         FROM mes_dispatch_lots lt
                         JOIN mes_materials m ON m.id = lt.material_id
                    LEFT JOIN mes_lines l ON l.id = lt.line_id
                        WHERE lt.load_id = %s
                        ORDER BY lt.packed_at""", (load_id,))
        lots = cur.fetchall()
    return {"load": load, "lots": lots}


# ════════════════════════════════════════════════════════════════════
#  Dashboard summaries (cheap roll-ups for landing tiles)
# ════════════════════════════════════════════════════════════════════
@router.get("/dashboard")
def store_dashboard(user=Depends(get_current_user)):
    """Compact stats for the Store landing page."""
    _ensure_tables()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT COUNT(*)::int AS total_materials
              FROM mes_materials WHERE is_active = TRUE
        """)
        total = cur.fetchone()
        cur.execute("""
            SELECT COALESCE(SUM(qty),0)::float AS in_today
              FROM mes_store_grn WHERE received_at::date = CURRENT_DATE
        """)
        in_today = cur.fetchone()
        cur.execute("""
            SELECT COALESCE(SUM(qty),0)::float AS out_today
              FROM mes_store_issues WHERE issued_at::date = CURRENT_DATE
        """)
        out_today = cur.fetchone()
    stock = stock_balance(user=user)
    low  = [s for s in stock if s["status"] in ("LOW", "OUT")]
    high = [s for s in stock if s["status"] == "HIGH"]
    return {
        "total_materials": total["total_materials"],
        "in_today":  in_today["in_today"],
        "out_today": out_today["out_today"],
        "low_count":  len(low),
        "high_count": len(high),
        "low_materials":  low[:10],
        "high_materials": high[:10],
    }


@dispatch_router.get("/dashboard")
def dispatch_dashboard(user=Depends(get_current_user)):
    """Compact stats for the Dispatch landing page."""
    _ensure_tables()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT COUNT(*)::int AS ready_lots,
                              COALESCE(SUM(qty_packed),0)::float AS ready_qty
                         FROM mes_dispatch_lots WHERE status='READY'""")
        ready = cur.fetchone()
        cur.execute("""SELECT COUNT(*)::int AS planned_loads
                         FROM mes_dispatch_loads WHERE status='PLANNED'""")
        planned = cur.fetchone()
        cur.execute("""SELECT COUNT(*)::int AS dispatched_today,
                              COALESCE(SUM((SELECT SUM(qty_packed) FROM mes_dispatch_lots WHERE load_id=ld.id)),0)::float AS qty_today
                         FROM mes_dispatch_loads ld
                        WHERE dispatched_at::date = CURRENT_DATE""")
        today = cur.fetchone()
    return {
        "ready_lots": ready["ready_lots"],
        "ready_qty":  ready["ready_qty"],
        "planned_loads":     planned["planned_loads"],
        "dispatched_today":  today["dispatched_today"],
        "qty_today":         today["qty_today"],
    }
