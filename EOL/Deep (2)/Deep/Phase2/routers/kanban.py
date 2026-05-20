"""
routers/kanban.py
=================
TPS-style Kanban / Dispatch tracking for finished-goods (FG) parts.

Source-of-truth for FG master + monthly plan is the operator's existing
Excel: "KANBAN POST 2025-26".  One sheet per month (APR-26, MAY-26, …)
with per-part rows × 3 dispatch windows per day (12:00, Shift A, Shift B).

Workflow:
  1. Admin imports Excel (specify which month sheet) → mes_fg_parts +
     mes_monthly_plan populated.
  2. Operator links each PLC model_number → an FG part once (admin UI).
  3. Auto-fire watcher runs every 60 s.  When the wall clock crosses
     a window boundary (12:00 / shift A end / shift B end), it:
       - groups cycles in [prev_window, this_window] by model_number
       - for each model that has a linked FG part, computes
         kanban_count = floor(produced / packing_std_qty) and
         dispatch_qty = kanban_count * packing_std_qty
       - inserts a row into mes_kanban_log
  4. Dashboard shows today's posts per part × window plus monthly
     plan vs achieved.

Tables
------
  mes_fg_parts          FG part master (from Excel)
  mes_monthly_plan      per-part per-month shift-A/B plan
  mes_fg_model_link     PLC model_number → FG part id
  mes_kanban_log        auto-fired window logs (kanban + dispatch qty)
"""
from __future__ import annotations

import io
import re
import threading
from datetime import datetime, date, timedelta, time as dt_time
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook

from database import get_conn, dict_cursor
from auth import require_admin, get_current_user

router = APIRouter(prefix="/api/kanban", tags=["kanban"])


# ════════════════════════════════════════════════════════════════════
#  Schema
# ════════════════════════════════════════════════════════════════════
def _ensure_tables() -> None:
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_fg_parts (
                id                  SERIAL PRIMARY KEY,
                tbdi_part_no        VARCHAR(64) UNIQUE NOT NULL,
                tbdi_new_part_no    VARCHAR(64),
                customer_part_no    VARCHAR(64),
                description         TEXT,
                model               VARCHAR(120),
                packing_std_qty     INTEGER NOT NULL DEFAULT 1,
                line_id             INTEGER,
                is_active           BOOLEAN NOT NULL DEFAULT TRUE,
                created_at          TIMESTAMP DEFAULT NOW(),
                updated_at          TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_monthly_plan (
                id            SERIAL PRIMARY KEY,
                fg_part_id    INTEGER NOT NULL REFERENCES mes_fg_parts(id),
                year_month    CHAR(7) NOT NULL,    -- 'YYYY-MM'
                shift_a_plan  INTEGER NOT NULL DEFAULT 0,
                shift_b_plan  INTEGER NOT NULL DEFAULT 0,
                total_plan    INTEGER NOT NULL DEFAULT 0,
                updated_at    TIMESTAMP DEFAULT NOW(),
                UNIQUE (fg_part_id, year_month)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_fg_model_link (
                id            SERIAL PRIMARY KEY,
                line_id       INTEGER NOT NULL,
                model_number  INTEGER NOT NULL,
                fg_part_id    INTEGER NOT NULL REFERENCES mes_fg_parts(id),
                UNIQUE (line_id, model_number)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_kanban_log (
                id              SERIAL PRIMARY KEY,
                fg_part_id      INTEGER NOT NULL REFERENCES mes_fg_parts(id),
                line_id         INTEGER,
                log_date        DATE NOT NULL,
                window_name     VARCHAR(20) NOT NULL,    -- '12PM' | 'SHIFT_A' | 'SHIFT_B'
                cycles_produced INTEGER NOT NULL DEFAULT 0,
                kanban_count    INTEGER NOT NULL DEFAULT 0,
                dispatch_qty    INTEGER NOT NULL DEFAULT 0,
                fired_at        TIMESTAMP NOT NULL DEFAULT NOW(),
                fired_by        VARCHAR(32) NOT NULL DEFAULT 'auto',  -- 'auto' | 'manual'
                notes           TEXT,
                UNIQUE (fg_part_id, log_date, window_name)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_kanban_log_date ON mes_kanban_log (log_date, line_id)")
        conn.commit()


# ════════════════════════════════════════════════════════════════════
#  FG Part Master
# ════════════════════════════════════════════════════════════════════
class FgPartUpsert(BaseModel):
    tbdi_part_no:      str
    tbdi_new_part_no:  Optional[str] = None
    customer_part_no:  Optional[str] = None
    description:       Optional[str] = None
    model:             Optional[str] = None
    packing_std_qty:   int = 1
    line_id:           Optional[int] = None
    is_active:         bool = True


@router.get("/parts")
def list_fg_parts(line_id: Optional[int] = None, user=Depends(get_current_user)):
    _ensure_tables()
    sql = """SELECT p.*, l.line_name
               FROM mes_fg_parts p
          LEFT JOIN mes_lines l ON l.id = p.line_id
              WHERE p.is_active = TRUE"""
    params: list = []
    if line_id is not None:
        sql += " AND p.line_id = %s"
        params.append(line_id)
    sql += " ORDER BY p.model, p.tbdi_part_no"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


@router.post("/parts", status_code=201)
def upsert_fg_part(body: FgPartUpsert, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_fg_parts
                (tbdi_part_no, tbdi_new_part_no, customer_part_no,
                 description, model, packing_std_qty, line_id, is_active, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON CONFLICT (tbdi_part_no) DO UPDATE
                SET tbdi_new_part_no = EXCLUDED.tbdi_new_part_no,
                    customer_part_no = EXCLUDED.customer_part_no,
                    description      = EXCLUDED.description,
                    model            = EXCLUDED.model,
                    packing_std_qty  = EXCLUDED.packing_std_qty,
                    line_id          = EXCLUDED.line_id,
                    is_active        = EXCLUDED.is_active,
                    updated_at       = NOW()
            RETURNING id
        """, (body.tbdi_part_no.strip(), body.tbdi_new_part_no,
              body.customer_part_no, body.description, body.model,
              max(1, body.packing_std_qty), body.line_id, body.is_active))
        new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id, "ok": True}


@router.delete("/parts/{part_id}")
def delete_fg_part(part_id: int, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE mes_fg_parts SET is_active=FALSE WHERE id=%s", (part_id,))
        conn.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════
#  Excel import — handles the KANBAN POST sheet layout
# ════════════════════════════════════════════════════════════════════
EXCEL_HEADER_ROW = 3      # row 3 has column names
EXCEL_DATA_START = 4      # rows 4+ have data
EXPECTED_COLS = {
    "description":      "PART DESCRIPTION",
    "tbdi_part_no":     "TBDI PART NO.",
    "tbdi_new_part_no": "TBDI NEW",                 # matches "TBDI NEW  PART NO."
    "customer_part_no": "CUSTOMER PART NO.",
    "model":            "MODEL",
    "packing_std_qty":  "PACKING STANDARD QTY",
    "shift_a_plan":     "MONTHLY PLAN SHIFT A",
    "shift_b_plan":     "MONTHLY PLAN SHIFT B",
    "total_plan":       "MONTHLY TOTAL PLAN",
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


def _find_col_indices(header_row) -> dict:
    """Find column indices for each expected field by fuzzy match."""
    idx = {}
    for ci, cell in enumerate(header_row):
        nm = _norm(cell)
        if not nm:
            continue
        for field, needle in EXPECTED_COLS.items():
            if field in idx:
                continue
            if _norm(needle) in nm:
                idx[field] = ci
                break
    return idx


@router.get("/parts/template")
def parts_template(admin=Depends(require_admin)):
    """Blank XLSX template matching the import format."""
    wb = Workbook(); ws = wb.active; ws.title = "FG Parts"
    headers = ["PART DESCRIPTION", "TBDI PART NO.", "TBDI NEW PART NO.",
               "CUSTOMER PART NO.", "MODEL", "PACKING STANDARD QTY.",
               "MONTHLY PLAN SHIFT A", "MONTHLY PLAN SHIFT B",
               "MONTHLY TOTAL PLAN A+B", "year_month (YYYY-MM)",
               "line_name (optional)"]
    ws.append(headers)
    ws.append(["YNC 4WAY INNER RH", "72130-X7P03-YNC", "72130-X7P03-YNC",
               "AYY8086", "SEAT SLIDER", 240, 4320, 0, 4320, "2026-05", "YNC-SS"])
    for i, w in enumerate([45, 22, 22, 20, 18, 12, 18, 18, 18, 16, 18], 1):
        ws.column_dimensions[chr(64+i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="kanban_parts_template.xlsx"'})


class KanbanImportResult(BaseModel):
    sheets:       List[str]
    inserted:     int
    updated:      int
    plan_rows:    int
    skipped:      int
    errors:       List[str]


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    sheets: Optional[str] = Query(None,
        description="Comma-separated sheet names to import (default: all monthly sheets)"),
    line_name: Optional[str] = Query(None,
        description="Default line_name to attach to imported parts"),
    admin=Depends(require_admin),
):
    """Import the KANBAN POST Excel.  Each monthly sheet is parsed for
    FG part master + monthly plan.  Sheet name → year_month resolution:
      'APR-26' → '2026-04', 'MAY-26' → '2026-05', etc.
    """
    _ensure_tables()
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not read workbook: {exc}")

    # Decide which sheets to process
    if sheets:
        wanted = [s.strip() for s in sheets.split(",") if s.strip()]
    else:
        # All sheets matching MMM-YY pattern
        wanted = [s for s in wb.sheetnames
                  if re.match(r"[A-Z]{3}-\d{2}", s.strip().upper())]
    if not wanted:
        raise HTTPException(400, "No monthly sheets found / specified")

    # Resolve line_id from name
    line_id = None
    if line_name:
        with get_conn() as conn:
            cur = dict_cursor(conn)
            cur.execute("SELECT id FROM mes_lines WHERE LOWER(line_name)=LOWER(%s)",
                        (line_name.strip(),))
            r = cur.fetchone()
            line_id = r["id"] if r else None

    MONTH_MAP = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,
                 "JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}
    inserted = updated = plan_rows = skipped = 0
    errors: List[str] = []
    processed_sheets = []

    with get_conn() as conn:
        cur = conn.cursor()
        for sname in wanted:
            sname_clean = sname.strip()
            if sname_clean not in wb.sheetnames:
                errors.append(f"Sheet '{sname_clean}' not found")
                continue
            ws = wb[sname_clean]
            # Resolve sheet name → year_month
            m = re.match(r"([A-Z]{3})-(\d{2})", sname_clean.upper())
            if not m:
                errors.append(f"Sheet '{sname_clean}' not a monthly sheet")
                continue
            mon = MONTH_MAP.get(m.group(1))
            yr  = 2000 + int(m.group(2))
            if not mon:
                errors.append(f"Unknown month token in '{sname_clean}'")
                continue
            year_month = f"{yr:04d}-{mon:02d}"

            # Find header columns
            try:
                header_row = next(ws.iter_rows(min_row=EXCEL_HEADER_ROW,
                                                max_row=EXCEL_HEADER_ROW,
                                                values_only=True))
            except StopIteration:
                errors.append(f"Sheet '{sname_clean}' has no header row")
                continue
            col_idx = _find_col_indices(header_row)
            if "tbdi_part_no" not in col_idx or "description" not in col_idx:
                errors.append(f"Sheet '{sname_clean}' missing required columns")
                continue

            # Parse rows
            for ri, row in enumerate(
                    ws.iter_rows(min_row=EXCEL_DATA_START, values_only=True),
                    start=EXCEL_DATA_START):
                try:
                    tbdi = str(row[col_idx["tbdi_part_no"]] or "").strip()
                    desc = str(row[col_idx["description"]] or "").strip()
                    if not tbdi or not desc:
                        continue
                    tbdi_new = (str(row[col_idx["tbdi_new_part_no"]] or "").strip()
                                if "tbdi_new_part_no" in col_idx else "")
                    cust_pn  = (str(row[col_idx["customer_part_no"]] or "").strip()
                                if "customer_part_no" in col_idx else "")
                    model    = (str(row[col_idx["model"]] or "").strip()
                                if "model" in col_idx else "")
                    pack_qty = int(row[col_idx["packing_std_qty"]] or 1) if "packing_std_qty" in col_idx else 1
                    plan_a   = int(row[col_idx["shift_a_plan"]] or 0) if "shift_a_plan" in col_idx else 0
                    plan_b   = int(row[col_idx["shift_b_plan"]] or 0) if "shift_b_plan" in col_idx else 0
                    plan_t   = int(row[col_idx["total_plan"]] or 0) if "total_plan" in col_idx else (plan_a + plan_b)

                    # Upsert mes_fg_parts
                    cur.execute("SELECT id FROM mes_fg_parts WHERE tbdi_part_no=%s", (tbdi,))
                    er = cur.fetchone()
                    cur.execute("""
                        INSERT INTO mes_fg_parts
                            (tbdi_part_no, tbdi_new_part_no, customer_part_no,
                             description, model, packing_std_qty, line_id,
                             is_active, updated_at)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,TRUE,NOW())
                        ON CONFLICT (tbdi_part_no) DO UPDATE
                            SET tbdi_new_part_no = EXCLUDED.tbdi_new_part_no,
                                customer_part_no = EXCLUDED.customer_part_no,
                                description      = EXCLUDED.description,
                                model            = EXCLUDED.model,
                                packing_std_qty  = EXCLUDED.packing_std_qty,
                                line_id          = COALESCE(EXCLUDED.line_id, mes_fg_parts.line_id),
                                updated_at       = NOW()
                        RETURNING id
                    """, (tbdi, tbdi_new or None, cust_pn or None,
                          desc, model or None, max(1, pack_qty), line_id))
                    fg_id = cur.fetchone()[0]
                    if er: updated += 1
                    else:  inserted += 1

                    # Upsert monthly plan
                    cur.execute("""
                        INSERT INTO mes_monthly_plan
                            (fg_part_id, year_month, shift_a_plan, shift_b_plan, total_plan, updated_at)
                        VALUES (%s,%s,%s,%s,%s,NOW())
                        ON CONFLICT (fg_part_id, year_month) DO UPDATE
                            SET shift_a_plan = EXCLUDED.shift_a_plan,
                                shift_b_plan = EXCLUDED.shift_b_plan,
                                total_plan   = EXCLUDED.total_plan,
                                updated_at   = NOW()
                    """, (fg_id, year_month, plan_a, plan_b, plan_t))
                    plan_rows += 1
                except Exception as exc:
                    errors.append(f"{sname_clean} row {ri}: {exc}")
                    skipped += 1
            processed_sheets.append(sname_clean)
        conn.commit()

    return {
        "sheets":   processed_sheets,
        "inserted": inserted,
        "updated":  updated,
        "plan_rows": plan_rows,
        "skipped":  skipped,
        "errors":   errors[:30],
    }


# ════════════════════════════════════════════════════════════════════
#  Monthly Plan
# ════════════════════════════════════════════════════════════════════
@router.get("/monthly-plan")
def list_monthly_plan(year_month: str = Query(...),
                       line_id: Optional[int] = None,
                       user=Depends(get_current_user)):
    _ensure_tables()
    if not re.match(r"\d{4}-\d{2}", year_month):
        raise HTTPException(400, "year_month must be YYYY-MM")
    sql = """SELECT mp.*, p.tbdi_part_no, p.tbdi_new_part_no,
                    p.customer_part_no, p.description, p.model,
                    p.packing_std_qty, p.line_id, l.line_name
               FROM mes_monthly_plan mp
               JOIN mes_fg_parts p ON p.id = mp.fg_part_id
          LEFT JOIN mes_lines l ON l.id = p.line_id
              WHERE mp.year_month = %s AND p.is_active = TRUE"""
    params = [year_month]
    if line_id is not None:
        sql += " AND p.line_id = %s"
        params.append(line_id)
    sql += " ORDER BY p.model, p.tbdi_part_no"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


# ════════════════════════════════════════════════════════════════════
#  Model → FG link (admin)
# ════════════════════════════════════════════════════════════════════
class ModelLinkBody(BaseModel):
    line_id:      int
    model_number: int
    fg_part_id:   int


@router.get("/model-links")
def list_model_links(line_id: Optional[int] = None, user=Depends(get_current_user)):
    _ensure_tables()
    sql = """SELECT ml.*, mm.model_name,
                    p.tbdi_part_no, p.description, p.packing_std_qty
               FROM mes_fg_model_link ml
               JOIN mes_fg_parts p ON p.id = ml.fg_part_id
          LEFT JOIN mes_model_mappings mm
                 ON mm.line_id = ml.line_id AND mm.model_number = ml.model_number
              WHERE 1=1"""
    params: list = []
    if line_id is not None:
        sql += " AND ml.line_id = %s"
        params.append(line_id)
    sql += " ORDER BY ml.line_id, ml.model_number"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


@router.post("/model-links", status_code=201)
def upsert_model_link(body: ModelLinkBody, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_fg_model_link (line_id, model_number, fg_part_id)
            VALUES (%s,%s,%s)
            ON CONFLICT (line_id, model_number)
                DO UPDATE SET fg_part_id = EXCLUDED.fg_part_id
            RETURNING id
        """, (body.line_id, body.model_number, body.fg_part_id))
        new_id = cur.fetchone()[0]
        conn.commit()
    return {"id": new_id, "ok": True}


@router.delete("/model-links/{link_id}")
def delete_model_link(link_id: int, admin=Depends(require_admin)):
    _ensure_tables()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM mes_fg_model_link WHERE id=%s", (link_id,))
        conn.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════
#  Kanban Log + Auto-fire watcher
# ════════════════════════════════════════════════════════════════════
@router.get("/log")
def list_kanban_log(date_from: Optional[str] = None,
                     date_to:   Optional[str] = None,
                     line_id:   Optional[int] = None,
                     user=Depends(get_current_user)):
    _ensure_tables()
    today = date.today().isoformat()
    d_from = date_from or today
    d_to   = date_to   or today
    sql = """SELECT kl.*, p.tbdi_part_no, p.description, p.model,
                    p.packing_std_qty, p.tbdi_new_part_no
               FROM mes_kanban_log kl
               JOIN mes_fg_parts p ON p.id = kl.fg_part_id
              WHERE kl.log_date BETWEEN %s AND %s"""
    params: list = [d_from, d_to]
    if line_id is not None:
        sql += " AND kl.line_id = %s"
        params.append(line_id)
    sql += " ORDER BY kl.log_date DESC, p.model, p.tbdi_part_no, kl.window_name"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


@router.get("/dashboard")
def dashboard(line_id: Optional[int] = None,
               on_date: Optional[str] = None,
               user=Depends(get_current_user)):
    """Compact today-snapshot for the Kanban dashboard:
       per part → 3 window dispatch_qty + total + monthly plan + achieved %"""
    _ensure_tables()
    d = on_date or date.today().isoformat()
    ym = d[:7]
    sql = """
        WITH parts AS (
            SELECT p.id AS fg_part_id, p.tbdi_part_no, p.tbdi_new_part_no,
                   p.customer_part_no, p.description, p.model,
                   p.packing_std_qty, p.line_id, l.line_name,
                   mp.shift_a_plan, mp.shift_b_plan, mp.total_plan
              FROM mes_fg_parts p
         LEFT JOIN mes_lines l ON l.id = p.line_id
         LEFT JOIN mes_monthly_plan mp
                ON mp.fg_part_id = p.id AND mp.year_month = %s
             WHERE p.is_active = TRUE
        ),
        today_log AS (
            SELECT fg_part_id, window_name,
                   SUM(kanban_count) AS kc, SUM(dispatch_qty) AS dq,
                   SUM(cycles_produced) AS cp
              FROM mes_kanban_log
             WHERE log_date = %s
          GROUP BY fg_part_id, window_name
        ),
        month_total AS (
            SELECT fg_part_id, SUM(dispatch_qty)::int AS dispatched
              FROM mes_kanban_log
             WHERE log_date >= date_trunc('month', %s::date)
               AND log_date <  date_trunc('month', %s::date) + INTERVAL '1 month'
          GROUP BY fg_part_id
        )
        SELECT p.*,
               COALESCE(t12.kc, 0)::int  AS kanban_12pm,
               COALESCE(t12.dq, 0)::int  AS dispatch_12pm,
               COALESCE(ta.kc,  0)::int  AS kanban_a,
               COALESCE(ta.dq,  0)::int  AS dispatch_a,
               COALESCE(tb.kc,  0)::int  AS kanban_b,
               COALESCE(tb.dq,  0)::int  AS dispatch_b,
               (COALESCE(t12.dq,0) + COALESCE(ta.dq,0) + COALESCE(tb.dq,0))::int AS total_today,
               COALESCE(mt.dispatched, 0)::int AS dispatched_mtd
          FROM parts p
     LEFT JOIN today_log t12 ON t12.fg_part_id = p.fg_part_id AND t12.window_name = '12PM'
     LEFT JOIN today_log ta  ON ta.fg_part_id  = p.fg_part_id AND ta.window_name  = 'SHIFT_A'
     LEFT JOIN today_log tb  ON tb.fg_part_id  = p.fg_part_id AND tb.window_name  = 'SHIFT_B'
     LEFT JOIN month_total mt ON mt.fg_part_id = p.fg_part_id
    """
    params: list = [ym, d, d, d]
    if line_id is not None:
        sql += " WHERE p.line_id = %s"
        params.append(line_id)
    sql += " ORDER BY p.model, p.tbdi_part_no"
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, tuple(params))
        return cur.fetchall()


class ManualLogBody(BaseModel):
    fg_part_id:   int
    log_date:     str           # YYYY-MM-DD
    window_name:  str           # 12PM | SHIFT_A | SHIFT_B
    kanban_count: int = 0
    dispatch_qty: int = 0
    notes:        Optional[str] = None


@router.post("/log", status_code=201)
def upsert_manual_log(body: ManualLogBody, user=Depends(get_current_user)):
    _ensure_tables()
    if body.window_name not in ("12PM", "SHIFT_A", "SHIFT_B"):
        raise HTTPException(400, "window_name must be 12PM|SHIFT_A|SHIFT_B")
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT line_id FROM mes_fg_parts WHERE id=%s", (body.fg_part_id,))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, "fg_part not found")
        lid = r["line_id"]
        cur2 = conn.cursor()
        cur2.execute("""
            INSERT INTO mes_kanban_log
                (fg_part_id, line_id, log_date, window_name,
                 cycles_produced, kanban_count, dispatch_qty, fired_by, notes)
            VALUES (%s,%s,%s,%s, 0,%s,%s,'manual',%s)
            ON CONFLICT (fg_part_id, log_date, window_name) DO UPDATE
                SET kanban_count = EXCLUDED.kanban_count,
                    dispatch_qty = EXCLUDED.dispatch_qty,
                    fired_by     = 'manual',
                    notes        = EXCLUDED.notes
            RETURNING id
        """, (body.fg_part_id, lid, body.log_date, body.window_name,
              body.kanban_count, body.dispatch_qty, body.notes))
        new_id = cur2.fetchone()[0]
        conn.commit()
    return {"id": new_id, "ok": True}


# ════════════════════════════════════════════════════════════════════
#  Auto-fire watcher  (background thread)
# ════════════════════════════════════════════════════════════════════
_STOP   = threading.Event()
_THREAD: Optional[threading.Thread] = None


def _window_boundaries(line_id: int) -> List[tuple]:
    """Return [(window_name, end_time)] for the line, sorted by end_time.
    12PM is always 12:00:00.  SHIFT_A end and SHIFT_B end come from
    mes_shift_configs."""
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT shift_name, end_time
                         FROM mes_shift_configs
                        WHERE line_id = %s AND COALESCE(is_production, true)
                          AND shift_name IN ('A','B')""", (line_id,))
        end_times = {r["shift_name"]: r["end_time"] for r in cur.fetchall()}
    windows = [("12PM", dt_time(12, 0, 0))]
    if "A" in end_times: windows.append(("SHIFT_A", end_times["A"]))
    if "B" in end_times: windows.append(("SHIFT_B", end_times["B"]))
    return sorted(windows, key=lambda x: x[1])


def _window_start_dt(window_name: str, end_dt: datetime,
                      line_id: int) -> datetime:
    """Look up previous-window end as this window's start.  Fallback:
    8 hours back (covers any single shift)."""
    boundaries = _window_boundaries(line_id)
    # Sort by end_time ascending
    end_dts = []
    base    = end_dt.date()
    for nm, et in boundaries:
        # Crosses-midnight handling: if et < 06:00, it's "next-day"
        d = base if et.hour >= 6 else base + timedelta(days=1)
        end_dts.append((nm, datetime.combine(d, et)))
    end_dts.sort(key=lambda x: x[1])
    # Find the boundary BEFORE end_dt
    prev = None
    for nm, dt_ in end_dts:
        if dt_ < end_dt:
            prev = dt_
    return prev or (end_dt - timedelta(hours=8))


def _fire_window(line_id: int, window_name: str, end_dt: datetime) -> int:
    """For the window ending at end_dt, count cycles in the window range
    and insert kanban_log rows for every linked (model → FG) pair.
    Returns number of log rows written."""
    with get_conn() as conn:
        cur = dict_cursor(conn)
        # Get the line's db_table_name + shift A/B for window range
        cur.execute("SELECT db_table_name FROM mes_lines WHERE id=%s", (line_id,))
        r = cur.fetchone()
        if not r:
            return 0
        tbl_log = r["db_table_name"] + "_ct_log"
        start_dt = _window_start_dt(window_name, end_dt, line_id)

        # Active model on the line — read from the line's dashboard
        # shift row (current_model_number is updated by the collector
        # every 2 s).  Fall back to NULL if no row for today yet.
        try:
            cur.execute(f"""SELECT current_model_number
                              FROM {r["db_table_name"]}
                             WHERE record_date = CURRENT_DATE
                          ORDER BY id DESC LIMIT 1""")
        except Exception:
            cur_model = None
        else:
            mr = cur.fetchone()
            cur_model = mr.get("current_model_number") if mr else None

        # Count cycles in the window (group by part_code prefix → use it
        # as a proxy for "model" when ct_log lacks model_number).  For
        # simplicity, attribute all cycles to the currently-linked FG
        # via the model_link table.
        cur.execute(f"""
            SELECT COUNT(*) AS cycles
              FROM {tbl_log}
             WHERE ts >= %s AND ts < %s
        """, (start_dt, end_dt))
        cycles = (cur.fetchone() or {}).get("cycles", 0)

        # All FG parts linked to this line (via model_link OR direct
        # line_id) — attribute cycles to each pro-rata if multiple
        # were configured.  Simplest: pick the FG linked to the
        # current model.  If none linked, fall back to first FG for line.
        target_fg = None
        if cur_model is not None:
            cur.execute("""SELECT fg_part_id FROM mes_fg_model_link
                            WHERE line_id=%s AND model_number=%s""",
                        (line_id, cur_model))
            r = cur.fetchone()
            if r: target_fg = r["fg_part_id"]
        if target_fg is None:
            cur.execute("""SELECT id FROM mes_fg_parts
                            WHERE line_id=%s AND is_active=TRUE
                            ORDER BY id LIMIT 1""", (line_id,))
            r = cur.fetchone()
            if r: target_fg = r["id"]
        if target_fg is None:
            return 0

        cur.execute("SELECT packing_std_qty FROM mes_fg_parts WHERE id=%s",
                    (target_fg,))
        pack_qty = max(1, int((cur.fetchone() or {}).get("packing_std_qty") or 1))

        kanban_count = cycles // pack_qty
        dispatch_qty = kanban_count * pack_qty
        log_date     = end_dt.date()

        cur2 = conn.cursor()
        cur2.execute("""
            INSERT INTO mes_kanban_log
                (fg_part_id, line_id, log_date, window_name,
                 cycles_produced, kanban_count, dispatch_qty,
                 fired_at, fired_by, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,NOW(),'auto',%s)
            ON CONFLICT (fg_part_id, log_date, window_name) DO UPDATE
                SET cycles_produced = EXCLUDED.cycles_produced,
                    kanban_count    = EXCLUDED.kanban_count,
                    dispatch_qty    = EXCLUDED.dispatch_qty,
                    fired_at        = NOW(),
                    fired_by        = 'auto',
                    notes           = EXCLUDED.notes
        """, (target_fg, line_id, log_date, window_name,
              cycles, kanban_count, dispatch_qty,
              f"auto-fired {start_dt:%H:%M}→{end_dt:%H:%M}"))
        conn.commit()
    return 1


def _scheduled_end_dts_today(line_id: int) -> List[tuple]:
    """Return [(window_name, end_datetime_today)] for the three windows.
    SHIFT_B endtime that's "before noon" wall-clock is treated as
    "tomorrow morning" — but only AFTER its corresponding shift-start
    has been crossed today (handled by the caller via < now filter)."""
    today = date.today()
    out: List[tuple] = []
    for nm, et in _window_boundaries(line_id):
        # 12PM and SHIFT_A end-times are always today.
        # SHIFT_B endtime (e.g. 03:15) is the NEXT day's early morning.
        if et.hour < 6 and nm == "SHIFT_B":
            d = today  # actually fires "tomorrow morning"; back-fill
                      # logic still uses today as log_date so the dashboard
                      # groups it under the production day that started B.
            out.append((nm, datetime.combine(today, et) + timedelta(days=1)))
        else:
            out.append((nm, datetime.combine(today, et)))
    return out


def _already_logged(line_id: int, fg_part_id: int, log_date: date,
                     window_name: str) -> bool:
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT 1 FROM mes_kanban_log
                        WHERE line_id=%s AND fg_part_id=%s
                          AND log_date=%s AND window_name=%s""",
                    (line_id, fg_part_id, log_date, window_name))
        return cur.fetchone() is not None


def _backfill_pending_windows(now_dt: datetime) -> int:
    """For every line, if a window's end_dt has passed today and no
    kanban_log row exists yet for today's (line, window), fire it now.
    Returns count of windows fired."""
    fired = 0
    try:
        with get_conn() as conn:
            cur = dict_cursor(conn)
            cur.execute("""SELECT id FROM mes_lines
                            WHERE db_table_name IS NOT NULL
                              AND db_table_name <> ''""")
            line_ids = [r["id"] for r in cur.fetchall()]
        for lid in line_ids:
            for nm, end_dt in _scheduled_end_dts_today(lid):
                if end_dt > now_dt:
                    continue   # boundary not yet reached
                # Has any row for (lid, today, nm) been written?
                with get_conn() as conn:
                    cur = conn.cursor()
                    cur.execute("""SELECT 1 FROM mes_kanban_log
                                    WHERE line_id=%s AND log_date=%s
                                      AND window_name=%s LIMIT 1""",
                                (lid, end_dt.date(), nm))
                    if cur.fetchone():
                        continue
                try:
                    n = _fire_window(lid, nm, end_dt)
                    if n:
                        fired += 1
                        print(f"[KANBAN] back-fired line={lid} {nm} @ {end_dt}")
                except Exception as exc:
                    print(f"[KANBAN] back-fire failed line={lid} {nm}: {exc}")
    except Exception as exc:
        print(f"[KANBAN] backfill scan error: {exc}")
    return fired


def _watcher_loop():
    """Every 60 s: scan all lines × all windows for today.  If a window's
    end-time has passed AND no log row written yet → fire.  This back-fill
    style handles the case where the watcher started AFTER a window had
    already passed (e.g. backend restart mid-shift)."""
    print("[KANBAN] Auto-fire watcher started — checks every 60 s")
    # Initial sweep — back-fill any past windows for today right away
    try:
        fired = _backfill_pending_windows(datetime.now())
        if fired:
            print(f"[KANBAN] startup back-fill: {fired} windows fired")
    except Exception as exc:
        print(f"[KANBAN] startup back-fill error: {exc}")
    while not _STOP.wait(60):
        try:
            _backfill_pending_windows(datetime.now())
        except Exception as exc:
            print(f"[KANBAN] watcher iteration error: {exc}")
    print("[KANBAN] watcher stopped")


# ── Manual "fire now" endpoint (admin) — useful for testing + recovery
@router.post("/fire-now")
def fire_now_endpoint(line_id: Optional[int] = None,
                       window_name: Optional[str] = None,
                       admin=Depends(require_admin)):
    """Manually trigger window auto-fire for today.  If no line_id
    given, fires for all lines.  If no window_name, fires every window
    whose end_dt has passed today.  Idempotent — won't double-write."""
    now_dt = datetime.now()
    fired_rows: List[dict] = []
    with get_conn() as conn:
        cur = dict_cursor(conn)
        if line_id is not None:
            cur.execute("SELECT id FROM mes_lines WHERE id=%s", (line_id,))
        else:
            cur.execute("""SELECT id FROM mes_lines
                            WHERE db_table_name IS NOT NULL
                              AND db_table_name <> ''""")
        line_ids = [r["id"] for r in cur.fetchall()]
    for lid in line_ids:
        for nm, end_dt in _scheduled_end_dts_today(lid):
            if window_name and nm != window_name.upper():
                continue
            if end_dt > now_dt:
                continue
            try:
                # Always fire (will UPDATE on conflict)
                n = _fire_window(lid, nm, end_dt)
                if n:
                    fired_rows.append({"line_id": lid, "window": nm, "end_dt": str(end_dt)})
            except Exception as exc:
                print(f"[KANBAN] manual fire failed line={lid} {nm}: {exc}")
    return {"fired": len(fired_rows), "rows": fired_rows}


def start_watcher():
    global _THREAD
    _ensure_tables()
    if _THREAD and _THREAD.is_alive():
        return
    _STOP.clear()
    _THREAD = threading.Thread(target=_watcher_loop,
                                name="kanban-auto-fire",
                                daemon=True)
    _THREAD.start()
