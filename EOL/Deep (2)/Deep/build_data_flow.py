"""
build_data_flow.py
==================
Single-page editable Excel workflow that mirrors the approved ASCII
layout exactly:

  • EOL Machine — centred, narrow box at the top
  • 3 attached components (PLC, Camera, ANDON) — spread with gaps
  • Collector — centred, narrow box
  • 5 stream boxes on the left + tall SIDE PANEL on the right
  • 5 department boxes at the bottom (only colour layer)

All boxes are merged-cell text — fully editable.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────
# Palette
# ─────────────────────────────────────────────────────────────────────
BW = {"title_bg":"1A202C","title_fg":"FFFFFF","body_bg":"FFFFFF",
      "body_fg":"1A202C","edge":"1A202C","subtitle":"4A5568"}
SIDE = {"title_bg":"374151","title_fg":"FFFFFF","body_bg":"F8FAFC",
        "body_fg":"1A202C","edge":"374151"}
DEPT_PRODUCTION  = {"title_bg":"15803D","title_fg":"FFFFFF","body_bg":"E5F5E5","body_fg":"14532D","edge":"15803D"}
DEPT_MAINTENANCE = {"title_bg":"B91C1C","title_fg":"FFFFFF","body_bg":"FDE5E5","body_fg":"7F1D1D","edge":"B91C1C"}
DEPT_QUALITY     = {"title_bg":"A16207","title_fg":"FFFFFF","body_bg":"FCEFCC","body_fg":"713F12","edge":"A16207"}
DEPT_ADMIN       = {"title_bg":"1E40AF","title_fg":"FFFFFF","body_bg":"DDE9F8","body_fg":"1E3A8A","edge":"1E40AF"}
DEPT_OPERATOR    = {"title_bg":"475569","title_fg":"FFFFFF","body_bg":"E2E8F0","body_fg":"1F2937","edge":"475569"}

# ─────────────────────────────────────────────────────────────────────
# Workbook setup — 18 columns
# ─────────────────────────────────────────────────────────────────────
NCOLS = 18
COL_WIDTH = 13

wb = Workbook()
ws = wb.active
ws.title = "MES Data Flow"
ws.sheet_view.showGridLines = False

for c in range(1, NCOLS + 1):
    ws.column_dimensions[get_column_letter(c)].width = COL_WIDTH

ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A3
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)
ws.print_options.horizontalCentered = True
ws.oddFooter.center.text = "MES Data Flow Diagram   ·   Internal — Bawal Plant   ·   Page &P"

# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def merge_box(row_top, col_left, row_bottom, col_right):
    return f"{get_column_letter(col_left)}{row_top}:{get_column_letter(col_right)}{row_bottom}"

def apply_box_borders(row_top, col_left, row_bottom, col_right, edge_color):
    side = Side(style="medium", color=edge_color)
    none = Side(style=None)
    for r in range(row_top, row_bottom + 1):
        for c in range(col_left, col_right + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                top    = side if r == row_top    else none,
                bottom = side if r == row_bottom else none,
                left   = side if c == col_left   else none,
                right  = side if c == col_right  else none,
            )

def draw_box(row_top, col_left, height, width, *, title, body, palette,
             title_size=11, body_size=10):
    row_bottom = row_top + height - 1
    col_right  = col_left + width - 1
    apply_box_borders(row_top, col_left, row_bottom, col_right, palette["edge"])

    ws.merge_cells(merge_box(row_top, col_left, row_top, col_right))
    t = ws.cell(row=row_top, column=col_left, value=title)
    t.font = Font(name="Arial", size=title_size, bold=True, color=palette["title_fg"])
    t.fill = fill(palette["title_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if height > 1:
        ws.merge_cells(merge_box(row_top + 1, col_left, row_bottom, col_right))
        b = ws.cell(row=row_top + 1, column=col_left, value=body or "")
        b.font = Font(name="Arial", size=body_size, color=palette["body_fg"])
        b.fill = fill(palette["body_bg"])
        b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return row_bottom

def draw_arrows_at(row, columns, height=22):
    for c in columns:
        cell = ws.cell(row=row, column=c, value="↓")
        cell.font = Font(name="Arial", size=14, bold=True, color="1A202C")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height

# ─────────────────────────────────────────────────────────────────────
# TITLE BAND
# ─────────────────────────────────────────────────────────────────────
row = 1
ws.merge_cells(merge_box(row, 1, row + 1, NCOLS))
t = ws.cell(row=row, column=1,
            value="MES DATA FLOW   —   Toyota Boshoku Device India")
t.font = Font(name="Arial", size=22, bold=True, color="FFFFFF")
t.fill = fill(BW["title_bg"])
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 28
ws.row_dimensions[row + 1].height = 28
row += 2

ws.merge_cells(merge_box(row, 1, row, NCOLS))
s = ws.cell(row=row, column=1,
            value="From the EOL machine and its attached components, through the collector, to every department's screen")
s.font = Font(name="Arial", size=11, italic=True, color=BW["subtitle"])
s.fill = fill("F7FAFC")
s.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 22
row += 1

ws.row_dimensions[row].height = 8
row += 1

# ─────────────────────────────────────────────────────────────────────
# LAYER 1 — EOL MACHINE   (centred,  cols 7–12)
# ─────────────────────────────────────────────────────────────────────
EOL_W = 6
EOL_LEFT = 7      # cols 7..12 = 6 cols, centred in 18-col grid
EOL_H = 4
eol_top = row
draw_box(eol_top, EOL_LEFT, EOL_H, EOL_W,
         title="EOL MACHINE",
         body=("(one production cell)\n"
               "Main Assembly  +  3 sub-stations"),
         palette=BW, title_size=14, body_size=10.5)
ws.row_dimensions[eol_top].height = 26
for r in range(eol_top + 1, eol_top + EOL_H):
    ws.row_dimensions[r].height = 18
row = eol_top + EOL_H

# Caption
ws.merge_cells(merge_box(row, 1, row, NCOLS))
cap = ws.cell(row=row, column=1, value="3 components are attached to this cell")
cap.font = Font(name="Arial", size=10, italic=True, color=BW["subtitle"])
cap.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 16
row += 1

# Arrows down to each component (3 arrows at columns 3.5, 9.5, 15.5)
draw_arrows_at(row, [3, 10, 16])
row += 1

# ─────────────────────────────────────────────────────────────────────
# LAYER 2 — PLC / CAMERA / ANDON   (3 boxes with gaps)
# ─────────────────────────────────────────────────────────────────────
attach_specs = [
    ( 2, 5, "PLC",
     "Mitsubishi controller\n\n"
     "carries:\n"
     "•  status — running / idle /\n"
     "    setup / breakdown\n"
     "•  OK / NG part counts\n"
     "•  cycle time\n"
     "•  current model\n"
     "•  Poka-Yoke checks\n"
     "•  per sub-station counts"),
    ( 8, 4, "CAMERA",
     "overhead NF2 camera\n\n"
     "carries:\n"
     "•  live video stream\n"
     "•  per-cycle recording\n"
     "    saved for later review"),
    (14, 5, "ANDON BOX",
     "tower light + buttons\n\n"
     "carries:\n"
     "•  visual alert state\n"
     "•  operator events:\n"
     "    call for help, raise issue,\n"
     "    manual reset\n"
     "•  audible alarm"),
]
ATTACH_H = 11
attach_top = row
attach_centers = []
for col_left, width, title, body in attach_specs:
    draw_box(attach_top, col_left, ATTACH_H, width,
             title=title, body=body, palette=BW,
             title_size=12, body_size=9.5)
    attach_centers.append(col_left + width // 2)
ws.row_dimensions[attach_top].height = 24
for r in range(attach_top + 1, attach_top + ATTACH_H):
    ws.row_dimensions[r].height = 16
row = attach_top + ATTACH_H

# Arrows down — 3 arrows under each attached box
draw_arrows_at(row, attach_centers)
row += 1

# ─────────────────────────────────────────────────────────────────────
# LAYER 3 — COLLECTOR   (centred,  cols 6–13)
# ─────────────────────────────────────────────────────────────────────
COLL_W = 8
COLL_LEFT = 6
COLL_H = 5
coll_top = row
draw_box(coll_top, COLL_LEFT, COLL_H, COLL_W,
         title="COLLECTOR",
         body=("reads every machine continuously,\n"
               "listens to events,\n"
               "logs everything to the database"),
         palette=BW, title_size=13, body_size=10)
ws.row_dimensions[coll_top].height = 26
for r in range(coll_top + 1, coll_top + COLL_H):
    ws.row_dimensions[r].height = 18
row = coll_top + COLL_H

# Caption + arrows down to streams
ws.merge_cells(merge_box(row, 1, row, NCOLS))
cap = ws.cell(row=row, column=1, value="5 processed data streams")
cap.font = Font(name="Arial", size=10, italic=True, color=BW["subtitle"])
cap.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 16
row += 1

# Arrows above each stream (5 arrows centred in their 2-col stream box)
stream_arrow_cols = [1, 3, 5, 7, 9]   # left-col of each stream pair
draw_arrows_at(row, [c + 1 for c in [1, 3, 5, 7, 9]])
row += 1

# ─────────────────────────────────────────────────────────────────────
# LAYER 4 — STREAMS  (cols 1–10)  +  SIDE PANEL  (cols 14–18)
# ─────────────────────────────────────────────────────────────────────
stream_specs = [
    ("LIVE OEE",
     "plan vs\nactual\n\nhourly\nrollup\n\nlosses\nbreak-up"),
    ("BREAKDOWN",
     "slip + cause\n+ duration\n\nMTBF\nMTTR\nLTTR"),
    ("POKA-YOKE\nALERTS",
     "bypass +\nstuck\n\nemail\nescalation"),
    ("PROCESS\nTRACKING",
     "per-minute\nsamples\nvs target\n\n→ see\nSIDE PANEL"),
    ("DEVIATION\nFLOW",
     "Maint →\nQuality\napproval"),
]
STREAM_H = 9
stream_top = row
col_left = 1
stream_centers = []
for title, body in stream_specs:
    draw_box(stream_top, col_left, STREAM_H, 2,
             title=title, body=body, palette=BW,
             title_size=10, body_size=9)
    stream_centers.append(col_left)   # left col of pair (stream box centre)
    col_left += 2

# Side panel — same height as streams (avoids row overlap with dept layer below)
SIDE_H = STREAM_H
draw_box(stream_top, 14, SIDE_H, 5,
         title="SIDE PANEL  —  per-sub-station capture",
         body=("Process Tracking captures the per-minute output\n"
               "of EVERY sub-station, rendered as a bar chart\n"
               "with a horizontal target line.\n\n"
               "Stations captured:\n"
               "•  Main Assembly\n"
               "•  Upper Rail Greasing\n"
               "•  Lock Bar Insert\n"
               "•  Lower Rail Greasing\n\n"
               "Drill:  Zone  →  Line  →  Machine\n"
               "Adaptive bucket: 1 / 5 / 10 / 30 / 60 min."),
         palette=SIDE, title_size=11, body_size=9.5)

ws.row_dimensions[stream_top].height = 30
for r in range(stream_top + 1, stream_top + STREAM_H):
    ws.row_dimensions[r].height = 17

# Stream layer ends at stream_top + STREAM_H - 1
row = stream_top + STREAM_H

# Arrows down to departments — 5 arrows under stream centres
# Stream boxes are at cols 1-2, 3-4, 5-6, 7-8, 9-10 → centres at 1,3,5,7,9 (left col)
# But arrow should appear in centre of merged cell pair (use left col)
draw_arrows_at(row, [c for c in [2, 4, 6, 8, 10]])  # left+1 of each pair
row += 1

# Side panel ends 4 rows past streams — but department row starts now.
# That's fine: side panel cols (14-18) are not used by stream's arrow row.
# Just continue.

# ─────────────────────────────────────────────────────────────────────
# LAYER 5 — DEPARTMENTS   (5 boxes spread across 18 cols)
# ─────────────────────────────────────────────────────────────────────
# Widths: 4 + 4 + 3 + 4 + 3 = 18
dept_specs = [
    (4, "PRODUCTION TEAM",
     "Monitors live production.\n"
     "Files the slip when the line stops.\n"
     "Reviews trends + AI assistant.\n\n"
     "Line leads and shop-floor staff",
     DEPT_PRODUCTION),
    (4, "MAINTENANCE TEAM",
     "Attends every breakdown.\n"
     "Resolves Poka-Yoke faults.\n"
     "Files CAPA on recurring issues.\n"
     "Raises deviations to Quality.\n\n"
     "Technicians and supervisors",
     DEPT_MAINTENANCE),
    (3, "QUALITY TEAM",
     "Approves deviation requests.\n"
     "Monitors bypass alerts.\n"
     "Files 4M Change Notes.\n\n"
     "Quality Sec Head &\nQuality Head",
     DEPT_QUALITY),
    (4, "ADMIN  /  PLANT HEAD",
     "Configures the entire platform.\n"
     "Grants per-page access.\n"
     "Has visibility into every screen.\n\n"
     "Platform owner",
     DEPT_ADMIN),
    (3, "OPERATOR",
     "Sees only assigned line.\n"
     "Read-only dashboard.\n"
     "Cannot reach admin pages.\n\n"
     "Single-line view",
     DEPT_OPERATOR),
]
DEPT_H = 8
dept_top = row
col_cursor = 1
for w, title, body, palette in dept_specs:
    draw_box(dept_top, col_cursor, DEPT_H, w,
             title=title, body=body, palette=palette,
             title_size=11, body_size=9.5)
    col_cursor += w
ws.row_dimensions[dept_top].height = 28
for r in range(dept_top + 1, dept_top + DEPT_H):
    ws.row_dimensions[r].height = 17
row = dept_top + DEPT_H

# ─────────────────────────────────────────────────────────────────────
# CROSS-FLOW + FOOTER
# ─────────────────────────────────────────────────────────────────────
row += 1
ws.merge_cells(merge_box(row, 1, row, NCOLS))
n = ws.cell(row=row, column=1,
            value=("Cross-flow:   Live OEE → Production · Admin     |     "
                   "Breakdown → Production (banner) + Maintenance (closure / CAPA) · Admin     |     "
                   "Poka-Yoke → Maintenance (fix) + Quality (watch + email) · Admin     |     "
                   "Process Tracking → Production (graphs) · Admin     |     "
                   "Deviation → Maintenance (raise) → Quality (approve) · Admin     |     "
                   "Operator sees only own assigned lines."))
n.font = Font(name="Arial", size=9, color=BW["subtitle"])
n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
n.fill = fill("F7FAFC")
ws.row_dimensions[row].height = 36
row += 1

ws.merge_cells(merge_box(row, 1, row, NCOLS))
fc = ws.cell(row=row, column=1,
             value="Internal document — Toyota Boshoku Device India Pvt. Ltd., Bawal")
fc.font = Font(name="Arial", size=8, italic=True, color="A0AEC0")
fc.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 16

# ─────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────
import time
XLSX_PATH = r"D:\EOL\EOL\Deep (2)\Deep\MES_Data_Flow.xlsx"
try:
    wb.save(XLSX_PATH)
except PermissionError:
    XLSX_PATH = XLSX_PATH.replace(".xlsx", f"_{int(time.time())}.xlsx")
    wb.save(XLSX_PATH)
    print(f"WARNING: original was locked — saved as: {XLSX_PATH}")
print(f"Excel written: {XLSX_PATH}")
print(f"Total rows used: {row}")
