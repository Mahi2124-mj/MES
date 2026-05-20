"""
import_and_merge.py
===================
Reads poka_yoke_full.xlsx (PLC actuals) + poka_yoke_export.xlsx (desired values),
merges them using the sensor→D.bit mapping, and outputs:
  - pokayoke_data.json   (for web dashboard)
  - pokayoke_compiled.xlsx (color-coded Excel)

Usage:
    python import_and_merge.py
    python import_and_merge.py --plc path/to/plc.xlsx --export path/to/export.xlsx
"""

import json, os, sys, argparse
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# HARD-CODED SENSOR → D.BIT MAPPING  (from PLC documentation)
# ══════════════════════════════════════════════════════════════
SENSOR_TO_DBIT = {
    "relocate pin":                          "D.081",
    "LH HARNES 1":                           "D.049",
    "RH HARNESS 1":                          "D.050",
    "LH HARNESS 2":                          "D.041",
    "RH HARNESS 2":                          "D.041",
    "Rr. Lwr Proctecter":                    "D.090",
    "Fr. Lwr Proctecter":                    "D.083",
    "rr upper proctector /shifting pos. ng": "D.097",
    "lh bending":                            "D.104",
    "Rh bending":                            "D.105",
    "pop revit 1":                           "D.045",
    "pop rivet 2":                           "D.046",
    "fr lighter protector1":                 "D.054",
    "fr lighter proctector2":                "D.075",
    "fr lighter proctector3":                "D.053",
    "Rr lighter proctector 1":               "D.074",
    "Rr lighter proctector 2":               "D.076",
    "E ring":                                "D.060",
    "pop jig":                               None,      # no D.bit mapping
    "Fr lighter proctector 4":               "D.055",
    "lh harness 3":                          "D.049",
    "bolt mixing":                           "D.042",
    "Rh harness bkt":                        "D.050",
    "ytb lh exp Harness bkt":                "D.049",
    "upper rail hole detection":             "D.096",
    "lower rail whole detection":            "D.089",
}

# PLC model column mapping:
#   Column index in Sheet1 (0-based col 3,4,5 → headers "9","10","11")
#   model_9  = TRACK ASSY FRONT SEAT YHB 4 WAY OTR
#   model_10 = TRACK ASSY FRONT SEAT YHB 4 WAY INR RH
#   model_11 = TRACK ASSY FRONT SEAT YHB 4 WAY INR LH
MODEL_COLUMNS = {
    "9":  {"model": "YHB", "model_type": "4 WAY OUTER", "type_side": "BOTH",
            "full_name": "TRACK ASSY FRONT SEAT YHB 4 WAY OTR"},
    "10": {"model": "YHB", "model_type": "4 WAY INNER", "type_side": "RH",
            "full_name": "TRACK ASSY FRONT SEAT YHB 4 WAY INR RH"},
    "11": {"model": "YHB", "model_type": "4 WAY INNER", "type_side": "LH",
            "full_name": "TRACK ASSY FRONT SEAT YHB 4 WAY INR LH"},
}

# Value encoding: PLC text → numeric
ACTUAL_MAP = {"on": 2, "off": 1, "pass": 0}
VALUE_LABELS = {0: "PASS", 1: "OFF", 2: "ON"}


# ══════════════════════════════════════════════════════════════
# READ PLC ACTUALS  (poka_yoke_full.xlsx)
# ══════════════════════════════════════════════════════════════
#
# Data lives in TWO places inside this file:
#   Sheet1     → sensor_name (col 1) + device_no (col 2) — NO model values
#   Inner AssY → PY No (col 1), and model actual values in cols 15/16/17
#                col 15 = model_9 (YHB OTR), col 16 = model_10 (YHB INR RH),
#                col 17 = model_11 (YHB INR LH) — values: on/off/pass or 0/1/2
#
# We use sensor_map to link Sheet1 sensor names → D.bits,
# and the PY No from Inner AssY to derive D.bits (PY-6067 → D.067).
# ══════════════════════════════════════════════════════════════
def _py_no_to_dbit(py_no):
    """Extract D.bit from PY number: TBDI-PE-PY-6067 → D.067"""
    import re
    m = re.search(r'(\d{4})$', py_no.strip())
    if m:
        num = m.group(1)        # e.g. "6067"
        suffix = num[1:]        # "067"
        return f"D.{suffix}"
    return None


def _parse_actual_value(raw):
    """Parse a single on/off/pass/0/1/2 value. Returns int or None."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    # Handle compound like "on/off" → take first part
    if '/' in s:
        s = s.split('/')[0].strip()
    if s in ACTUAL_MAP:
        return ACTUAL_MAP[s]
    # Try numeric
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def read_plc_actuals(plc_path, sensor_map):
    """
    Returns dict keyed by D.bit:
      { "D.041": { "sensor_name": "...", "device_no": "404",
                    "actuals": {"9": 2, "10": 1, "11": 0} } }
    """
    result = {}
    if not plc_path or not os.path.exists(plc_path):
        return result

    wb = openpyxl.load_workbook(plc_path, data_only=True)

    # ── Step 1: Read Sheet1 for sensor→device mapping ──────────
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        for r in range(1, ws.max_row + 1):
            sensor_name = ws.cell(r, 1).value
            device_no   = ws.cell(r, 2).value
            if not sensor_name:
                continue
            sensor_clean = str(sensor_name).strip()

            d_bit = None
            for key, val in sensor_map.items():
                if key.strip().lower() == sensor_clean.lower():
                    d_bit = val
                    break
            if not d_bit:
                continue

            if d_bit not in result:
                result[d_bit] = {
                    "sensor_name": sensor_clean,
                    "device_no": str(device_no or "").strip(),
                    "actuals": {},
                }

    # ── Step 2: Read Inner AssY for actual on/off/pass values ──
    #   Col 15 = model_9, Col 16 = model_10, Col 17 = model_11
    if "Inner AssY" in wb.sheetnames:
        ws = wb["Inner AssY"]
        for r in range(2, ws.max_row + 1):
            py_no_raw = ws.cell(r, 1).value
            if not py_no_raw:
                continue
            py_no = str(py_no_raw).strip()
            if py_no.lower() in ("py no", "py no ", ""):
                continue

            # Check if any of cols 15-17 have data
            has_data = False
            for c in [15, 16, 17]:
                if ws.cell(r, c).value is not None:
                    has_data = True
                    break
            if not has_data:
                continue

            # Derive D.bit from PY number
            d_bit = _py_no_to_dbit(py_no)
            if not d_bit:
                continue

            py_name = str(ws.cell(r, 2).value or "").strip()

            if d_bit not in result:
                result[d_bit] = {
                    "sensor_name": py_name,
                    "device_no": "",
                    "actuals": {},
                }

            # Col 15→model "9", Col 16→model "10", Col 17→model "11"
            for col, model_key in [(15, "9"), (16, "10"), (17, "11")]:
                val = _parse_actual_value(ws.cell(r, col).value)
                if val is not None and model_key not in result[d_bit]["actuals"]:
                    result[d_bit]["actuals"][model_key] = val

    return result


# ══════════════════════════════════════════════════════════════
# READ EXPORT FILE  (poka_yoke_export.xlsx)
# ══════════════════════════════════════════════════════════════
def read_export(export_path):
    wb = openpyxl.load_workbook(export_path, data_only=True)

    # ── MODEL MASTER ──
    model_master = []
    ws = wb["MODEL MASTER"]
    for r in range(2, ws.max_row + 1):
        mn = ws.cell(r, 1).value
        if not mn:
            continue
        model_master.append({
            "model_name": str(mn).strip(),
            "type":       str(ws.cell(r, 2).value or "").strip(),
            "old_model_no": str(ws.cell(r, 3).value or "").strip(),
            "model":      str(ws.cell(r, 4).value or "").strip(),
        })

    # ── POKA YOKE MASTER ──
    py_master = []
    seen = set()
    ws = wb["POKA YOKE MASTER"]
    for r in range(2, ws.max_row + 1):
        pno = ws.cell(r, 1).value
        if not pno:
            continue
        pno = str(pno).strip()
        if pno in seen:
            continue
        seen.add(pno)
        py_master.append({
            "py_no":      pno,
            "py_name":    str(ws.cell(r, 2).value or "").strip(),
            "model_type": str(ws.cell(r, 3).value or "").strip(),
            "machine":    str(ws.cell(r, 4).value or "").strip(),
        })

    # ── FINAL SEAT (assignments) ──
    assignments = []
    ws = wb["final seat"]
    for r in range(2, ws.max_row + 1):
        pno = ws.cell(r, 2).value
        if not pno:
            continue
        dv_raw = ws.cell(r, 11).value
        try:
            dv = int(float(str(dv_raw)))
        except (ValueError, TypeError):
            dv = None

        assignments.append({
            "id":           str(ws.cell(r, 1).value or r - 1),
            "py_no":        str(pno).strip(),
            "py_name":      str(ws.cell(r, 3).value or "").strip(),
            "type_side":    str(ws.cell(r, 4).value or "").strip(),
            "model_type":   str(ws.cell(r, 5).value or "").strip(),
            "model_name":   str(ws.cell(r, 6).value or "").strip(),
            "type2":        str(ws.cell(r, 7).value or "").strip(),
            "old_model_no": str(ws.cell(r, 8).value or "").strip(),
            "model":        str(ws.cell(r, 9).value or "").strip(),
            "d_bit":        str(ws.cell(r, 10).value or "").strip(),
            "desired":      dv,
            "machine":      str(ws.cell(r, 12).value or "").strip(),
        })

    return model_master, py_master, assignments


# ══════════════════════════════════════════════════════════════
# MERGE:  match final-seat rows to PLC actuals
# ══════════════════════════════════════════════════════════════
def resolve_model_column(row, model_columns):
    """
    Given a final-seat row, determine which PLC model column (9/10/11) it maps to.
    Returns the column key ("9","10","11") or None.
    """
    code   = row["model"].upper()
    mtype  = row["model_type"].upper()
    side   = row["type_side"].upper()

    for col_key, col_info in model_columns.items():
        if col_info["model"].upper() != code:
            continue
        ci_type = col_info["model_type"].upper()
        ci_side = col_info["type_side"].upper()

        if ci_type not in mtype and mtype not in ci_type:
            continue

        # For INNER: must match side; for OUTER: side is BOTH
        if "INNER" in ci_type:
            if ci_side == side or side == "BOTH" or ci_side == "BOTH":
                return col_key
        else:
            # OUTER — side doesn't matter
            return col_key

    return None


def merge(plc_actuals, model_master, py_master, assignments, sensor_map, model_columns):
    compiled = []

    for row in assignments:
        d_bit = row["d_bit"]
        actual = None
        status = "NO_DATA"

        col_key = resolve_model_column(row, model_columns)
        if col_key and d_bit in plc_actuals:
            plc_entry = plc_actuals[d_bit]
            if col_key in plc_entry["actuals"]:
                actual = plc_entry["actuals"][col_key]

        desired = row["desired"]
        if actual is not None and desired is not None:
            status = "MATCH" if actual == desired else "MISMATCH"
        elif actual is not None:
            status = "ACTUAL_ONLY"
        else:
            status = "NO_DATA"

        compiled.append({
            "id":           row["id"],
            "py_no":        row["py_no"],
            "py_name":      row["py_name"],
            "type_side":    row["type_side"],
            "model_type":   row["model_type"],
            "model_name":   row["model_name"],
            "type2":        row["type2"],
            "old_model_no": row["old_model_no"],
            "model":        row["model"],
            "d_bit":        d_bit,
            "desired":      desired,
            "actual":       actual,
            "status":       status,
            "machine":      row["machine"],
            "plc_column":   col_key,
            "sensor_name":  plc_actuals.get(d_bit, {}).get("sensor_name"),
        })

    # Stats
    match = sum(1 for c in compiled if c["status"] == "MATCH")
    mismatch = sum(1 for c in compiled if c["status"] == "MISMATCH")
    nodata = sum(1 for c in compiled if c["status"] in ("NO_DATA", "ACTUAL_ONLY"))

    return {
        "models":          model_master,
        "poka_yokes":      py_master,
        "sensor_mapping":  {k: v for k, v in sensor_map.items() if v},
        "compiled":        compiled,
        "model_columns":   model_columns,
        "last_updated":    datetime.now().isoformat(),
        "plc_actuals_raw": {k: v for k, v in plc_actuals.items()},
        "stats": {
            "total":          len(compiled),
            "match":          match,
            "mismatch":       mismatch,
            "no_data":        nodata,
            "total_models":   len(model_master),
            "total_py":       len(py_master),
            "model_codes":    sorted(set(m["model"] for m in model_master if m["model"])),
        },
    }


# ══════════════════════════════════════════════════════════════
# GENERATE COMPILED EXCEL
# ══════════════════════════════════════════════════════════════
def generate_excel(data, output_path):
    wb = openpyxl.Workbook()

    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    grey_fill   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_header(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            mx = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 45)

    # ── Sheet 1: Dashboard View ──
    ws1 = wb.active
    ws1.title = "Dashboard View"
    h1 = ["Model Code", "Model Name", "Type", "Side", "PY No", "PY Name",
          "D Bit", "Machine", "Desired", "Actual", "Status", "Sensor Name"]
    write_header(ws1, h1)

    for i, c in enumerate(data["compiled"], 2):
        ws1.cell(i, 1,  c["model"])
        ws1.cell(i, 2,  c["model_name"])
        ws1.cell(i, 3,  c["model_type"])
        ws1.cell(i, 4,  c["type_side"])
        ws1.cell(i, 5,  c["py_no"])
        ws1.cell(i, 6,  c["py_name"])
        ws1.cell(i, 7,  c["d_bit"])
        ws1.cell(i, 8,  c["machine"])
        ws1.cell(i, 9,  VALUE_LABELS.get(c["desired"], "—") if c["desired"] is not None else "—")
        ws1.cell(i, 10, VALUE_LABELS.get(c["actual"], "—") if c["actual"] is not None else "—")
        ws1.cell(i, 11, c["status"])
        ws1.cell(i, 12, c.get("sensor_name") or "—")

        fill = grey_fill
        if c["status"] == "MATCH":    fill = green_fill
        elif c["status"] == "MISMATCH": fill = red_fill
        for col in range(1, len(h1) + 1):
            ws1.cell(i, col).fill = fill
            ws1.cell(i, col).border = thin_border

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}{len(data['compiled']) + 1}"
    auto_width(ws1)

    # ── Sheet 2: Model Master ──
    ws2 = wb.create_sheet("Model Master")
    h2 = ["Model Name", "Type", "Old Model No", "Model Code"]
    write_header(ws2, h2)
    for i, m in enumerate(data["models"], 2):
        ws2.cell(i, 1, m["model_name"])
        ws2.cell(i, 2, m["type"])
        ws2.cell(i, 3, m["old_model_no"])
        ws2.cell(i, 4, m["model"])
        for col in range(1, 5):
            ws2.cell(i, col).border = thin_border
    ws2.auto_filter.ref = f"A1:D{len(data['models']) + 1}"
    auto_width(ws2)

    # ── Sheet 3: All Poka Yokes ──
    ws3 = wb.create_sheet("All Poka Yokes")
    h3 = ["PY No", "PY Name", "Model Type", "Machine/Fixture"]
    write_header(ws3, h3)
    for i, p in enumerate(data["poka_yokes"], 2):
        ws3.cell(i, 1, p["py_no"])
        ws3.cell(i, 2, p["py_name"])
        ws3.cell(i, 3, p["model_type"])
        ws3.cell(i, 4, p["machine"])
        for col in range(1, 5):
            ws3.cell(i, col).border = thin_border
    ws3.auto_filter.ref = f"A1:D{len(data['poka_yokes']) + 1}"
    auto_width(ws3)

    wb.save(output_path)
    print(f"  [OK] Excel: {output_path}")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Poka Yoke Import & Merge")
    parser.add_argument("--plc",    default=None)
    parser.add_argument("--export", default=None)
    parser.add_argument("--out-json", default="pokayoke_data.json")
    parser.add_argument("--out-xlsx", default="pokayoke_compiled.xlsx")
    args = parser.parse_args()

    base = os.path.dirname(os.path.abspath(__file__))

    # Auto-detect files
    plc_path = args.plc
    if not plc_path:
        for p in [os.path.join(base, "poka_yoke_full.xlsx"),
                  os.path.join(base, "..", "Phase2", "poka_yoke_full.xlsx")]:
            if os.path.exists(p):
                plc_path = p
                break

    export_path = args.export
    if not export_path:
        for p in [os.path.join(base, "poka_yoke_export.xlsx"),
                  os.path.join(base, "..", "POKA-YOKE", "data", "poka_yoke_export.xlsx")]:
            if os.path.exists(p):
                export_path = p
                break

    if not export_path:
        print("[ERROR] Cannot find poka_yoke_export.xlsx. Use --export.")
        sys.exit(1)

    print(f"  PLC file:    {plc_path or '(not found — actuals will be NO_DATA)'}")
    print(f"  Export file: {export_path}")

    # Read
    plc_actuals = read_plc_actuals(plc_path, SENSOR_TO_DBIT)
    model_master, py_master, assignments = read_export(export_path)

    print(f"  Models: {len(model_master)} | PY Master: {len(py_master)} | Assignments: {len(assignments)}")
    if plc_actuals:
        print(f"  PLC sensors matched: {len(plc_actuals)} D.bits with actual values")

    # Merge
    merged = merge(plc_actuals, model_master, py_master, assignments,
                    SENSOR_TO_DBIT, MODEL_COLUMNS)

    s = merged["stats"]
    print(f"  Result: {s['match']} MATCH | {s['mismatch']} MISMATCH | {s['no_data']} NO_DATA | {s['total']} total")

    # Save JSON
    json_path = os.path.join(base, args.out_json)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(merged, f, indent=2, ensure_ascii=False, default=str)
    print(f"  [OK] JSON: {json_path}")

    # Save Excel
    xlsx_path = os.path.join(base, args.out_xlsx)
    generate_excel(merged, xlsx_path)

    print("[DONE]")


if __name__ == "__main__":
    main()
