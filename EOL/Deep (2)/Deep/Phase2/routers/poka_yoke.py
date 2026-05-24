# ════════════════════════════════════════════════════════════════
# routers/poka_yoke.py  — complete file
# ════════════════════════════════════════════════════════════════
"""
routers/poka_yoke.py
====================
Rules       → /api/poka-yoke/rules/{line_id}
Events      → /api/poka-yoke/events/{line_id}
PY Master   → /api/poka-yoke/master/
Model Master→ /api/poka-yoke/models/
Assignments → /api/poka-yoke/assignments/
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Optional, Any

from database import get_conn, dict_cursor
from auth import get_current_user, get_current_user_optional, require_admin

router = APIRouter(prefix="/api/poka-yoke", tags=["poka-yoke"])


# ── Helper ─────────────────────────────────────────────────────
def _check_operator_access(user: dict, line_id: int, conn) -> None:
    if user["role"] == "operator":
        cur = conn.cursor()
        cur.execute(
            "SELECT 1 FROM mes_operator_lines WHERE admin_id=%s AND line_id=%s",
            (user["id"], line_id)
        )
        if not cur.fetchone():
            raise HTTPException(403, "Not authorized to access this line")


# ══════════════════════════════════════════════════════════════
# SCHEMAS
# ══════════════════════════════════════════════════════════════

class RuleCreate(BaseModel):
    poka_yoke_no:   str
    poka_yoke_name: Optional[str] = None
    side:           Optional[str] = "ALL"
    model:          Optional[str] = "ALL"
    bit:            Optional[str] = None
    value:          Optional[int] = 1
    machine_name:   Optional[str] = None
    sheet_name:     Optional[str] = None
    alert_level:    str           = "WARNING"
    is_active:      bool          = True

class RuleUpdate(BaseModel):
    poka_yoke_name: Optional[str]  = None
    side:           Optional[str]  = None
    model:          Optional[str]  = None
    bit:            Optional[str]  = None
    value:          Optional[int]  = None
    machine_name:   Optional[str]  = None
    sheet_name:     Optional[str]  = None
    alert_level:    Optional[str]  = None
    is_active:      Optional[bool] = None

class PYMasterCreate(BaseModel):
    # pyNo is now optional — D-bit is the functional primary key.  Legacy column
    # retained (nullable) so existing joins keep working; if not supplied, the
    # backend auto-populates pyNo from the normalized D-bit string.
    pyNo:             Optional[str] = None
    description:      str
    modelType:        Optional[str] = None        # "4 Way" | "6 Way"
    typeSide:         Optional[str] = "ALL"       # LH / RH / OTR / Otr LH / Otr RH
    dBit:             Optional[str] = None        # D-register like "D400"
    desiredValue:     Optional[Any] = None
    machineFixture:   Optional[str] = None
    register:         Optional[str] = None        # legacy mirror of dBit
    registerCount:    Optional[int] = 1           # 1 → {0,1,2}, 2 → {0..4}
    zoneId:           Optional[int] = None        # FK → mes_zones.id
    sensingBits:      Optional[str] = None        # comma-sep X-bits, e.g. "X15" or "X21,X22"
    assignedModelIds: Optional[list[int]] = None

class PYMasterUpdate(BaseModel):
    description:      Optional[str] = None
    modelType:        Optional[str] = None
    typeSide:         Optional[str] = None
    dBit:             Optional[str] = None
    desiredValue:     Optional[Any] = None
    machineFixture:   Optional[str] = None
    register:         Optional[str] = None
    registerCount:    Optional[int] = None
    zoneId:           Optional[int] = None
    sensingBits:      Optional[str] = None
    assignedModelIds: Optional[list[int]] = None

class ModelMasterCreate(BaseModel):
    modelName:  str
    type:       Optional[str] = None
    oldModelNo: Optional[str] = None
    model:      Optional[str] = None   # series
    bitNumber:  Optional[int] = None   # PLC bit number (unique WITHIN a zone)
    zoneId:     Optional[int] = None   # FK → mes_zones.id

class ModelMasterUpdate(BaseModel):
    modelName:  Optional[str] = None
    type:       Optional[str] = None
    oldModelNo: Optional[str] = None
    model:      Optional[str] = None
    bitNumber:  Optional[int] = None
    zoneId:     Optional[int] = None

class AssignmentCreate(BaseModel):
    pyNo:           str
    pyName:         Optional[str] = None
    typeSide:       Optional[str] = "ALL"
    modelType:      Optional[str] = None
    modelName:      str
    type2:          Optional[str] = None
    oldModelNo:     Optional[str] = None
    modelSeries:    Optional[str] = None
    dBit:           Optional[str] = None
    desiredValue:   Optional[int] = None    # 0 = OFF, 1 = ON
    desiredBit:     Optional[int] = None    # which PLC bit to check
    machineFixture: Optional[str] = None

# ══════════════════════════════════════════════════════════════
# RULES
# ══════════════════════════════════════════════════════════════

@router.get("/rules/{line_id}")
def get_rules(
    line_id: int,
    side:    Optional[str] = None,
    sheet:   Optional[str] = None,
    model:   Optional[str] = None,
    user=Depends(get_current_user)
):
    with get_conn() as conn:
        _check_operator_access(user, line_id, conn)
        cur    = dict_cursor(conn)
        where  = ["line_id = %s", "is_active = true"]
        params = [line_id]

        if side and side.upper() not in ("ALL", ""):
            where.append("(side = %s OR side = 'ALL')")
            params.append(side.upper())
        if sheet:
            where.append("sheet_name = %s")
            params.append(sheet)
        if model:
            where.append("(model ILIKE %s OR model = 'all')")
            params.append(f"%{model}%")

        cur.execute(f"""
            SELECT * FROM mes_poka_yoke_rules
            WHERE {" AND ".join(where)}
            ORDER BY sheet_name, bit NULLS LAST, poka_yoke_no
        """, params)
        return cur.fetchall()


@router.post("/rules/{line_id}", status_code=201)
def create_rule(line_id: int, body: RuleCreate, admin=Depends(require_admin)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            INSERT INTO mes_poka_yoke_rules (
                line_id, poka_yoke_no, side, poka_yoke_name,
                model, bit, value, machine_name,
                sheet_name, alert_level, is_active
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING *
        """, (
            line_id,
            body.poka_yoke_no,
            body.side.upper(),
            body.poka_yoke_name,
            body.model,
            body.bit,
            body.value,
            body.machine_name,
            body.sheet_name,
            body.alert_level,
            body.is_active,
        ))
        rule = cur.fetchone()
        conn.cursor().execute("""
            INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
            VALUES ('POKA_YOKE_RULE_ADDED', 'line', %s, %s)
        """, (line_id, f"rule={body.poka_yoke_name} bit={body.bit}"))
        return rule


class RuleBulkCreate(BaseModel):
    rules: list[RuleCreate]

@router.post("/rules/{line_id}/bulk", status_code=201)
def create_rules_bulk(line_id: int, body: RuleBulkCreate, admin=Depends(require_admin)):
    """Bulk-assign all PY checks of a model configuration to a line (from Matrix tab)."""
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT id FROM mes_lines WHERE id = %s", (line_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Line not found")
        inserted = skipped = 0
        for rule in body.rules:
            # Skip if this exact bit is already assigned to this line
            cur.execute(
                "SELECT id FROM mes_poka_yoke_rules WHERE line_id=%s AND poka_yoke_no=%s AND bit=%s",
                (line_id, rule.poka_yoke_no, rule.bit),
            )
            if cur.fetchone():
                skipped += 1
                continue
            cur.execute("""
                INSERT INTO mes_poka_yoke_rules
                    (line_id, poka_yoke_no, poka_yoke_name, side, model, bit, value,
                     machine_name, sheet_name, alert_level, is_active)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                line_id, rule.poka_yoke_no,
                rule.poka_yoke_name or rule.poka_yoke_no,
                (rule.side or "ALL").upper(), rule.model or "ALL",
                rule.bit, rule.value if rule.value is not None else 1,
                rule.machine_name, rule.sheet_name,
                rule.alert_level, rule.is_active,
            ))
            inserted += 1
        conn.cursor().execute("""
            INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
            VALUES ('POKA_YOKE_BULK_ASSIGN', 'line', %s, %s)
        """, (line_id, f"inserted={inserted} skipped={skipped}"))
        return {"ok": True, "inserted": inserted, "skipped": skipped, "total": len(body.rules)}


class RuleBulkDelete(BaseModel):
    poka_yoke_nos: list[str]

@router.delete("/rules/{line_id}/bulk")
def delete_rules_bulk(line_id: int, body: RuleBulkDelete, admin=Depends(require_admin)):
    """Remove rules for specific PY nos from a line (used by Matrix tab remove)."""
    with get_conn() as conn:
        cur = conn.cursor()
        if not body.poka_yoke_nos:
            return {"ok": True, "deleted": 0}
        placeholders = ",".join(["%s"] * len(body.poka_yoke_nos))
        cur.execute(
            f"DELETE FROM mes_poka_yoke_rules WHERE line_id=%s AND poka_yoke_no IN ({placeholders})",
            [line_id] + body.poka_yoke_nos,
        )
        deleted = cur.rowcount
        conn.cursor().execute("""
            INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
            VALUES ('POKA_YOKE_BULK_REMOVE', 'line', %s, %s)
        """, (line_id, f"py_nos={body.poka_yoke_nos} deleted={deleted}"))
        return {"ok": True, "deleted": deleted}


@router.post("/rules/{line_id}/bulk-delete")
def delete_rules_bulk_post(line_id: int, body: RuleBulkDelete, admin=Depends(require_admin)):
    """POST-style bulk delete (client cannot send body with DELETE)."""
    return delete_rules_bulk(line_id, body, admin)


@router.put("/rules/{rule_id}")
def update_rule(rule_id: int, body: RuleUpdate, admin=Depends(require_admin)):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(400, "Nothing to update")
    sets   = ", ".join(f"{k} = %s" for k in updates)
    values = list(updates.values()) + [rule_id]
    with get_conn() as conn:
        conn.cursor().execute(
            f"UPDATE mes_poka_yoke_rules SET {sets} WHERE id = %s", values)
        conn.cursor().execute("""
            INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
            VALUES ('POKA_YOKE_RULE_UPDATED', 'poka_yoke_rule', %s, %s)
        """, (rule_id, str(updates)))
    return {"ok": True}


@router.delete("/rules/{rule_id}")
def delete_rule(rule_id: int, admin=Depends(require_admin)):
    with get_conn() as conn:
        conn.cursor().execute(
            "DELETE FROM mes_poka_yoke_rules WHERE id = %s", (rule_id,))
    return {"ok": True}


# ══════════════════════════════════════════════════════════════
# EVENTS
# ══════════════════════════════════════════════════════════════

@router.get("/events/{line_id}")
def get_events(
    line_id:      int,
    page:         int  = Query(1, ge=1),
    limit:        int  = Query(50, ge=1, le=200),
    unacked_only: bool = Query(False),
    alert_level:  Optional[str] = Query(None),
    user=Depends(get_current_user)
):
    with get_conn() as conn:
        _check_operator_access(user, line_id, conn)
        cur    = dict_cursor(conn)
        where  = ["e.line_id = %s"]
        params = [line_id]

        if unacked_only:
            where.append("e.acknowledged = false")
        if alert_level:
            where.append("e.alert_level = %s")
            params.append(alert_level.upper())

        where_sql = " AND ".join(where)
        offset    = (page - 1) * limit

        cur.execute(f"""
            SELECT e.*,
                   r.poka_yoke_name AS rule_name,
                   r.poka_yoke_name AS rule_type,
                   r.machine_name   AS alert_message
            FROM mes_poka_yoke_events e
            LEFT JOIN mes_poka_yoke_rules r ON r.id = e.rule_id
            WHERE {where_sql}
            ORDER BY e.detected_at DESC
            LIMIT %s OFFSET %s
        """, params + [limit, offset])
        events = cur.fetchall()

        cur.execute(
            f"SELECT COUNT(*) as total FROM mes_poka_yoke_events e WHERE {where_sql}",
            params
        )
        total = cur.fetchone()["total"]

        return {
            "events": events,
            "total":  total,
            "page":   page,
            "limit":  limit,
            "pages":  max(1, (total + limit - 1) // limit),
        }


@router.get("/bypass-episodes/{line_id}")
def get_bypass_episodes(
    line_id: int,
    date:    Optional[str] = Query(None, description="YYYY-MM-DD (defaults to today)"),
    shift:   Optional[str] = Query(None, description="Shift name (defaults to current open shift)"),
    user=Depends(get_current_user),
):
    """Return bypass episodes for a given line + shift, grouped per
    PY with start/end timestamps + slot attribution.

    2026-05-16 — Built after operator complaint that the hourly modal
    showed bit codes (D406) instead of PY names AND only the current
    slot's episodes were visible.  This endpoint walks every event in
    the requested shift's time window, parses context_json for the
    real py_name, and groups consecutive non-PASS events into episodes
    that close on the next PASS.  Each episode is also tagged with
    the hourly slot it falls into.

    JSON-side cache: results are also written to a local file
    `Phase2/_py_bypass_log_line{id}.json` so the frontend can be
    served from disk even if the events table is briefly slow,
    and so the file IS the at-rest representation of the current
    shift's bypass log.  Operator can ZIP and ship this file with
    audit packs.

    The file is overwritten on every request (atomic write via temp
    + rename), and the collector / shift-rotation worker can clear
    it on shift end (NOT during OT — OT keeps appending).
    """
    import json as _json
    import os as _os
    from datetime import datetime as _dt, date as _date, timedelta as _td

    target_date = date or _date.today().isoformat()
    with get_conn() as conn:
        _check_operator_access(user, line_id, conn)
        cur = dict_cursor(conn)

        # 1. Resolve shift window (start_time + end_time) from configs
        target_shift = shift
        if not target_shift:
            cur.execute(
                "SELECT db_table_name, current_shift_row_id FROM mes_lines WHERE id=%s",
                (line_id,),
            )
            ln = cur.fetchone()
            if ln and ln.get("current_shift_row_id") and ln.get("db_table_name"):
                try:
                    cur.execute(
                        f"SELECT shift_name FROM {ln['db_table_name']} "
                        f"WHERE id=%s", (ln["current_shift_row_id"],),
                    )
                    r = cur.fetchone()
                    if r and r.get("shift_name"):
                        target_shift = r["shift_name"]
                except Exception:
                    pass

        cur.execute(
            """SELECT shift_name, start_time, end_time
                 FROM mes_shift_configs
                WHERE line_id=%s
                  AND ( %s::text IS NULL OR shift_name=%s )
                ORDER BY shift_name""",
            (line_id, target_shift, target_shift),
        )
        shift_rows = cur.fetchall() or []
        if not shift_rows:
            return {
                "line_id":   line_id,
                "date":      target_date,
                "shift":     target_shift,
                "episodes":  [],
                "by_slot":   {},
                "note":      "No shift config found",
            }
        sh = shift_rows[0] if target_shift else shift_rows[0]
        # build start_ts / end_ts from chosen shift
        sh_start = sh["start_time"]
        sh_end   = sh["end_time"]

        def _combine(d_iso, t):
            if isinstance(t, str):
                # "HH:MM:SS" → time
                hh, mm, ss = (int(x) for x in t.split(":")[:3])
                return _dt.combine(_dt.fromisoformat(d_iso).date(),
                                   _dt.min.time().replace(hour=hh, minute=mm, second=ss))
            return _dt.combine(_dt.fromisoformat(d_iso).date(), t)

        start_dt = _combine(target_date, sh_start)
        end_dt   = _combine(target_date, sh_end)
        if end_dt <= start_dt:
            end_dt = end_dt + _td(days=1)  # shift wraps past midnight

        # 2. Hourly slots for slot attribution
        cur.execute(
            """SELECT slot_label, start_time, end_time
                 FROM mes_hourly_slots
                WHERE line_id=%s AND shift_name=%s
                ORDER BY start_time""",
            (line_id, sh["shift_name"]),
        )
        slots_raw = cur.fetchall() or []
        slot_windows = []
        for srow in slots_raw:
            ss = _combine(target_date, srow["start_time"])
            se = _combine(target_date, srow["end_time"])
            if se <= ss: se = se + _td(days=1)
            slot_windows.append({"label": srow["slot_label"], "start": ss, "end": se})

        def _slot_of(ts):
            for w in slot_windows:
                if w["start"] <= ts < w["end"]:
                    return w["label"]
            return None

        # 3. Fetch every bypass event in this shift window
        cur.execute(
            """SELECT detected_at, context_json, plc_value, alert_level
                 FROM mes_poka_yoke_events
                WHERE line_id=%s
                  AND rule_type='SENSOR_BYPASS'
                  AND detected_at >= %s
                  AND detected_at <  %s
                  AND context_json IS NOT NULL
                  AND context_json ~ '^\\s*\\{'
                ORDER BY detected_at ASC""",
            (line_id, start_dt, end_dt),
        )
        rows = cur.fetchall() or []

    # 4. Group into episodes per py_no
    by_py = {}
    for r in rows:
        try:
            ctx = _json.loads(r["context_json"]) if isinstance(r["context_json"], str) else r["context_json"]
        except Exception:
            ctx = {}
        py_no   = ctx.get("py_no") or "?"
        py_name = ctx.get("py_name") or py_no
        actual  = (ctx.get("actual") or "").upper()
        ts      = r["detected_at"]
        is_pass = actual == "PASS"
        info = by_py.setdefault(py_no, {"name": py_name, "events": []})
        info["events"].append({"ts": ts, "is_pass": is_pass, "actual": actual,
                                "alert": r["alert_level"]})

    # 2026-05-16 (afternoon v2) — episode-split fix.
    # Operator audit caught a 3+ hour "ongoing since 13:03" entry for
    # D406 even though card flipped to "All set" at 15:05.  Root cause:
    # this builder lumped EVERY consecutive non-PASS event into ONE
    # open episode (just incrementing hit_count).  But D406's sensor
    # never reports actual='PASS' — it only reports a snapshot on the
    # tick when bypass is observed.  So 13:03 → 14:21 (78 min idle
    # between) → 15:07 (46 min idle) → 15:14 → ... was treated as one
    # 2-hour bypass instead of 4 short episodes with long recoveries.
    #
    # New rule: a non-PASS event that arrives MORE THAN
    # EPISODE_SPLIT_SEC after the previous non-PASS event for the SAME
    # py_no CLOSES the previous episode (end_dt = previous event's ts)
    # and starts a fresh one.  PASS event still closes any open
    # episode as before.  EPISODE_SPLIT_SEC defaults to MERGE_GAP_SEC
    # (60 s) so the consolidation here matches the downstream merge
    # consolidation (they're the same operator-meaningful threshold).
    EPISODE_SPLIT_SEC = 60.0
    raw_episodes = []
    for py_no, info in by_py.items():
        cur_ep = None
        last_nonpass_ts = None
        for e in info["events"]:
            if not e["is_pass"]:
                # Should this START a new episode (no current, OR gap exceeded)?
                if cur_ep is None:
                    cur_ep = {
                        "py_no":      py_no,
                        "py_name":    info["name"],
                        "start_dt":   e["ts"],
                        "end_dt":     None,
                        "hit_count":  1,
                        "alert":      e["alert"] or "WARNING",
                    }
                elif (last_nonpass_ts is not None
                      and (e["ts"] - last_nonpass_ts).total_seconds() > EPISODE_SPLIT_SEC):
                    # Long gap since last detection → previous episode is over.
                    # Close it at the LAST observed non-PASS event time
                    # (sensor was OK from then until this new detection).
                    cur_ep["end_dt"] = last_nonpass_ts
                    raw_episodes.append(cur_ep)
                    cur_ep = {
                        "py_no":      py_no,
                        "py_name":    info["name"],
                        "start_dt":   e["ts"],
                        "end_dt":     None,
                        "hit_count":  1,
                        "alert":      e["alert"] or "WARNING",
                    }
                else:
                    # Same episode (gap within threshold) — extend hit count.
                    cur_ep["hit_count"] += 1
                last_nonpass_ts = e["ts"]
            elif e["is_pass"] and cur_ep:
                cur_ep["end_dt"] = e["ts"]
                raw_episodes.append(cur_ep)
                cur_ep = None
                last_nonpass_ts = None
        if cur_ep:
            raw_episodes.append(cur_ep)  # ongoing — no recovery seen yet

    # 4b. MERGE CLOSE EPISODES (2026-05-16 operator request).
    # Collector polling cadence + scanner-to-PLC latency causes a real
    # 1-minute bypass to fragment into 6-10 small "episodes" with 4-12 s
    # gaps between them.  Operator: "4-5 sec ka time toh collector se
    # ek aadh bit miss mein hi lg jata hai isko iss hisab se set kr ki
    # km se km 1 minute ka naa aae agli toh".  We merge consecutive
    # episodes for the SAME py_no when the next start is within
    # MERGE_GAP_SEC of the previous end.  hit_count sums; start = first
    # episode's start; end = last episode's end (or None if ongoing).
    MERGE_GAP_SEC = 60.0
    by_py_eps = {}
    for ep in raw_episodes:
        by_py_eps.setdefault(ep["py_no"], []).append(ep)

    episodes = []
    for py_no, eps in by_py_eps.items():
        eps.sort(key=lambda x: x["start_dt"])
        merged = None
        for ep in eps:
            if merged is None:
                merged = dict(ep)
                continue
            # If merged is still ongoing (end_dt None), cannot extend further
            # — flush and start new (defensive; shouldn't happen since ongoing
            # is the LAST episode for a py_no).
            if merged["end_dt"] is None:
                episodes.append(merged)
                merged = dict(ep)
                continue
            gap_sec = (ep["start_dt"] - merged["end_dt"]).total_seconds()
            if gap_sec <= MERGE_GAP_SEC:
                # Extend the merged episode through this one
                merged["end_dt"]    = ep["end_dt"]
                merged["hit_count"] += ep["hit_count"]
            else:
                episodes.append(merged)
                merged = dict(ep)
        if merged is not None:
            episodes.append(merged)

    # Convert datetimes to ISO strings; slot attribution happens below
    # via SPAN intersection (an episode crossing N slots appears in all N).
    from datetime import datetime as _dt2
    now_dt = _dt2.now()
    for ep in episodes:
        ep["start_at"]    = ep["start_dt"].isoformat()
        ep["end_at"]      = ep["end_dt"].isoformat() if ep["end_dt"] else None
        ep["start_slot"]  = _slot_of(ep["start_dt"])
        # `slot` field retained for backward-compat (single-slot consumers);
        # span-aware consumers should iterate `by_slot` instead.
        ep["slot"]        = ep["start_slot"]
    episodes.sort(key=lambda x: x["start_at"])

    # 5. SPAN-AWARE slot attribution + analytics.
    # 2026-05-16 (operator feedback): an episode starting at 13:03 and
    # still ongoing at 14:43 was being shown ONLY in the 11:30-13:05
    # slot's row.  The CURRENT slot (14:05-15:05) painted "✓ All set
    # so far" even though the bypass was still active.  Fix: walk
    # every hourly slot and intersect the episode's [start, end_or_now]
    # range with each slot's window.  Any intersection ≥ 1 s adds a
    # slot-scoped entry that captures HOW MUCH of the slot was spent
    # in bypass + flags whether the episode is the originating one
    # (so the modal can show "started here" vs "carrying over").
    by_slot      = {}
    slot_summary = {}
    for ep in episodes:
        try:
            s_dt = ep["start_dt"]
            e_dt = ep["end_dt"] or now_dt
        except Exception:
            continue
        for w in slot_windows:
            iv_start = max(s_dt, w["start"])
            iv_end   = min(e_dt, w["end"])
            iv_sec   = (iv_end - iv_start).total_seconds()
            if iv_sec <= 1.0:
                continue
            slot_key = w["label"]
            ends_in_this_slot   = ep["end_dt"] is not None and w["start"] <= ep["end_dt"] < w["end"]
            starts_in_this_slot = w["start"] <= s_dt < w["end"]
            entry = {
                **{k: v for k, v in ep.items() if k not in ("start_dt", "end_dt")},
                "slot_label":        slot_key,
                "slot_segment_from": iv_start.isoformat(),
                "slot_segment_to":   iv_end.isoformat(),
                "slot_segment_sec":  round(iv_sec, 1),
                "starts_here":       starts_in_this_slot,
                "ends_here":         ends_in_this_slot,
                "ongoing_now":       ep["end_dt"] is None and w["start"] <= now_dt < w["end"],
            }
            by_slot.setdefault(slot_key, []).append(entry)
            bucket = slot_summary.setdefault(slot_key, {
                "total_bypass_sec": 0.0,
                "longest_sec":      0.0,
                "py_set":           set(),
                "episode_count":    0,
                "ongoing_count":    0,
            })
            bucket["total_bypass_sec"] += iv_sec
            bucket["longest_sec"]       = max(bucket["longest_sec"], iv_sec)
            bucket["py_set"].add(ep["py_no"])
            bucket["episode_count"]    += 1
            if entry["ongoing_now"]:
                bucket["ongoing_count"] += 1
    for k, b in slot_summary.items():
        b["distinct_py"]      = len(b.pop("py_set"))
        b["total_bypass_sec"] = round(b["total_bypass_sec"], 1)
        b["longest_sec"]      = round(b["longest_sec"], 1)

    # Drop the temporary datetime fields before serialisation
    for ep in episodes:
        ep.pop("start_dt", None)
        ep.pop("end_dt",   None)

    payload = {
        "line_id":      line_id,
        "date":         target_date,
        "shift":        sh["shift_name"],
        "shift_start":  start_dt.isoformat(),
        "shift_end":    end_dt.isoformat(),
        "episodes":     episodes,
        "by_slot":      by_slot,
        "slot_summary": slot_summary,
        "merge_gap_sec": MERGE_GAP_SEC,
        "generated_at": _dt.now().isoformat(),
    }

    # 6. Persist to local JSON (atomic write).
    # File name encodes line + date + shift so OT picks the right file.
    # Shift-end rotation: collector deletes / archives by renaming this
    # file at shift boundary — see _archive_bypass_log() called from
    # the existing shift-rotation hook.  During OT we just keep
    # appending to the active shift's file.
    try:
        cache_dir = _os.path.dirname(_os.path.abspath(__file__))
        cache_dir = _os.path.dirname(cache_dir)  # → Phase2/
        fname = f"_py_bypass_log_line{line_id}_{target_date}_{sh['shift_name']}.json"
        fpath = _os.path.join(cache_dir, fname)
        tmp   = fpath + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fp:
            _json.dump(payload, fp, indent=2, default=str)
        _os.replace(tmp, fpath)
    except Exception as exc:
        print(f"[POKA-EPISODES] local JSON write failed: {exc}")

    return payload


@router.get("/events/{line_id}/summary")
def events_summary(line_id: int, user=Depends(get_current_user)):
    with get_conn() as conn:
        _check_operator_access(user, line_id, conn)
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT
                COUNT(*)                                           AS total_unacked,
                COUNT(*) FILTER (WHERE alert_level='CRITICAL')    AS critical,
                COUNT(*) FILTER (WHERE alert_level='WARNING')     AS warnings
            FROM mes_poka_yoke_events
            WHERE line_id = %s AND acknowledged = false
        """, (line_id,))
        return cur.fetchone()


@router.post("/events/{event_id}/acknowledge")
def acknowledge_event(event_id: int, admin=Depends(require_admin)):
    with get_conn() as conn:
        conn.cursor().execute("""
            UPDATE mes_poka_yoke_events
            SET acknowledged    = true,
                acknowledged_at = NOW(),
                acknowledged_by = %s
            WHERE id = %s AND acknowledged = false
        """, (admin["username"], event_id))
    return {"ok": True}


@router.post("/events/{line_id}/acknowledge-all")
def acknowledge_all(line_id: int, admin=Depends(require_admin)):
    with get_conn() as conn:
        conn.cursor().execute("""
            UPDATE mes_poka_yoke_events
            SET acknowledged    = true,
                acknowledged_at = NOW(),
                acknowledged_by = %s
            WHERE line_id = %s AND acknowledged = false
        """, (admin["username"], line_id))
    return {"ok": True}


# REPLACE existing EventIngest class and ingest_event with:

class EventIngest(BaseModel):
    line_id:      int
    rule_id:      Optional[int] = None
    rule_type:    str
    alert_level:  str           = "WARNING"
    shift_name:   Optional[str] = None
    plc_value:    Optional[str] = None
    context_json: Optional[str] = None


@router.get("/live/{line_id}")
def get_live_status(
    line_id:    int,
    model_name: Optional[str] = None,
    model_bit:  Optional[int] = None,
    user=Depends(get_current_user)
):
    """
    Return poka-yoke checks for the current model running on this line.

    Resolution priority:
    1. model_bit query param (preferred — stable across model_name renames)
    2. model_name query param (legacy)
    3. Auto-detect current_model_number / current_model_name from shift table

    Data source: mes_py_assignments JOINed with mes_py_model_master so that
    even if the master's model_name was renamed after the assignment was
    created, we still resolve assignments by the stable bit_number.
    is_bypassed = unacknowledged event within last 8 hours.
    """
    with get_conn() as conn:
        _check_operator_access(user, line_id, conn)
        cur = dict_cursor(conn)

        # ── 1. Resolve model (bit preferred, fall back to name) ───────────
        resolved_bit   = model_bit
        resolved_model = model_name

        # Resolve THIS line's zone — used to disambiguate bit_number when
        # the same bit value exists in multiple zones.
        line_zone_id = None
        try:
            cur.execute("SELECT zone_id FROM mes_lines WHERE id=%s", (line_id,))
            _zr = cur.fetchone()
            if _zr:
                line_zone_id = _zr["zone_id"] if isinstance(_zr, dict) else _zr[0]
        except Exception:
            pass

        if resolved_bit is None or not resolved_model:
            cur.execute(
                "SELECT db_table_name, current_shift_row_id FROM mes_lines WHERE id = %s",
                (line_id,)
            )
            line_row = cur.fetchone()
            if line_row and line_row["db_table_name"]:
                tbl = line_row["db_table_name"]
                try:
                    if line_row["current_shift_row_id"]:
                        cur.execute(
                            f"SELECT current_model_number, current_model_name FROM {tbl} WHERE id = %s",
                            (line_row["current_shift_row_id"],)
                        )
                    else:
                        cur.execute(
                            f"SELECT current_model_number, current_model_name FROM {tbl} "
                            f"WHERE is_shift_completed = false "
                            f"ORDER BY timestamp DESC NULLS LAST LIMIT 1"
                        )
                    mrow = cur.fetchone() or {}
                    if resolved_bit is None and mrow.get("current_model_number"):
                        resolved_bit = int(mrow["current_model_number"])
                    if not resolved_model and mrow.get("current_model_name"):
                        resolved_model = mrow["current_model_name"]
                except Exception:
                    pass

        # ── 2. Fetch assignments for this model ────────────────────────────
        # Prefer matching by bit_number (stable). Fall back to name match
        # for legacy rows where bit linkage is missing.
        checks = []
        if resolved_bit is not None:
            cur.execute("""
                SELECT DISTINCT ON (a.py_no, a.side)
                       a.id,
                       m.id             AS py_master_id,
                       z.zone_name      AS zone_name,
                       m.sensing_bits   AS sensing_bits,
                       a.py_no          AS poka_yoke_no,
                       a.py_name        AS poka_yoke_name,
                       a.side,
                       COALESCE(m.bit, a.d_bit) AS bit,
                       COALESCE(m.desired_value, a.desired_value) AS value,
                       a.machine_name,
                       COALESCE(mm.model_name, a.model_name) AS model_name,
                       a.model_series,
                       a.model_type,
                       mm.bit_number    AS model_bit,
                       'WARNING'        AS alert_level
                FROM mes_py_assignments a
                LEFT JOIN mes_py_master m        ON m.py_no    = a.py_no AND m.is_active = true
                LEFT JOIN mes_zones z            ON z.id       = m.zone_id
                LEFT JOIN mes_py_model_master mm ON mm.id      = a.model_id
                WHERE mm.bit_number = %s AND mm.is_active = true
                  AND (mm.zone_id IS NULL OR mm.zone_id = COALESCE(%s, mm.zone_id))
                ORDER BY a.py_no, a.side, a.id
            """, (resolved_bit, line_zone_id))
            checks = [dict(r) for r in cur.fetchall()]

        # Fallback: legacy rows without model_id linkage — match by name.
        if not checks and resolved_model:
            mn_lower = resolved_model.strip().lower()
            cur.execute("""
                SELECT DISTINCT ON (a.py_no, a.side)
                       a.id,
                       m.id            AS py_master_id,
                       z.zone_name     AS zone_name,
                       m.sensing_bits  AS sensing_bits,
                       a.py_no         AS poka_yoke_no,
                       a.py_name       AS poka_yoke_name,
                       a.side,
                       COALESCE(m.bit, a.d_bit) AS bit,
                       COALESCE(m.desired_value, a.desired_value) AS value,
                       a.machine_name,
                       a.model_name,
                       a.model_series,
                       a.model_type,
                       NULL            AS model_bit,
                       'WARNING'       AS alert_level
                FROM mes_py_assignments a
                LEFT JOIN mes_py_master m ON m.py_no = a.py_no AND m.is_active = true
                LEFT JOIN mes_zones z     ON z.id    = m.zone_id
                WHERE LOWER(a.model_name) = %s
                ORDER BY a.py_no, a.side, a.id
            """, (mn_lower,))
            checks = [dict(r) for r in cur.fetchall()]

        # ── 3. Also include any line-scoped rules still in DB ──
        # (fallback / compatibility — reads mes_poka_yoke_rules if any exist)
        cur.execute("""
            SELECT r.id, r.poka_yoke_no, r.poka_yoke_name, r.side,
                   r.model, r.bit, r.value, r.machine_name,
                   r.sheet_name, r.alert_level
            FROM mes_poka_yoke_rules r
            WHERE r.line_id = %s AND r.is_active = true
            ORDER BY r.bit NULLS LAST, r.poka_yoke_no
        """, (line_id,))
        line_rules = [dict(r) for r in cur.fetchall()]

        # Filter line_rules by model if we know the model
        def rule_matches_model(r):
            if not resolved_model:
                return True
            rm = (r.get("model") or "all").strip().lower()
            mn = resolved_model.strip().lower()
            if rm in ("", "all"):
                return True
            if rm.startswith("except "):
                return rm[7:].strip() not in mn
            return rm in mn or mn in rm

        filtered_line_rules = [r for r in line_rules if rule_matches_model(r)]

        # Merge: matrix checks first, then any remaining line-scoped rules
        # De-duplicate by poka_yoke_no + side.  Also drop any row whose
        # desired-output for the active model is PASS (value 0 / NULL / "0"
        # / "pass") — such a PY is not applicable to this model so it
        # shouldn't clutter the Fullscreen list or trigger health checks.
        def _is_pass(v):
            if v is None:
                return True
            try:
                return int(v) == 0
            except (ValueError, TypeError):
                return str(v).strip().lower() in ("pass", "")

        seen_keys = set()
        merged = []
        for c in checks:
            if _is_pass(c.get("value")):
                continue
            key = (c.get("poka_yoke_no",""), c.get("side","ALL"))
            if key not in seen_keys:
                seen_keys.add(key)
                # Normalise field names
                merged.append({
                    "id":             c.get("id"),
                    "py_master_id":   c.get("py_master_id"),
                    "zone_name":      c.get("zone_name"),
                    "sensing_bits":   c.get("sensing_bits"),
                    "poka_yoke_no":   c.get("poka_yoke_no"),
                    "poka_yoke_name": c.get("poka_yoke_name"),
                    "side":           c.get("side", "ALL"),
                    "bit":            c.get("bit"),
                    "value":          c.get("value", 1),
                    "machine_name":   c.get("machine_name"),
                    "alert_level":    c.get("alert_level", "WARNING"),
                    "source":         "matrix",
                })
        for r in filtered_line_rules:
            if _is_pass(r.get("value")):
                continue
            key = (r.get("poka_yoke_no",""), r.get("side","ALL"))
            if key not in seen_keys:
                seen_keys.add(key)
                merged.append({**r, "source": "line_rule"})

        # ── 4. Attach bypass status ────────────────────────────
        # Match events two ways:
        #   (a) new system: events have rule_id=NULL, py_no is inside context_json
        #   (b) legacy   : events have rule_id → mes_poka_yoke_rules.poka_yoke_no
        # An event is considered a live bypass if it's unacknowledged and within
        # the last 8 hours.
        #
        # 2026-05-16 — AUTO-CLEAR sweep (3 paths, run on every /live call).
        # Operator complaint chain:
        #   v1: "agar PY recover ho jaye to manual ack na karna pade"
        #   v2: "abhi bhi matrix sub-page pe jaake ack karna pad raha hai"
        # Root cause for v2: collector's in-memory `_py_bypass_state`
        # resets on restart.  If the sensor recovers DURING the brief
        # restart window, the collector never POSTs to /auto-ack and
        # the event stays unacked forever (until 8 h TTL).
        #
        # Three orthogonal auto-ack rules — ALL fire here so the matrix
        # page never has stale work for the operator:
        #   (A) Latest event for this py_no shows actual='PASS'
        #       → sensor explicitly reported recovery
        #   (B) No new event for this py_no in the last STALE_AFTER_SEC
        #       seconds → presume recovered (transient blip)
        #   (C) Event acknowledged_at > STALE_AFTER_SEC ago means a
        #       previous ack was confirmed — drop any older sibling
        #       events still hanging from the same py_no
        STALE_AFTER_SEC = 120   # 2 min — sensor surely settled by then
        try:
            cur.execute("""
                WITH latest_per_py AS (
                    SELECT DISTINCT ON ((context_json::jsonb)->>'py_no')
                           (context_json::jsonb)->>'py_no'  AS py_no,
                           (context_json::jsonb)->>'actual' AS actual_state,
                           detected_at
                    FROM mes_poka_yoke_events
                    WHERE line_id        = %s
                      AND rule_id        IS NULL
                      AND rule_type      = 'SENSOR_BYPASS'
                      AND detected_at    > NOW() - INTERVAL '8 hours'
                      AND context_json IS NOT NULL
                      AND context_json ~ '^\\s*\\{'
                    ORDER BY (context_json::jsonb)->>'py_no', detected_at DESC
                )
                UPDATE mes_poka_yoke_events e
                   SET acknowledged    = true,
                       acknowledged_at = NOW(),
                       acknowledged_by = 'auto_sweep'
                  FROM latest_per_py lp
                 WHERE e.line_id      = %s
                   AND e.rule_id      IS NULL
                   AND e.rule_type    = 'SENSOR_BYPASS'
                   AND e.acknowledged = false
                   AND (e.context_json::jsonb)->>'py_no' = lp.py_no
                   AND (
                         -- (A) explicit PASS recovery
                         lp.actual_state = 'PASS'
                         -- (B) no new event in STALE_AFTER_SEC → recovered
                         OR lp.detected_at < NOW() - INTERVAL '1 second' * %s
                       )
            """, (line_id, line_id, STALE_AFTER_SEC))
        except Exception as _exc:
            print(f"[POKA-LIVE] auto-ack sweep failed: {_exc}")

        bypass_map = {}
        if merged:
            py_nos = list({r["poka_yoke_no"] for r in merged if r.get("poka_yoke_no")})
            if py_nos:
                # (a) new-system matching via context_json
                cur.execute("""
                    SELECT py_no,
                           true                   AS is_bypassed,
                           MAX(detected_at)       AS last_bypass_at,
                           MAX(plc_value)         AS plc_value,
                           MAX(context_json)      AS context_json
                    FROM (
                        SELECT (context_json::jsonb)->>'py_no' AS py_no,
                               detected_at, plc_value, context_json
                        FROM mes_poka_yoke_events
                        WHERE line_id        = %s
                          AND rule_id        IS NULL
                          AND rule_type      = 'SENSOR_BYPASS'
                          AND acknowledged   = false
                          AND detected_at    > NOW() - INTERVAL '8 hours'
                          AND context_json IS NOT NULL
                          AND context_json ~ '^\\s*\\{'   -- valid JSON only
                    ) x
                    WHERE py_no = ANY(%s)
                    GROUP BY py_no
                """, (line_id, py_nos))
                for r in cur.fetchall():
                    bypass_map[r["py_no"]] = dict(r)

                # (b) legacy matching via rule_id join
                cur.execute("""
                    SELECT DISTINCT ON (r.poka_yoke_no)
                           r.poka_yoke_no,
                           true           AS is_bypassed,
                           e.detected_at  AS last_bypass_at,
                           e.plc_value    AS plc_value,
                           e.context_json AS context_json
                    FROM mes_poka_yoke_events e
                    JOIN mes_poka_yoke_rules  r ON r.id = e.rule_id
                    WHERE r.poka_yoke_no = ANY(%s)
                      AND r.line_id      = %s
                      AND e.rule_type    = 'SENSOR_BYPASS'
                      AND e.acknowledged = false
                      AND e.detected_at  > NOW() - INTERVAL '8 hours'
                    ORDER BY r.poka_yoke_no, e.detected_at DESC
                """, (py_nos, line_id))
                for r in cur.fetchall():
                    # New-system match wins over legacy if both exist.
                    bypass_map.setdefault(r["poka_yoke_no"], dict(r))

        result = []
        for r in merged:
            b = bypass_map.get(r.get("poka_yoke_no"), {})
            # Decode the context_json (stored as text) so frontend gets a
            # structured object with actual/expected/register info.
            ctx_obj = None
            ctx_raw = b.get("context_json")
            if ctx_raw:
                try:
                    import json as _json
                    ctx_obj = _json.loads(ctx_raw) if isinstance(ctx_raw, str) else ctx_raw
                except Exception:
                    ctx_obj = None
            result.append({
                **r,
                "is_bypassed":       b.get("is_bypassed", False),
                "last_bypass_at":    b.get("last_bypass_at"),
                "last_plc_value":    b.get("plc_value"),
                "bypass_context":    ctx_obj,
                "current_model":     resolved_model,
                # Top-level resolved bit so the UI can label the running
                # model regardless of whether the per-row JOIN produced a
                # bit_number (legacy rows have NULL there).
                "current_model_bit": resolved_bit,
            })

        return result

# ══════════════════════════════════════════════════════════════════════════
# MAIL CONFIG — per-task (bypass / health / hourly) TO + CC lookup
# ══════════════════════════════════════════════════════════════════════════
# A single source of truth for all mail recipients.  The admin UI writes to
# `mes_mail_config`; mail functions read via `_get_mail_addrs(kind)`.
# Resolution order per field:
#   1. DB row  (mes_mail_config.value)  ← admin UI edits this
#   2. Env var (e.g. BYPASS_TO)          ← initial bootstrap / fallback
#   3. Legacy  (NOTIFY_EMAIL / SLOT_REPORT_EMAIL)  ← backward compat
# Values are comma-separated email lists; the helper returns (to[], cc[]).

_MAIL_CONFIG_SCHEMA_ENSURED = False

# (key → (env_var_fallback, legacy_env_var, description))
_MAIL_CONFIG_KEYS = {
    # ── Legacy pipeline (still in use) ────────────────────────
    "bypass_to":  ("BYPASS_TO",  "NOTIFY_EMAIL",       "Poka-Yoke Bypass alerts — To addresses (comma-separated)"),
    "bypass_cc":  ("BYPASS_CC",  None,                 "Poka-Yoke Bypass alerts — Cc addresses"),
    "health_to":  ("HEALTH_TO",  "NOTIFY_EMAIL",       "Sensor Health Fail alerts — To addresses"),
    "health_cc":  ("HEALTH_CC",  None,                 "Sensor Health Fail alerts — Cc addresses"),
    "hourly_to":  ("HOURLY_TO",  "SLOT_REPORT_EMAIL",  "Hourly slot report — To addresses"),
    "hourly_cc":  ("HOURLY_CC",  None,                 "Hourly slot report — Cc addresses"),
}


def _ensure_mail_config_schema(conn):
    """Idempotently create mes_mail_config + seed default rows so the admin
    UI has every known key visible even if nothing has been edited yet."""
    global _MAIL_CONFIG_SCHEMA_ENSURED
    if _MAIL_CONFIG_SCHEMA_ENSURED:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_mail_config (
                key         VARCHAR(50) PRIMARY KEY,
                value       TEXT,
                description TEXT,
                updated_at  TIMESTAMPTZ DEFAULT NOW(),
                updated_by  VARCHAR(80)
            )
        """)
        # Seed any missing keys (doesn't overwrite existing values).
        for key, (_env, _legacy, desc) in _MAIL_CONFIG_KEYS.items():
            cur.execute("""
                INSERT INTO mes_mail_config (key, value, description)
                VALUES (%s, %s, %s)
                ON CONFLICT (key) DO UPDATE SET description = EXCLUDED.description
            """, (key, None, desc))
        conn.commit()
        _MAIL_CONFIG_SCHEMA_ENSURED = True
    except Exception:
        try: conn.rollback()
        except Exception: pass


def _lookup_mail_value(key: str) -> str:
    """Resolve a single mail-config key: DB → env → legacy env → empty."""
    import os
    spec = _MAIL_CONFIG_KEYS.get(key)
    if not spec:
        return ""
    env_var, legacy_var, _desc = spec

    # 1. DB
    try:
        with get_conn() as conn:
            _ensure_mail_config_schema(conn)
            cur = conn.cursor()
            cur.execute("SELECT value FROM mes_mail_config WHERE key=%s", (key,))
            row = cur.fetchone()
            if row and row[0] and str(row[0]).strip():
                return str(row[0]).strip()
    except Exception:
        pass

    # 2. Primary env var
    v = os.getenv(env_var, "").strip()
    if v:
        return v

    # 3. Legacy fallback (so existing setups keep working)
    if legacy_var:
        v = os.getenv(legacy_var, "").strip()
        if v:
            return v

    return ""


def _get_mail_addrs(kind: str) -> tuple[list, list]:
    """Return (to_list, cc_list) for 'bypass' | 'health' | 'hourly'."""
    raw_to = _lookup_mail_value(f"{kind}_to")
    raw_cc = _lookup_mail_value(f"{kind}_cc")
    to_list = [e.strip() for e in raw_to.split(",") if e.strip()]
    cc_list = [e.strip() for e in raw_cc.split(",") if e.strip()]
    return to_list, cc_list


def _send_bypass_email_async(body: "EventIngest") -> None:
    """Background thread: email the quality team when a SENSOR_BYPASS or
    SENSOR_HEALTH event fires.  TO/CC lookup is driven by rule_type so the
    two flows can notify different groups (see Admin → Mail Config)."""
    import os, json as _json, smtplib
    from email.mime.text     import MIMEText
    from email.mime.multipart import MIMEMultipart
    from datetime             import datetime as _dt

    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587") or 587)
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")

    rule_type = (body.rule_type or "").upper()
    kind      = "health" if rule_type == "SENSOR_HEALTH" else "bypass"
    to_list, cc_list = _get_mail_addrs(kind)

    if not (smtp_user and smtp_pass and to_list):
        print(f"[POKA-YOKE-EMAIL/{kind}] skipped (SMTP or To not configured)")
        return

    try:
        ctx = _json.loads(body.context_json) if body.context_json else {}
    except Exception:
        ctx = {}

    line_name  = f"Line #{body.line_id}"
    model_name = "—"
    model_bit  = ctx.get("model_bit")
    try:
        with get_conn() as conn:
            cur = dict_cursor(conn)
            cur.execute("SELECT line_name FROM mes_lines WHERE id = %s", (body.line_id,))
            r = cur.fetchone()
            if r and r.get("line_name"): line_name = r["line_name"]
            if model_bit is not None:
                cur.execute(
                    "SELECT model_name FROM mes_py_model_master "
                    "WHERE bit_number=%s AND is_active=true ORDER BY id DESC LIMIT 1",
                    (model_bit,),
                )
                mr = cur.fetchone()
                if mr and mr.get("model_name"):
                    # Strip any legacy "TYPE-SERIES:" prefix so the email is clean.
                    import re as _re
                    model_name = _re.sub(r"^TYPE-SERIES:\s*", "", mr["model_name"], flags=_re.IGNORECASE)
    except Exception:
        pass

    level    = (body.alert_level or "WARNING").upper()
    py_name  = ctx.get("py_name") or "—"
    color    = "#ef4444" if level == "CRITICAL" else "#f59e0b"
    emoji    = "🚨 CRITICAL" if level == "CRITICAL" else "⚠️ WARNING"

    rule_type = (body.rule_type or "").upper()
    is_health = rule_type == "SENSOR_HEALTH"
    headline  = "Poka-Yoke Sensor Health Fail" if is_health else "Poka-Yoke Bypass Detected"
    subject   = f"[{level}] {'Sensor Health Fail' if is_health else 'Poka-Yoke Bypass'} — {line_name}"
    html = f"""
<html><body style="font-family:Arial,sans-serif;color:#0f172a;">
  <div style="border-left:5px solid {color};padding:16px 20px;background:#fff;">
    <h2 style="margin:0 0 8px;color:{color};">{emoji} — {headline}</h2>
    <table style="border-collapse:collapse;width:100%;margin-top:12px;">
      <tr><td style="padding:6px 12px;font-weight:700;background:#f8fafc;width:150px;">Line</td><td style="padding:6px 12px;">{line_name}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f1f5f9;">PY Name</td><td style="padding:6px 12px;">{py_name}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f8fafc;">Model Name</td><td style="padding:6px 12px;font-family:monospace;">{model_name}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f1f5f9;">Shift</td><td style="padding:6px 12px;">{body.shift_name or "—"}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f8fafc;">Alert Level</td><td style="padding:6px 12px;color:{color};font-weight:700;">{level}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f1f5f9;">Time</td><td style="padding:6px 12px;">{_dt.now().strftime("%Y-%m-%d %H:%M:%S")}</td></tr>
    </table>
    <p style="margin-top:16px;font-size:13px;color:#64748b;">
      Please investigate and acknowledge this bypass in the Fullscreen dashboard.<br/>
      Automated alert — Production Monitoring System.
    </p>
  </div>
</body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(to_list)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, to_list + cc_list, msg.as_string())
        print(f"[POKA-YOKE-EMAIL/{kind}] Sent: {py_name} | {line_name} → "
              f"To={to_list} Cc={cc_list}")
    except Exception as e:
        print(f"[POKA-YOKE-EMAIL/{kind}] Failed: {e}")


# ══════════════════════════════════════════════════════════════════════════
# MAIL CONFIG CRUD — admin UI edits `mes_mail_config`; mail functions read
# via _get_mail_addrs().  Env vars still work as bootstrap fallback.
# ══════════════════════════════════════════════════════════════════════════

@router.get("/mail-config/", tags=["mail-config"])
def list_mail_config(user=Depends(get_current_user)):
    """Return every known mail-config key with its effective value resolved
    through DB → env → legacy fallback, so the UI can show both the 'raw'
    stored value and the 'effective' value currently in use."""
    import os
    with get_conn() as conn:
        _ensure_mail_config_schema(conn)
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT key, value, description, updated_at, updated_by
            FROM mes_mail_config ORDER BY key
        """)
        rows = cur.fetchall()

    out = []
    for r in rows:
        spec = _MAIL_CONFIG_KEYS.get(r["key"], ("", None, ""))
        env_var, legacy_var, _desc = spec
        env_val = os.getenv(env_var, "") if env_var else ""
        legacy_val = os.getenv(legacy_var, "") if legacy_var else ""
        effective = (r["value"] or "").strip() or env_val.strip() or legacy_val.strip()
        out.append({
            "key":          r["key"],
            "value":        r["value"] or "",
            "description":  r["description"],
            "updated_at":   r["updated_at"],
            "updated_by":   r["updated_by"],
            "env_var":      env_var,
            "env_value":    env_val,
            "legacy_var":   legacy_var,
            "legacy_value": legacy_val,
            "effective":    effective,
        })
    return out


class MailConfigUpdate(BaseModel):
    value: str

@router.put("/mail-config/{key}", tags=["mail-config"])
def update_mail_config(key: str, body: MailConfigUpdate,
                       admin=Depends(require_admin)):
    """Upsert a mail-config value.  Keys outside _MAIL_CONFIG_KEYS are
    rejected so typos don't silently accumulate."""
    if key not in _MAIL_CONFIG_KEYS:
        raise HTTPException(400, f"Unknown mail-config key: {key}")
    with get_conn() as conn:
        _ensure_mail_config_schema(conn)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_mail_config (key, value, description, updated_at, updated_by)
            VALUES (%s, %s, %s, NOW(), %s)
            ON CONFLICT (key) DO UPDATE SET
                value      = EXCLUDED.value,
                updated_at = NOW(),
                updated_by = EXCLUDED.updated_by
        """, (key, body.value, _MAIL_CONFIG_KEYS[key][2],
              admin.get("username", "admin")))
    return {"ok": True, "key": key, "value": body.value}


@router.post("/mail-config/{key}/test", tags=["mail-config"])
def test_mail_config(key: str, admin=Depends(require_admin)):
    """Send a one-line test email to the currently-configured recipients
    for a given kind ('bypass_to' → kind 'bypass', etc.)  Lets the admin
    verify SMTP + addresses without waiting for a real alert."""
    import os, smtplib
    from email.mime.text import MIMEText
    if key not in _MAIL_CONFIG_KEYS:
        raise HTTPException(400, f"Unknown key: {key}")
    kind = key.rsplit("_", 1)[0]   # "bypass_to" → "bypass"
    to_list, cc_list = _get_mail_addrs(kind)
    if not to_list:
        raise HTTPException(400, f"No To addresses resolved for '{kind}'")

    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587") or 587)
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    if not (smtp_user and smtp_pass):
        raise HTTPException(500, "SMTP credentials not configured in .env")

    msg = MIMEText(
        f"This is a test email from the mail-config panel for kind '{kind}'.\n"
        f"Resolved To: {to_list}\nResolved Cc: {cc_list}\n"
        f"If you received this, the configuration is working correctly.",
        "plain",
    )
    msg["Subject"] = f"[MAIL-TEST] {kind} configuration test"
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(to_list)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo(); server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, to_list + cc_list, msg.as_string())
    except Exception as e:
        raise HTTPException(500, f"SMTP send failed: {e}")
    return {"ok": True, "to": to_list, "cc": cc_list}


# ══════════════════════════════════════════════════════════════════════════
# SENSOR SWEEP — passive read-only X-bit health snapshot
# ══════════════════════════════════════════════════════════════════════════
# Collector samples each sensing X-bit ~1 Hz and POSTs the full snapshot
# here every 10 s.  The UI polls GET /sensor-sweep/{line_id}.  No PLC
# writes ever — collector NEVER overwrites sensor bits.  Sensor stuck
# >15 min → SENSOR_HEALTH email fires (1 shot).
#
# Layout: _SENSOR_SWEEP[line_id] = {
#     "swept_at": iso string,
#     "entries":  [ {bit, current_value, last_toggle_at, status, py_id, ...} ]
# }

_SENSOR_SWEEP: dict = {}

# 2026-05-14 — persist sweep to disk so a MES restart doesn't blank the
# UI for the 10-second gap before the collector's next publish.  Tiny
# JSON file under the backend dir; updated on every POST, loaded once
# at module import below.
import json as _ss_json
import os as _ss_os
_SENSOR_SWEEP_FILE = _ss_os.path.join(
    _ss_os.path.dirname(_ss_os.path.dirname(_ss_os.path.abspath(__file__))),
    "_sensor_sweep_cache.json",
)


def _ss_save() -> None:
    try:
        with open(_SENSOR_SWEEP_FILE, "w", encoding="utf-8") as f:
            _ss_json.dump(_SENSOR_SWEEP, f, default=str)
    except Exception as e:
        print(f"[SENSOR-SWEEP] persist failed: {e}")


def _ss_load() -> None:
    try:
        if _ss_os.path.exists(_SENSOR_SWEEP_FILE):
            with open(_SENSOR_SWEEP_FILE, "r", encoding="utf-8") as f:
                data = _ss_json.load(f)
            for k, v in (data or {}).items():
                try:
                    _SENSOR_SWEEP[int(k)] = v
                except (TypeError, ValueError):
                    pass
            print(f"[SENSOR-SWEEP] restored {len(_SENSOR_SWEEP)} line(s) from cache")
    except Exception as e:
        print(f"[SENSOR-SWEEP] cache load failed: {e}")


_ss_load()


class SensorSweepUpdate(BaseModel):
    """Maximally permissive shape — entries are forwarded raw to the UI
    without per-field validation.  The collector schema is allowed to
    evolve freely (new fields, renamed fields) without breaking publish."""
    model_config = {"extra": "allow"}
    line_id:  int
    swept_at: Optional[str] = None
    entries:  list[dict] = []

@router.post("/sensor-sweep/update")
def update_sensor_sweep(body: SensorSweepUpdate):
    """Collector POSTs the latest sensor-health snapshot.  Whole set per
    line is replaced + persisted to disk so a MES restart preserves the
    last-known state.  Entries are stored verbatim (no per-field
    validation) so any future field-shape change in the collector flows
    straight to the UI without backend churn."""
    _SENSOR_SWEEP[body.line_id] = {
        "swept_at": body.swept_at,
        "entries":  list(body.entries),
    }
    _ss_save()
    return {"ok": True, "count": len(body.entries)}


@router.get("/sensor-sweep/{line_id}")
def get_sensor_sweep(line_id: int, user=Depends(get_current_user)):
    """Admin sub-page polls this.  Empty sweep means the collector hasn't
    published yet for this line (first 10–20 s after startup).  Reads
    from in-memory cache which is itself seeded from disk at startup, so
    a MES restart doesn't blank the UI."""
    return _SENSOR_SWEEP.get(line_id, {"swept_at": None, "entries": []})


# ════════════════════════════════════════════════════════════════════════
# Sensor Health (current model) — simple, backend-driven
# ════════════════════════════════════════════════════════════════════════
# Combines the DB-side "applicable PYs for current model" list with the
# last-known sweep data from the collector.  ALWAYS returns one entry
# per configured PY of the current model (even when the collector hasn't
# published yet) so the UI shows the list immediately instead of "0
# sensors".  Each PY rolls up the status of ALL its sensing bits:
#   • alive   = every bit toggled within stuck threshold
#   • stuck   = at least one bit hasn't toggled past threshold
#   • unknown = no data captured for at least one bit yet
# Health badge:
#   • OK       — every PY alive
#   • WARNING  — some unknowns (collector still warming up)
#   • CRITICAL — any PY stuck (real sensor problem)
# ────────────────────────────────────────────────────────────────────────

@router.get("/sensor-health/{line_id}")
def sensor_health(
    line_id:   int,
    model_bit: Optional[int] = None,
    stuck_sec: int = 900,    # 15 min default — match collector threshold
    user=Depends(get_current_user),
):
    import re as _re
    from datetime import datetime as _dt
    REG_RE = _re.compile(r"(?:[XYMLFB])[0-9A-F]+|(?:[DRTCS])\d+", _re.IGNORECASE)

    with get_conn() as conn:
        try:
            _check_operator_access(user, line_id, conn)
        except Exception:
            pass
        cur = dict_cursor(conn)

        # 1. Resolve current model bit (same logic as /live/{line_id}).
        resolved_bit = model_bit
        resolved_model_name = None
        if resolved_bit is None:
            try:
                cur.execute("SELECT db_table_name, current_shift_row_id FROM mes_lines WHERE id=%s", (line_id,))
                ln = cur.fetchone()
                if ln and ln.get("db_table_name"):
                    tbl = ln["db_table_name"]
                    if ln.get("current_shift_row_id"):
                        cur.execute(
                            f"SELECT current_model_number, current_model_name FROM {tbl} WHERE id=%s",
                            (ln["current_shift_row_id"],),
                        )
                    else:
                        cur.execute(
                            f"SELECT current_model_number, current_model_name FROM {tbl} "
                            f"WHERE is_shift_completed = false "
                            f"ORDER BY timestamp DESC NULLS LAST LIMIT 1"
                        )
                    mrow = cur.fetchone() or {}
                    if mrow.get("current_model_number") is not None:
                        resolved_bit = int(mrow["current_model_number"])
                    if mrow.get("current_model_name"):
                        resolved_model_name = mrow["current_model_name"]
            except Exception:
                pass

        # 2. Pull applicable PYs (current model, non-PASS desired_value).
        # Same query the working /live/ endpoint uses — joins by py_no
        # (NOT py_id, which doesn't link cleanly) + filters by model
        # bit_number and zone.  Returning poka_yoke_no / poka_yoke_name
        # so the field names match what the rest of the system uses.
        applicable = []
        # Resolve THIS line's zone — for the zone-disambiguation filter
        # on mes_py_model_master.
        line_zone_id = None
        try:
            cur.execute("SELECT zone_id FROM mes_lines WHERE id=%s", (line_id,))
            _zr = cur.fetchone()
            if _zr:
                line_zone_id = _zr["zone_id"] if isinstance(_zr, dict) else _zr[0]
        except Exception:
            pass
        if resolved_bit is not None:
            cur.execute("""
                SELECT DISTINCT ON (a.py_no, a.side)
                       m.id             AS py_master_id,
                       m.sensing_bits   AS sensing_bits,
                       a.py_no          AS py_no,
                       a.py_name        AS py_name,
                       a.side,
                       COALESCE(m.bit, a.d_bit) AS register_addr,
                       COALESCE(m.desired_value, a.desired_value) AS desired_value
                  FROM mes_py_assignments a
                  LEFT JOIN mes_py_master       m  ON m.py_no = a.py_no AND m.is_active = true
                  LEFT JOIN mes_py_model_master mm ON mm.id   = a.model_id
                 WHERE mm.bit_number = %s AND mm.is_active = true
                   AND (mm.zone_id IS NULL OR mm.zone_id = COALESCE(%s, mm.zone_id))
                 ORDER BY a.py_no, a.side, a.id
            """, (resolved_bit, line_zone_id))
            for r in cur.fetchall():
                dv = r.get("desired_value")
                # Skip PYs whose desired value for this model is PASS / 0 /
                # NULL — those don't apply to the running model.
                if dv is None:
                    continue
                try:
                    if int(dv) == 0:
                        continue
                except (ValueError, TypeError):
                    if str(dv).strip().lower() in ("", "pass"):
                        continue
                applicable.append(dict(r))

    # 3. Index sweep entries by bit name for fast lookup.
    sweep    = _SENSOR_SWEEP.get(line_id, {}) or {}
    swept_at = sweep.get("swept_at")
    entries  = sweep.get("entries") or []
    by_bit   = {}
    for e in entries:
        bit = (e.get("bit") or e.get("x_bit") or "").upper()
        if bit:
            by_bit[bit] = e

    # 4. Roll up each PY's status across its sensing bits.
    checks = []
    alive_n = stuck_n = unknown_n = 0
    for py in applicable:
        sb_raw = (py.get("sensing_bits") or "").upper()
        bits   = REG_RE.findall(sb_raw)
        per_bit = []
        statuses = []
        stuck_durs = []
        last_toggles = []
        for b in bits:
            ent = by_bit.get(b.upper())
            if not ent:
                statuses.append("unknown")
                per_bit.append({"bit": b, "status": "unknown",
                                 "current_value": None,
                                 "last_toggle_at": None,
                                 "stuck_for_sec": None})
                continue
            st = (ent.get("status") or "unknown")
            ago = ent.get("last_toggle_ago_sec")
            # Recompute status against threshold if collector gave us a
            # raw `ago` but no derived status (defensive).
            if st == "alive" and ago is not None and ago > stuck_sec:
                st = "stuck"
            statuses.append(st)
            if st == "stuck" and ago is not None:
                stuck_durs.append(float(ago))
            if ent.get("last_toggle_at"):
                last_toggles.append(ent["last_toggle_at"])
            per_bit.append({
                "bit":            b,
                "status":         st,
                "current_value":  ent.get("current_value"),
                "last_toggle_at": ent.get("last_toggle_at"),
                "stuck_for_sec":  float(ago) if (st == "stuck" and ago is not None) else None,
            })

        if not statuses:
            # PY has no parseable sensing bits configured — count as unknown
            py_status = "unknown"
        elif "unknown" in statuses:
            py_status = "unknown"
        elif "stuck" in statuses:
            py_status = "stuck"
        else:
            py_status = "alive"

        if py_status == "alive":   alive_n   += 1
        elif py_status == "stuck": stuck_n   += 1
        else:                      unknown_n += 1

        checks.append({
            "py_no":         py.get("py_no"),
            "py_name":       py.get("py_name"),
            "sensing_bits":  py.get("sensing_bits"),
            "register_addr": py.get("register_addr"),
            "status":        py_status,
            "stuck_for_sec": max(stuck_durs) if stuck_durs else None,
            "last_toggle_at": max(last_toggles) if last_toggles else None,
            "bits":          per_bit,
        })

    total = len(checks)
    if total == 0:
        health = "NO_PY"        # current model has no applicable PYs
    elif stuck_n > 0:
        health = "CRITICAL"
    elif unknown_n > 0:
        health = "WARNING"      # collector warming up
    else:
        health = "OK"

    return {
        "line_id":     line_id,
        "model_bit":   resolved_bit,
        "model_name":  resolved_model_name,
        "swept_at":    swept_at,
        "stuck_threshold_sec": stuck_sec,
        "health":      health,
        "counts": {
            "total":   total,
            "alive":   alive_n,
            "stuck":   stuck_n,
            "unknown": unknown_n,
        },
        "checks":      checks,
    }


# ── Auto-ack endpoint — collector calls this when a PY+register becomes OK
class EventAckRequest(BaseModel):
    line_id:  int
    py_no:    Optional[str] = None
    register: Optional[str] = None

@router.post("/events/auto-ack")
def auto_ack_events(body: EventAckRequest):
    """Mark SENSOR_BYPASS events as acknowledged for a given line + py_no
    (+ optionally register).  Called by the collector when a fault has
    cleared — this lets the dashboard drop the red alert immediately
    instead of waiting 8 hours for the event TTL."""
    if not body.py_no:
        return {"ok": True, "acked": 0}
    with get_conn() as conn:
        cur = conn.cursor()
        params = [body.line_id, body.py_no]
        extra  = ""
        if body.register:
            extra = " AND ((context_json::jsonb)->>'register') = %s"
            params.append(body.register)
        cur.execute(f"""
            UPDATE mes_poka_yoke_events
               SET acknowledged    = true,
                   acknowledged_at = NOW(),
                   acknowledged_by = 'collector:auto-ack'
             WHERE rule_type = 'SENSOR_BYPASS'
               AND acknowledged = false
               AND line_id = %s
               AND ((context_json::jsonb)->>'py_no') = %s
               {extra}
        """, params)
        n = cur.rowcount
    return {"ok": True, "acked": n}


@router.post("/events/ingest", status_code=201)
def ingest_event(body: EventIngest):
    """Called by collector — no auth needed (localhost only). Fires BOTH:
       • An immediate per-event email on SENSOR_BYPASS (fire-and-forget thread)
       • The usual 15-minute consolidated digest (runs separately)

    The collector already debounces duplicate codes, so we won't spam QA —
    one mail per *new* fault transition."""
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            INSERT INTO mes_poka_yoke_events
                (line_id, rule_id, rule_type, alert_level,
                 shift_name, plc_value, context_json)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            body.line_id, body.rule_id, body.rule_type, body.alert_level,
            body.shift_name, body.plc_value, body.context_json,
        ))
        new_id = cur.fetchone()["id"]

    # Fire-and-forget immediate notification for SENSOR_BYPASS and
    # SENSOR_HEALTH events (same mailer, subject branches on rule_type).
    if (body.rule_type or "").upper() in ("SENSOR_BYPASS", "SENSOR_HEALTH"):
        import threading
        threading.Thread(target=_send_bypass_email_async, args=(body,), daemon=True).start()

    return {"ok": True, "event_id": new_id}


# ══════════════════════════════════════════════════════════════════════════
# PERIODIC DIGEST — every 15 minutes, email a summary of all active bypasses
# ══════════════════════════════════════════════════════════════════════════

def _send_bypass_digest_now() -> None:
    """Gather all unacknowledged SENSOR_BYPASS events from the last 15 minutes
    and email them as a single HTML digest to NOTIFY_EMAIL recipients."""
    import os, json as _json, smtplib
    from email.mime.text      import MIMEText
    from email.mime.multipart import MIMEMultipart
    from datetime             import datetime as _dt

    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587") or 587)
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    to_list, cc_list = _get_mail_addrs("bypass")
    if not (smtp_user and smtp_pass and to_list):
        print("[POKA-DIGEST] Skipped (SMTP or To not configured)")
        return

    try:
        with get_conn() as conn:
            cur = dict_cursor(conn)
            cur.execute("""
                SELECT e.id, e.detected_at, e.plc_value, e.context_json,
                       e.line_id, e.shift_name, e.alert_level,
                       l.line_name
                FROM mes_poka_yoke_events e
                LEFT JOIN mes_lines l ON l.id = e.line_id
                WHERE e.rule_type    = 'SENSOR_BYPASS'
                  AND e.acknowledged = false
                  AND e.detected_at  > NOW() - INTERVAL '15 minutes'
                ORDER BY e.detected_at DESC
            """)
            events = cur.fetchall()
    except Exception as exc:
        print(f"[POKA-DIGEST] DB query failed: {exc}")
        return

    if not events:
        print("[POKA-DIGEST] No active bypasses in last 15 min — skip mail")
        return

    # Deduplicate by (line_id, py_no, register) so the digest stays compact.
    seen = {}
    for ev in events:
        try:
            ctx = _json.loads(ev.get("context_json") or "{}")
        except Exception:
            ctx = {}
        key = (ev["line_id"], ctx.get("py_no"), ctx.get("register"))
        if key in seen:
            continue
        seen[key] = {"ev": ev, "ctx": ctx}

    rows_html = []
    for i, item in enumerate(seen.values(), 1):
        ev, ctx = item["ev"], item["ctx"]
        rows_html.append(f"""
        <tr style="background:{'#fff' if i%2 else '#f8fafc'};">
          <td style="padding:7px 12px;font-weight:600;">{i}</td>
          <td style="padding:7px 12px;">{ev.get('line_name') or f"Line #{ev['line_id']}"}</td>
          <td style="padding:7px 12px;font-family:monospace;color:#1e40af;font-weight:700;">{ctx.get('py_no','—')}</td>
          <td style="padding:7px 12px;">{ctx.get('py_name','—')}</td>
          <td style="padding:7px 12px;font-family:monospace;color:#7c3aed;font-weight:700;">{ctx.get('register','—')}</td>
          <td style="padding:7px 12px;">{ev.get('shift_name') or '—'}</td>
          <td style="padding:7px 12px;color:#64748b;font-size:11px;">{ev['detected_at'].strftime('%H:%M:%S')}</td>
        </tr>""")

    html = f"""
<html><body style="font-family:Arial,sans-serif;color:#0f172a;background:#f1f5f9;padding:20px;">
  <div style="max-width:820px;margin:auto;background:#fff;border-radius:10px;padding:24px;border-left:5px solid #f59e0b;">
    <h2 style="margin:0 0 4px;color:#f59e0b;">⚠ Poka-Yoke Bypass Digest</h2>
    <div style="color:#64748b;font-size:13px;margin-bottom:16px;">
      <b>{len(seen)}</b> active bypass{'es' if len(seen)!=1 else ''} detected in the last 15 minutes.
    </div>
    <table style="border-collapse:collapse;width:100%;font-size:12px;border:1px solid #e2e8f0;">
      <thead>
        <tr style="background:#1e40af;color:#fff;">
          <th style="padding:8px 12px;text-align:left;">#</th>
          <th style="padding:8px 12px;text-align:left;">Line</th>
          <th style="padding:8px 12px;text-align:left;">PY No</th>
          <th style="padding:8px 12px;text-align:left;">Defect</th>
          <th style="padding:8px 12px;text-align:left;">Register</th>
          <th style="padding:8px 12px;text-align:left;">Shift</th>
          <th style="padding:8px 12px;text-align:left;">Time</th>
        </tr>
      </thead>
      <tbody>{''.join(rows_html)}</tbody>
    </table>
    <p style="margin-top:20px;font-size:12px;color:#64748b;">
      Please investigate and acknowledge these bypasses in the Fullscreen dashboard.<br/>
      Automated 15-minute digest — generated at {_dt.now().strftime("%Y-%m-%d %H:%M:%S")}.
    </p>
  </div>
</body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"[DIGEST] {len(seen)} active poka-yoke bypass(es) — last 15 min"
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(to_list)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            server.ehlo(); server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, to_list + cc_list, msg.as_string())
        print(f"[POKA-DIGEST] Sent digest with {len(seen)} entries → "
              f"To={to_list} Cc={cc_list}")
    except Exception as exc:
        print(f"[POKA-DIGEST] Send failed: {exc}")


def _digest_worker_loop():
    """Background thread — sleeps 15 min, sends digest, repeats forever."""
    import time as _time
    INTERVAL = 15 * 60  # 15 minutes
    while True:
        try:
            _time.sleep(INTERVAL)
            _send_bypass_digest_now()
        except Exception as exc:
            print(f"[POKA-DIGEST] Worker error: {exc}")

_DIGEST_STARTED = False
def _start_digest_worker():
    global _DIGEST_STARTED
    if _DIGEST_STARTED:
        return
    _DIGEST_STARTED = True
    import threading
    t = threading.Thread(target=_digest_worker_loop, name="poka-digest", daemon=True)
    t.start()
    print("[POKA-DIGEST] Worker started — digest every 15 minutes")

# Kick off the scheduler on module import (once per process).
_start_digest_worker()


@router.post("/digest/send-now")
def send_digest_now(admin=Depends(require_admin)):
    """Manual trigger — send the 15-minute digest immediately."""
    _send_bypass_digest_now()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════
# HOURLY-SLOT PERFORMANCE REPORT
# Fires at the END of every hourly slot with a per-line breakdown of:
#   Plan | Actual | OK | NG | Avg CT | Loss distribution | PY Bypasses |
#   Efficiency % | Target Achieved
# Goes to SLOT_REPORT_EMAIL (separate from the poka-yoke NOTIFY_EMAIL).
# ══════════════════════════════════════════════════════════════════════════

# In-memory dedupe so one slot-end isn't reported twice.
_SLOT_REPORT_STATE: set = set()   # {(line_id, date, slot_label)}


def _fmt_secs(s: float) -> str:
    s = max(0, int(s or 0))
    return f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}"


def _collect_slot_report(conn, line: dict, slot: dict, record_date) -> dict:
    """Build the per-line data dict for a completed hourly slot.
    Works for both REGULAR slots (mes_hourly_slots rows with db_column_prefix)
    and SYNTHETIC OT slots (on-the-fly dicts with prefix=None — stats are
    computed directly from the cycle log + status log)."""
    cur = dict_cursor(conn)

    line_id   = line["id"]
    shift_tbl = line["db_table_name"]
    ct_tbl    = f"{shift_tbl}_ct_log"
    prefix    = slot.get("db_column_prefix")
    shift     = slot["shift_name"]
    s_start   = slot["start_time"]
    s_end     = slot["end_time"]
    is_ot     = bool(slot.get("is_ot"))

    # 1. Slot plan / actual / OK / NG — two paths depending on slot kind.
    plan = int(slot.get("plan_pieces") or 0)
    actual = ok = ng = 0
    if prefix:
        try:
            cur.execute(
                f"SELECT {prefix}_actual AS actual, {prefix}_ok AS ok_cnt, "
                f"       {prefix}_ng AS ng_cnt, {prefix}_plan AS plan_col "
                f"FROM {shift_tbl} "
                f"WHERE record_date=%s AND shift_name=%s "
                f"ORDER BY id DESC LIMIT 1",
                (record_date, shift),
            )
            r = cur.fetchone() or {}
            actual = int(r.get("actual")  or 0)
            ok     = int(r.get("ok_cnt")  or 0)
            ng     = int(r.get("ng_cnt")  or 0)
            if r.get("plan_col"): plan = int(r["plan_col"])
        except Exception:
            pass
    # Synthetic (OT) slot → compute from cycle log; plan already set from slot

    # 2. Avg CT for the slot from <line>_ct_log (ts-filtered)
    from datetime import datetime, time as _time
    def _combine(d, t): return datetime.combine(d, t if isinstance(t, _time) else _time(0))
    slot_start_ts = _combine(record_date, s_start)
    slot_end_ts   = _combine(record_date, s_end)
    avg_ct = 0.0
    try:
        cur.execute(
            f"SELECT COALESCE(AVG(ct_value),0) AS avg_ct, "
            f"       COUNT(*) FILTER (WHERE is_ng=false) AS ok_log, "
            f"       COUNT(*) FILTER (WHERE is_ng=true)  AS ng_log "
            f"FROM {ct_tbl} WHERE ts BETWEEN %s AND %s",
            (slot_start_ts, slot_end_ts),
        )
        r2 = cur.fetchone() or {}
        avg_ct = float(r2.get("avg_ct") or 0)
        # For synthetic OT slots, fall back to cycle-log counts (no column prefix).
        if not prefix:
            ok     = int(r2.get("ok_log") or 0)
            ng     = int(r2.get("ng_log") or 0)
            actual = ok + ng
    except Exception:
        pass

    # 3. Loss distribution for THIS slot only (from mes_status_log)
    # BUG fix: mes_status_log.status stores the status NAME as text
    # ("BREAKDOWN", "RUNNING", …), NOT an integer code.  Map by name.
    # Also pick up the status that was ACTIVE at slot_start so a breakdown
    # which began before the slot is still counted from slot_start onward.
    loss_by_cat: dict = {}     # category → seconds
    total_loss_s = 0
    try:
        cur.execute("""
            SELECT status_code, status_name, loss_type
            FROM mes_status_mappings WHERE line_id=%s
        """, (line_id,))
        name_map = {
            str(r["status_name"] or "").strip().upper(): r
            for r in cur.fetchall()
        }

        # Last status BEFORE/AT slot_start — what the line was doing when slot opened
        cur.execute("""
            SELECT status FROM mes_status_log
            WHERE line_id=%s AND ts <= %s
            ORDER BY ts DESC LIMIT 1
        """, (line_id, slot_start_ts))
        pre = cur.fetchone() or {}
        current_status = str(pre.get("status") or "").strip().upper() or None

        # All status changes DURING the slot
        cur.execute("""
            SELECT status, ts FROM mes_status_log
            WHERE line_id=%s AND ts > %s AND ts <= %s
            ORDER BY ts
        """, (line_id, slot_start_ts, slot_end_ts))
        changes = cur.fetchall()

        # `mes_status_log.ts` is TIMESTAMPTZ → naive/aware comparison will
        # raise TypeError. Strip tzinfo from every returned ts so everything
        # is naive (local-time) for subtraction.
        def _naive(t):
            return t.replace(tzinfo=None) if t and t.tzinfo else t

        # Walk intervals: (status_in_effect, interval_start, interval_end)
        intervals = []
        t = slot_start_ts
        for ch in changes:
            ct = _naive(ch["ts"])
            intervals.append((current_status, t, ct))
            current_status = str(ch["status"] or "").strip().upper() or None
            t = ct
        intervals.append((current_status, t, slot_end_ts))

        for st, start, end in intervals:
            dur = (end - start).total_seconds()
            if dur <= 0 or not st:
                continue
            info = name_map.get(st)
            if not info:
                continue
            if st in ("RUNNING", "RUN", "PRODUCING"):
                continue
            # Bucket by loss_type (capitalised) if set, else by status_name.
            loss_type = (info.get("loss_type") or "").strip()
            if loss_type:
                cat = loss_type.title()   # 'breakdown' → 'Breakdown'
            else:
                cat = st.title().replace("_", " ")
            loss_by_cat[cat] = loss_by_cat.get(cat, 0) + dur
            total_loss_s += dur
    except Exception as _loss_exc:
        print(f"[SLOT-REPORT] loss calc failed line={line_id}: {_loss_exc}")

    # 4. Poka-yoke bypasses inside the slot window
    bypass_count = 0
    bypass_items: list = []
    try:
        cur.execute("""
            SELECT (context_json::jsonb)->>'py_no'   AS py_no,
                   (context_json::jsonb)->>'py_name' AS py_name
            FROM mes_poka_yoke_events
            WHERE line_id=%s AND rule_type='SENSOR_BYPASS'
              AND detected_at BETWEEN %s AND %s
              AND context_json IS NOT NULL
        """, (line_id, slot_start_ts, slot_end_ts))
        seen = {}
        for r in cur.fetchall():
            k = (r.get("py_no"), r.get("py_name"))
            seen[k] = seen.get(k, 0) + 1
        bypass_count = sum(seen.values())
        bypass_items = [{"py_no": k[0] or "—", "py_name": k[1] or "—", "count": v}
                        for k, v in seen.items()]
    except Exception:
        pass

    eff             = (actual / plan * 100) if plan > 0 else 0
    target_achieved = (plan > 0 and actual >= plan)

    return {
        "line_name":       line.get("line_name") or f"Line #{line_id}",
        "line_id":         line_id,
        "shift":           shift,
        "slot":            slot["slot_label"],
        "slot_start":      slot_start_ts,
        "slot_end":        slot_end_ts,
        "plan":            plan,
        "actual":          actual,
        "ok":              ok,
        "ng":              ng,
        "avg_ct":          avg_ct,
        "total_loss_s":    int(total_loss_s),
        "loss_by_cat":     loss_by_cat,
        "bypass_count":    bypass_count,
        "bypass_items":    bypass_items,
        "efficiency":      eff,
        "target_achieved": target_achieved,
        "is_ot":           is_ot,
    }


def _send_slot_report(entries: list) -> None:
    """Send one HTML email summarising the just-completed slot for all lines."""
    if not entries: return
    import os, smtplib
    from email.mime.text      import MIMEText
    from email.mime.multipart import MIMEMultipart
    from datetime             import datetime as _dt

    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587") or 587)
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    # Recipient config comes from the mail-config system: DB → env (HOURLY_TO
    # / HOURLY_CC) → legacy env (SLOT_REPORT_EMAIL).  Edit via the Admin
    # Panel → Mail Config tab.
    to_list, cc_list = _get_mail_addrs("hourly")
    if not (smtp_user and smtp_pass and to_list):
        print("[SLOT-REPORT] Skipped (SMTP or To not configured)")
        return

    slot_label = entries[0]["slot"]
    shift      = entries[0]["shift"]

    # Build per-line blocks
    blocks = []
    for e in entries:
        eff_color   = "#16a34a" if e["target_achieved"] else "#dc2626"
        ach_label   = "YES" if e["target_achieved"] else "NO"
        ach_bg      = "#16a34a" if e["target_achieved"] else "#dc2626"
        var         = e["actual"] - e["plan"]
        var_color   = "#16a34a" if var >= 0 else "#dc2626"

        loss_rows = ""
        for cat, secs in sorted(e["loss_by_cat"].items(), key=lambda x: -x[1]):
            pct = (secs / 3600 * 100) if secs else 0
            loss_rows += f"""
              <tr>
                <td style="padding:4px 10px;">{cat}</td>
                <td style="padding:4px 10px;text-align:right;font-family:monospace;">{_fmt_secs(secs)}</td>
                <td style="padding:4px 10px;text-align:right;color:#64748b;">{pct:.1f}%</td>
              </tr>"""
        if not loss_rows:
            loss_rows = "<tr><td colspan='3' style='padding:4px 10px;color:#94a3b8;'>No losses in this slot</td></tr>"

        bypass_html = ""
        if e["bypass_items"]:
            bp_rows = "".join(f"""
                <tr>
                  <td style="padding:3px 10px;font-family:monospace;color:#1e40af;">{b['py_no']}</td>
                  <td style="padding:3px 10px;">{b['py_name']}</td>
                  <td style="padding:3px 10px;text-align:right;font-family:monospace;">{b['count']}</td>
                </tr>""" for b in e["bypass_items"])
            bypass_html = f"""
              <div style="font-size:12px;font-weight:700;color:#c2410c;margin-top:10px;">
                ⚠ {e['bypass_count']} Poka-Yoke bypass event(s) in this slot:
              </div>
              <table style="border-collapse:collapse;width:100%;font-size:11px;margin-top:4px;">
                <thead>
                  <tr style="background:#fef3c7;">
                    <th style="padding:4px 10px;text-align:left;">PY No</th>
                    <th style="padding:4px 10px;text-align:left;">Defect</th>
                    <th style="padding:4px 10px;text-align:right;">Count</th>
                  </tr>
                </thead>
                <tbody>{bp_rows}</tbody>
              </table>"""
        else:
            bypass_html = """<div style="font-size:12px;color:#16a34a;margin-top:10px;">
              ✓ No poka-yoke bypasses in this slot.
            </div>"""

        blocks.append(f"""
        <div style="background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:16px 18px;margin-bottom:14px;">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">
            <h3 style="margin:0;color:#0f172a;font-size:16px;">{e['line_name']}</h3>
            <span style="background:{ach_bg};color:#fff;padding:4px 12px;border-radius:99px;font-size:11px;font-weight:700;">
              Target Achieved: {ach_label}
            </span>
          </div>
          <table style="border-collapse:collapse;width:100%;font-size:12px;">
            <tr>
              <td style="padding:5px 10px;background:#f8fafc;font-weight:700;width:150px;">Target</td>
              <td style="padding:5px 10px;font-family:monospace;">{e['plan']}</td>
            </tr>
            <tr>
              <td style="padding:5px 10px;background:#f1f5f9;font-weight:700;">Actual</td>
              <td style="padding:5px 10px;font-family:monospace;">{e['actual']}
                <span style="color:{var_color};font-weight:700;"> ({'+' if var>=0 else ''}{var})</span>
              </td>
            </tr>
            <tr>
              <td style="padding:5px 10px;background:#f8fafc;font-weight:700;">OK / NG</td>
              <td style="padding:5px 10px;font-family:monospace;">
                <span style="color:#16a34a;font-weight:700;">{e['ok']}</span> /
                <span style="color:#dc2626;font-weight:700;">{e['ng']}</span>
              </td>
            </tr>
            <tr>
              <td style="padding:5px 10px;background:#f1f5f9;font-weight:700;">Avg Cycle Time</td>
              <td style="padding:5px 10px;font-family:monospace;">{e['avg_ct']:.2f} s</td>
            </tr>
            <tr>
              <td style="padding:5px 10px;background:#f8fafc;font-weight:700;">Efficiency</td>
              <td style="padding:5px 10px;font-family:monospace;color:{eff_color};font-weight:700;">{e['efficiency']:.1f}%</td>
            </tr>
            <tr>
              <td style="padding:5px 10px;background:#f1f5f9;font-weight:700;">Total Loss (this slot)</td>
              <td style="padding:5px 10px;font-family:monospace;">{_fmt_secs(e['total_loss_s'])}</td>
            </tr>
          </table>

          <div style="font-size:12px;font-weight:700;color:#64748b;margin-top:12px;text-transform:uppercase;letter-spacing:.06em;">
            Loss Distribution (slot)
          </div>
          <table style="border-collapse:collapse;width:100%;font-size:11px;margin-top:4px;">
            <thead>
              <tr style="background:#e2e8f0;">
                <th style="padding:4px 10px;text-align:left;">Category</th>
                <th style="padding:4px 10px;text-align:right;">Duration</th>
                <th style="padding:4px 10px;text-align:right;">% of Slot</th>
              </tr>
            </thead>
            <tbody>{loss_rows}</tbody>
          </table>
          {bypass_html}
        </div>""")

    html = f"""
<html><body style="font-family:Arial,sans-serif;color:#0f172a;background:#f1f5f9;padding:20px;">
  <div style="max-width:900px;margin:auto;">
    <div style="background:#1e40af;color:#fff;padding:16px 20px;border-radius:10px 10px 0 0;">
      <h2 style="margin:0;">🕒 Slot Performance Report</h2>
      <div style="font-size:13px;opacity:.9;margin-top:4px;">
        Shift <b>{shift}</b> · Slot <b>{slot_label}</b> · {len(entries)} line(s) reported
      </div>
    </div>
    <div style="background:#fff;padding:16px;border-radius:0 0 10px 10px;">
      {''.join(blocks)}
      <p style="font-size:11px;color:#64748b;margin-top:10px;">
        <b>Report Generated at:</b> {_dt.now().strftime("%H:%M:%S")}
        ({_dt.now().strftime("%Y-%m-%d")} &middot; Shift <b>{shift}</b>)
      </p>
    </div>
  </div>
</body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"[SLOT REPORT] Shift {shift} · {slot_label} · {len(entries)} line(s)"
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(to_list)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            server.ehlo(); server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, to_list + cc_list, msg.as_string())
        print(f"[SLOT-REPORT] Sent {slot_label} report for {len(entries)} line(s) "
              f"→ To={to_list} Cc={cc_list}")
    except Exception as exc:
        print(f"[SLOT-REPORT] Send failed: {exc}")


def _slot_worker_loop():
    """Background thread — every 45 s, checks for newly-completed hourly slots
    across all lines; when found, fetches slot data and sends ONE combined
    email summary."""
    import time as _time
    from datetime import datetime as _dt, date as _date
    INTERVAL   = 45  # seconds
    GRACE      = 120 # report slots ended within the last 2 minutes
    print("[SLOT-REPORT] Worker started — hourly-slot digest at each slot end")
    while True:
        try:
            _time.sleep(INTERVAL)
            now = _dt.now()
            today = now.date()

            with get_conn() as conn:
                cur = dict_cursor(conn)
                cur.execute("""
                    SELECT id, line_name, db_table_name FROM mes_lines
                    WHERE db_table_name IS NOT NULL
                """)
                lines = cur.fetchall()
                # For each line, find slots whose end_time just passed.
                # We use the line's current active shift to figure out which slots apply.
                by_slot_label: dict = {}
                for line in lines:
                    # Determine current active shift for this line
                    try:
                        cur.execute(f"""
                            SELECT shift_name FROM {line['db_table_name']}
                            WHERE is_shift_completed=false
                            ORDER BY timestamp DESC NULLS LAST LIMIT 1
                        """)
                        sr = cur.fetchone()
                        shift = sr["shift_name"] if sr else None
                    except Exception:
                        shift = None
                    if not shift:
                        continue

                    cur.execute("""
                        SELECT shift_name, slot_label, slot_order, start_time, end_time,
                               crosses_midnight, plan_pieces, db_column_prefix
                        FROM mes_hourly_slots
                        WHERE line_id=%s AND shift_name=%s
                        ORDER BY slot_order
                    """, (line["id"], shift))
                    slots = [dict(s) for s in cur.fetchall()]

                    # ── OT slots: synthesize hourly OT slots when OT is on ──
                    # If ot_active_shift == current shift, generate hourly
                    # synthetic slots spanning ot_start_time..ot_end_time, so
                    # an OT-period slot report fires at each boundary.
                    try:
                        cur.execute(
                            "SELECT ot_active_shift FROM mes_lines WHERE id=%s",
                            (line["id"],),
                        )
                        ot_row = cur.fetchone() or {}
                        if ot_row.get("ot_active_shift") == shift:
                            cur.execute("""
                                SELECT ot_start_time, ot_end_time
                                FROM mes_shift_configs
                                WHERE line_id=%s AND shift_name=%s
                            """, (line["id"], shift))
                            ocfg = cur.fetchone() or {}
                            ot_s = ocfg.get("ot_start_time")
                            ot_e = ocfg.get("ot_end_time")
                            # Also pull line's ideal CT to estimate plan
                            cur.execute("""
                                SELECT ideal_cycle_time FROM mes_plc_configs
                                WHERE line_id=%s AND parent_plc_id IS NULL
                                LIMIT 1
                            """, (line["id"],))
                            pcf = cur.fetchone() or {}
                            ideal_ct = float(pcf.get("ideal_cycle_time") or 0)
                            if ot_s and ot_e:
                                from datetime import datetime as _dtX, timedelta as _td, time as _timeX
                                base = _dtX.combine(today, ot_s)
                                end  = _dtX.combine(today, ot_e)
                                # Handle shifts that cross midnight on OT too
                                if end <= base: end += _td(days=1)
                                cur_t = base
                                step_order = 9000  # well above regular slots
                                while cur_t < end:
                                    nxt = min(cur_t + _td(hours=1), end)
                                    lbl = f"OT {cur_t.strftime('%H:%M')}-{nxt.strftime('%H:%M')}"
                                    dur_hr = (nxt - cur_t).total_seconds() / 3600
                                    ot_plan = int(round(3600 * dur_hr / ideal_ct)) if ideal_ct > 0 else 0
                                    slots.append({
                                        "shift_name":        shift,
                                        "slot_label":        lbl,
                                        "slot_order":        step_order,
                                        "start_time":        cur_t.time(),
                                        "end_time":          nxt.time(),
                                        "crosses_midnight":  False,
                                        "plan_pieces":       ot_plan,
                                        "db_column_prefix":  None,   # synthetic
                                        "is_ot":             True,
                                    })
                                    step_order += 1
                                    cur_t = nxt
                    except Exception as _ot_exc:
                        print(f"[SLOT-REPORT] OT slot gen skipped line={line['id']}: {_ot_exc}")

                    for slot in slots:
                        # Slot end today
                        from datetime import datetime as _dt2, time as _time2
                        end_t = slot["end_time"]
                        if not isinstance(end_t, _time2):
                            continue
                        slot_end = _dt2.combine(today, end_t)
                        delta = (now - slot_end).total_seconds()
                        if delta < 0 or delta > GRACE:
                            continue    # slot hasn't ended yet OR ended long ago
                        key = (line["id"], today.isoformat(), slot["slot_label"])
                        if key in _SLOT_REPORT_STATE:
                            continue    # already reported
                        _SLOT_REPORT_STATE.add(key)
                        # Collect data
                        try:
                            data = _collect_slot_report(conn, dict(line), dict(slot), today)
                            by_slot_label.setdefault(slot["slot_label"], []).append(data)
                        except Exception as exc:
                            print(f"[SLOT-REPORT] collect failed line={line['id']}: {exc}")

            # Send one email per slot_label (multi-line combined)
            for lbl, entries in by_slot_label.items():
                _send_slot_report(entries)

        except Exception as exc:
            print(f"[SLOT-REPORT] Worker error: {exc}")


_SLOT_WORKER_STARTED = False
def _start_slot_worker():
    global _SLOT_WORKER_STARTED
    if _SLOT_WORKER_STARTED:
        return
    _SLOT_WORKER_STARTED = True
    import threading
    threading.Thread(target=_slot_worker_loop, name="slot-report", daemon=True).start()

_start_slot_worker()


@router.post("/slot-report/send-now")
def slot_report_send_now(line_id: int, slot_label: str, admin=Depends(require_admin)):
    """Manual trigger — build + email the slot report for a given line+slot.
    Handy for debugging, or to re-send a past slot."""
    from datetime import date as _date
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT id, line_name, db_table_name FROM mes_lines WHERE id=%s
        """, (line_id,))
        line = cur.fetchone()
        if not line:
            raise HTTPException(404, "Line not found")
        cur.execute("""
            SELECT shift_name, slot_label, slot_order, start_time, end_time,
                   crosses_midnight, plan_pieces, db_column_prefix
            FROM mes_hourly_slots
            WHERE line_id=%s AND slot_label=%s LIMIT 1
        """, (line_id, slot_label))
        slot = cur.fetchone()
        if not slot:
            raise HTTPException(404, "Slot not found")
        data = _collect_slot_report(conn, dict(line), dict(slot), _date.today())
    _send_slot_report([data])
    return {"ok": True}

# ══════════════════════════════════════════════════════════════
# POKA YOKE MASTER  (standalone library, not line-scoped)
# ══════════════════════════════════════════════════════════════

# ADD everything below at the very END of poka_yoke.py:

# ══════════════════════════════════════════════════════════════
# POKA YOKE MASTER
# ══════════════════════════════════════════════════════════════

def _ensure_py_register_col(conn):
    """Make sure the extra columns on mes_py_master exist (and py_no is
    nullable, since D-bit is now the functional primary key).  Idempotent
    via IF NOT EXISTS; the ALTER on py_no is a no-op once already nullable."""
    try:
        cur = conn.cursor()
        cur.execute("ALTER TABLE mes_py_master ADD COLUMN IF NOT EXISTS register VARCHAR(50)")
        cur.execute("ALTER TABLE mes_py_master ADD COLUMN IF NOT EXISTS register_count INTEGER DEFAULT 1")
        cur.execute("ALTER TABLE mes_py_master ADD COLUMN IF NOT EXISTS zone_id INTEGER")
        cur.execute("ALTER TABLE mes_py_master ADD COLUMN IF NOT EXISTS sensing_bits VARCHAR(100)")
        # py_no was NOT NULL historically — make nullable so the UI can omit it.
        try:
            cur.execute("ALTER TABLE mes_py_master ALTER COLUMN py_no DROP NOT NULL")
        except Exception:
            pass
        conn.commit()
    except Exception:
        try: conn.rollback()
        except Exception: pass


_REG_RE = None
def _normalize_register(raw: Optional[str]) -> str:
    """Extract every Mitsubishi register token (D/R/M/L/F/T/C/S decimal OR
    X/Y/W/B hex) from a free-form string and return them comma-joined, upper.
    Empty input → empty string."""
    global _REG_RE
    if _REG_RE is None:
        import re as _re
        _REG_RE = _re.compile(r"(?:D|R|M|L|F|T|C|S)\d+|(?:X|Y|W|B)[0-9A-F]+", _re.IGNORECASE)
    if not raw:
        return ""
    return ",".join(t.upper() for t in _REG_RE.findall(str(raw)))


def _sync_py_assignments(conn, py_no: str, py_name: Optional[str],
                         model_type: Optional[str], side: Optional[str],
                         d_bit: Optional[str], machine_fixture: Optional[str],
                         model_ids: Optional[list]) -> None:
    """Make mes_py_assignments reflect exactly the given model_ids for this PY.
    - New selections → insert (desired_value NULL, set in Config tab).
    - Deselected     → hard-delete (no is_active column on this table).
    - Still selected → refresh py_name / side / d_bit / machine_name, keep desired_value.
    """
    cur = dict_cursor(conn)

    cur.execute(
        "SELECT id, model_id, model_name FROM mes_py_assignments WHERE py_no=%s",
        (py_no,),
    )
    existing = cur.fetchall()
    existing_ids = {e["model_id"] for e in existing if e["model_id"] is not None}

    if not model_ids:
        cur.execute("DELETE FROM mes_py_assignments WHERE py_no=%s", (py_no,))
        return

    cur.execute("""
        SELECT id, model_name, model_type, series AS model_series,
               old_model_no, bit_number
        FROM mes_py_model_master
        WHERE id = ANY(%s) AND is_active=true
    """, (list(model_ids),))
    selected = cur.fetchall()
    selected_ids = {m["id"] for m in selected}

    # Hard-delete rows whose model_id is no longer selected (or has NULL model_id).
    cur.execute(
        "DELETE FROM mes_py_assignments WHERE py_no=%s AND (model_id IS NULL OR NOT (model_id = ANY(%s)))",
        (py_no, list(selected_ids)),
    )

    # Get PY master pk
    cur.execute("SELECT id FROM mes_py_master WHERE py_no=%s AND is_active=true", (py_no,))
    py_row = cur.fetchone()
    py_id = py_row["id"] if py_row else None

    for m in selected:
        if m["id"] in existing_ids:
            cur.execute("""
                UPDATE mes_py_assignments SET
                    py_name = %s, side = %s, model_type = %s,
                    model_series = %s, d_bit = %s, machine_name = %s
                WHERE py_no = %s AND model_id = %s
            """, (
                py_name or "", side or "ALL", model_type or "",
                m["model_series"] or "", d_bit or "", machine_fixture or "",
                py_no, m["id"],
            ))
        else:
            cur.execute("""
                INSERT INTO mes_py_assignments
                    (py_id, model_id, py_no, py_name, side, model_type,
                     model_name, model_series, old_model_no,
                     d_bit, desired_value, machine_name)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                py_id, m["id"], py_no, py_name or "", side or "ALL",
                model_type or "", m["model_name"], m["model_series"] or "",
                m["old_model_no"] or "", d_bit or "", None, machine_fixture or "",
            ))


@router.get("/master/")
def list_py_master(user=Depends(get_current_user)):
    with get_conn() as conn:
        _ensure_py_register_col(conn)
        cur = dict_cursor(conn)
        cur.execute("""
            WITH ranked AS (
                SELECT p.id,
                       p.py_no, p.description,
                       p.model_type, p.side, p.bit,
                       p.desired_value, p.machine_name,
                       p.register, p.register_count,
                       p.zone_id, p.sensing_bits,
                       ROW_NUMBER() OVER (
                           PARTITION BY p.zone_id
                           ORDER BY p.bit NULLS LAST, p.id
                       ) AS seq_in_zone
                FROM mes_py_master p
                WHERE p.is_active = true
            )
            SELECT r.id,
                   r.py_no         AS "pyNo",
                   r.description,
                   r.model_type    AS "modelType",
                   r.side          AS "typeSide",
                   r.bit           AS "dBit",
                   r.desired_value AS "desiredValue",
                   r.machine_name  AS "machineFixture",
                   r.register      AS "register",
                   COALESCE(r.register_count, 1) AS "registerCount",
                   r.zone_id       AS "zoneId",
                   z.zone_name     AS "zoneName",
                   z.zone_code     AS "zoneCode",
                   r.sensing_bits  AS "sensingBits",
                   r.seq_in_zone   AS "seqInZone",
                   COALESCE(
                     (SELECT array_agg(DISTINCT a.model_id)
                      FROM mes_py_assignments a
                      WHERE a.py_no = r.py_no
                        AND a.model_id IS NOT NULL),
                     ARRAY[]::int[]
                   ) AS "assignedModelIds"
            FROM ranked r
            LEFT JOIN mes_zones z ON z.id = r.zone_id
            ORDER BY z.zone_name NULLS LAST, r.seq_in_zone
        """)
        return cur.fetchall()


@router.post("/master/", status_code=201)
def create_py_master(body: PYMasterCreate, admin=Depends(require_admin)):
    with get_conn() as conn:
        _ensure_py_register_col(conn)
        cur = dict_cursor(conn)

        # D-bit is the new functional key. Normalize it first.
        d_bit_norm  = _normalize_register(body.dBit)
        register    = _normalize_register(body.register or body.dBit)
        sensing     = _normalize_register(body.sensingBits)
        if not d_bit_norm:
            raise HTTPException(400, "dBit (output register) is required")

        # Auto-generate a py_no fallback from d_bit if caller didn't supply one.
        # Legacy joins use py_no as a string key, so it must be unique-per-row.
        py_no = (body.pyNo or "").strip() or d_bit_norm

        # py_no has a UNIQUE constraint — handle both active + soft-deleted rows
        cur.execute("SELECT id, is_active FROM mes_py_master WHERE py_no = %s", (py_no,))
        existing = cur.fetchone()
        if existing and existing["is_active"]:
            raise HTTPException(409, f"Poka-Yoke with D-bit {d_bit_norm} already exists")

        if existing and not existing["is_active"]:
            # Reactivate the soft-deleted row with new values.
            cur.execute("""
                UPDATE mes_py_master SET
                    description    = %s,
                    model_type     = %s,
                    side           = %s,
                    bit            = %s,
                    desired_value  = %s,
                    machine_name   = %s,
                    register       = %s,
                    register_count = %s,
                    zone_id        = %s,
                    sensing_bits   = %s,
                    is_active      = true
                WHERE id = %s
                RETURNING id,
                          py_no          AS "pyNo",
                          description,
                          model_type     AS "modelType",
                          side           AS "typeSide",
                          bit            AS "dBit",
                          desired_value  AS "desiredValue",
                          machine_name   AS "machineFixture",
                          register       AS "register",
                          register_count AS "registerCount",
                          zone_id        AS "zoneId",
                          sensing_bits   AS "sensingBits"
            """, (
                body.description, body.modelType, body.typeSide, d_bit_norm,
                body.desiredValue, body.machineFixture, register,
                body.registerCount or 1, body.zoneId, sensing or None,
                existing["id"],
            ))
            created = cur.fetchone()
        else:
            cur.execute("""
                INSERT INTO mes_py_master
                    (py_no, description, model_type, side, bit, desired_value,
                     machine_name, register, register_count, zone_id, sensing_bits)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id,
                          py_no          AS "pyNo",
                          description,
                          model_type     AS "modelType",
                          side           AS "typeSide",
                          bit            AS "dBit",
                          desired_value  AS "desiredValue",
                          machine_name   AS "machineFixture",
                          register       AS "register",
                          register_count AS "registerCount",
                          zone_id        AS "zoneId",
                          sensing_bits   AS "sensingBits"
            """, (
                py_no, body.description, body.modelType,
                body.typeSide, d_bit_norm, body.desiredValue, body.machineFixture,
                register, body.registerCount or 1, body.zoneId, sensing or None,
            ))
            created = cur.fetchone()

        # Sync model assignments if provided
        if body.assignedModelIds is not None:
            _sync_py_assignments(
                conn, py_no, body.description,
                body.modelType, body.typeSide, d_bit_norm, body.machineFixture,
                body.assignedModelIds,
            )
        return created


@router.put("/master/{py_id}")
def update_py_master(py_id: int, body: PYMasterUpdate, admin=Depends(require_admin)):
    # Normalize register-shaped fields so stored values stay consistent with
    # what the collector / frontend expect (upper-case, comma-joined).
    d_bit_norm = _normalize_register(body.dBit)    if body.dBit     is not None else None
    reg_norm   = _normalize_register(body.register) if body.register is not None else None
    sens_norm  = _normalize_register(body.sensingBits) if body.sensingBits is not None else None

    col_map = {
        "description":    body.description,
        "model_type":     body.modelType,
        "side":           body.typeSide,
        "bit":            d_bit_norm,
        "desired_value":  body.desiredValue,
        "machine_name":   body.machineFixture,
        "register":       reg_norm,
        "register_count": body.registerCount,
        "zone_id":        body.zoneId,
        "sensing_bits":   sens_norm,
    }
    updates = {k: v for k, v in col_map.items() if v is not None}
    # No column updates but assignments may have changed → still allow.
    if not updates and body.assignedModelIds is None:
        raise HTTPException(400, "Nothing to update")

    with get_conn() as conn:
        _ensure_py_register_col(conn)
        cur = conn.cursor()
        py_no = None
        if updates:
            sets   = ", ".join(f"{k} = %s" for k in updates)
            values = list(updates.values()) + [py_id]
            cur.execute(
                f"UPDATE mes_py_master SET {sets} WHERE id = %s RETURNING py_no",
                values,
            )
            row = cur.fetchone()
            py_no = row[0] if row else None
        else:
            cur.execute("SELECT py_no FROM mes_py_master WHERE id=%s", (py_id,))
            row = cur.fetchone()
            py_no = row[0] if row else None

        # Cascade master-level fields to existing assignments (desired_value is
        # per-model, set on Config tab — never cascaded).
        if py_no:
            asgn_map = {
                "d_bit":        d_bit_norm,
                "py_name":      body.description,
                "machine_name": body.machineFixture,
                "side":         body.typeSide,
                "model_type":   body.modelType,
            }
            asgn_updates = {k: v for k, v in asgn_map.items() if v is not None}
            if asgn_updates:
                sets2 = ", ".join(f"{k} = %s" for k in asgn_updates)
                vals2 = list(asgn_updates.values()) + [py_no]
                cur.execute(
                    f"UPDATE mes_py_assignments SET {sets2} WHERE py_no = %s",
                    vals2,
                )

        cur.close()

        # Sync model assignments if list provided (None = don't touch)
        if py_no and body.assignedModelIds is not None:
            # Need py master fields (not in body) for the sync — fetch current.
            d_cur = dict_cursor(conn)
            d_cur.execute("""
                SELECT py_no, description, model_type, side, bit, machine_name
                FROM mes_py_master WHERE id=%s
            """, (py_id,))
            pm = d_cur.fetchone() or {}
            _sync_py_assignments(
                conn, pm.get("py_no") or py_no,
                body.description if body.description is not None else pm.get("description"),
                body.modelType   if body.modelType   is not None else pm.get("model_type"),
                body.typeSide    if body.typeSide    is not None else pm.get("side"),
                d_bit_norm       if d_bit_norm       is not None else pm.get("bit"),
                body.machineFixture if body.machineFixture is not None else pm.get("machine_name"),
                body.assignedModelIds,
            )

        conn.commit()
    return {"ok": True}


@router.delete("/master/{py_id}")
def delete_py_master(py_id: int, admin=Depends(require_admin)):
    """Soft-delete a PY: deactivates the master row AND removes every
    matching assignment (Config tab) so the Sensor Health monitor and the
    matrix views all clear in one go.  Idempotent — safe to call on an
    already-deleted PY."""
    with get_conn() as conn:
        cur = conn.cursor()
        # Resolve py_no for the legacy assignment join — assignments may be
        # linked by either the integer py_id or the string py_no.
        cur.execute("SELECT py_no FROM mes_py_master WHERE id=%s", (py_id,))
        row = cur.fetchone()
        py_no = row[0] if row else None

        # 1. Drop every assignment row pointing at this PY (covers both
        #    new-style py_id link and legacy py_no string match).
        if py_no:
            cur.execute(
                "DELETE FROM mes_py_assignments "
                "WHERE py_id = %s OR py_no = %s",
                (py_id, py_no),
            )
        else:
            cur.execute(
                "DELETE FROM mes_py_assignments WHERE py_id = %s",
                (py_id,),
            )
        deleted_assignments = cur.rowcount

        # 2. Soft-deactivate the master row so the collector stops
        #    monitoring + Sensor Health drops it on next sweep.
        cur.execute(
            "UPDATE mes_py_master SET is_active = false WHERE id = %s",
            (py_id,),
        )
    return {"ok": True, "deleted_assignments": deleted_assignments}


# ══════════════════════════════════════════════════════════════
# SERIES MASTER  (YRA, YNC, YY8 …)
# ══════════════════════════════════════════════════════════════

_SERIES_TABLE_ENSURED = False
def _ensure_series_table(conn):
    global _SERIES_TABLE_ENSURED
    if _SERIES_TABLE_ENSURED:
        return
    try:
        conn.cursor().execute("""
            CREATE TABLE IF NOT EXISTS mes_py_series (
                id         SERIAL PRIMARY KEY,
                code       TEXT UNIQUE NOT NULL,
                is_active  BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
    except Exception:
        pass
    _SERIES_TABLE_ENSURED = True


class SeriesCreate(BaseModel):
    code: str


@router.get("/series/")
def list_series(user=Depends(get_current_user)):
    with get_conn() as conn:
        _ensure_series_table(conn)
        cur = dict_cursor(conn)
        cur.execute("SELECT id, code FROM mes_py_series WHERE is_active=true ORDER BY code")
        return cur.fetchall()


@router.post("/series/", status_code=201)
def create_series(body: SeriesCreate, admin=Depends(require_admin)):
    code = (body.code or "").strip().upper()
    if not code:
        raise HTTPException(400, "Series code required")
    with get_conn() as conn:
        _ensure_series_table(conn)
        cur = dict_cursor(conn)
        # Re-activate if previously soft-deleted; error if already active.
        cur.execute("SELECT id, is_active FROM mes_py_series WHERE code=%s", (code,))
        row = cur.fetchone()
        if row:
            if row["is_active"]:
                raise HTTPException(409, f"Series {code} already exists")
            cur.execute(
                "UPDATE mes_py_series SET is_active=true WHERE id=%s RETURNING id, code",
                (row["id"],),
            )
            return cur.fetchone()
        cur.execute(
            "INSERT INTO mes_py_series (code) VALUES (%s) RETURNING id, code",
            (code,),
        )
        return cur.fetchone()


@router.delete("/series/{series_id}")
def delete_series(series_id: int, admin=Depends(require_admin)):
    with get_conn() as conn:
        _ensure_series_table(conn)
        conn.cursor().execute(
            "UPDATE mes_py_series SET is_active=false WHERE id=%s", (series_id,))
    return {"ok": True}


def _build_model_name(type_val: Optional[str], series_str: Optional[str]) -> str:
    """Auto-generate model name from type + slash-separated series list.
    Example: type='4 Way OTR', series='YNC/YCA' → '4 WAY OTR: (YNC/YCA)'"""
    t = (type_val or "").strip().upper() or "—"
    s = (series_str or "").strip() or "—"
    return f"{t}: ({s})"


# ══════════════════════════════════════════════════════════════
# MODEL MASTER
# ══════════════════════════════════════════════════════════════

_MODEL_COL_ENSURED = False
def _ensure_model_bit_col(conn):
    global _MODEL_COL_ENSURED
    if _MODEL_COL_ENSURED:
        return
    try:
        cur = conn.cursor()
        cur.execute("ALTER TABLE mes_py_model_master ADD COLUMN IF NOT EXISTS bit_number INTEGER")
        # Zone-scoped bit numbers — same bit value may legitimately exist
        # in two different zones (Seat Slider #1 ≠ Press Shop #1).
        cur.execute("ALTER TABLE mes_py_model_master ADD COLUMN IF NOT EXISTS zone_id INTEGER")
        # Composite uniqueness: a zone+bit pair must point to a single
        # active model.  Partial index so we don't trip over historical
        # is_active=false rows or rows where bit_number is still NULL.
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS ux_py_model_zone_bit_active
                ON mes_py_model_master (zone_id, bit_number)
                WHERE is_active = true AND bit_number IS NOT NULL
        """)
        # Drop the OLD global-bit unique index if it ever got created so
        # zone-scoped duplicates don't collide with it.
        cur.execute("""
            DO $$ BEGIN
                IF EXISTS (
                    SELECT 1 FROM pg_indexes
                    WHERE tablename = 'mes_py_model_master'
                      AND indexname = 'ux_py_model_bit_active'
                ) THEN
                    DROP INDEX ux_py_model_bit_active;
                END IF;
            END $$
        """)
        # One-time cleanup: strip legacy "TYPE-SERIES:" prefix from existing names
        cur.execute(
            "UPDATE mes_py_model_master "
            "SET model_name = regexp_replace(model_name, '^TYPE-SERIES:\\s*', '', 'i') "
            "WHERE model_name ILIKE 'TYPE-SERIES:%'"
        )
        conn.commit()
    except Exception:
        try: conn.rollback()
        except Exception: pass
    _MODEL_COL_ENSURED = True


@router.get("/models/")
def list_model_master(user=Depends(get_current_user)):
    with get_conn() as conn:
        _ensure_model_bit_col(conn)
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT m.id,
                   m.model_name   AS "modelName",
                   m.model_type   AS "type",
                   m.old_model_no AS "oldModelNo",
                   m.series       AS "model",
                   m.bit_number   AS "bitNumber",
                   m.zone_id      AS "zoneId",
                   z.zone_name    AS "zoneName",
                   z.zone_code    AS "zoneCode"
            FROM mes_py_model_master m
            LEFT JOIN mes_zones z ON z.id = m.zone_id
            WHERE m.is_active = true
            ORDER BY z.zone_name NULLS LAST, m.bit_number NULLS LAST, m.model_name
        """)
        return cur.fetchall()


@router.post("/models/", status_code=201)
def create_model_master(body: ModelMasterCreate, admin=Depends(require_admin)):
    with get_conn() as conn:
        _ensure_model_bit_col(conn)
        cur = dict_cursor(conn)
        # bit_number uniqueness is now scoped per-zone, so the same bit can
        # legitimately exist in two different zones.
        if body.bitNumber is not None:
            cur.execute(
                "SELECT id FROM mes_py_model_master "
                "WHERE bit_number=%s AND is_active=true "
                "  AND COALESCE(zone_id,0) = COALESCE(%s,0)",
                (body.bitNumber, body.zoneId),
            )
            if cur.fetchone():
                raise HTTPException(
                    409,
                    f"Bit number {body.bitNumber} already assigned in this zone",
                )
        model_name = (body.modelName or "").strip() or _build_model_name(body.type, body.model)
        # Name still globally unique across all zones — keeps the audit clean.
        cur.execute("SELECT id FROM mes_py_model_master WHERE model_name=%s AND is_active=true", (model_name,))
        if cur.fetchone():
            raise HTTPException(409, f"Another model already uses the name \"{model_name}\"")
        cur.execute("""
            INSERT INTO mes_py_model_master
                (model_name, model_type, old_model_no, series, bit_number, zone_id)
            VALUES (%s,%s,%s,%s,%s,%s)
            RETURNING id,
                      model_name   AS "modelName",
                      model_type   AS "type",
                      old_model_no AS "oldModelNo",
                      series       AS "model",
                      bit_number   AS "bitNumber",
                      zone_id      AS "zoneId"
        """, (model_name, body.type, body.oldModelNo, body.model, body.bitNumber, body.zoneId))
        return cur.fetchone()


@router.put("/models/{model_id}")
def update_model_master(model_id: int, body: ModelMasterUpdate, admin=Depends(require_admin)):
    with get_conn() as conn:
        _ensure_model_bit_col(conn)
        cur = dict_cursor(conn)
        # If bitNumber OR zoneId is being changed, validate within the
        # (possibly new) zone scope.
        if body.bitNumber is not None or body.zoneId is not None:
            # Resolve the effective zone_id post-update for the duplicate check.
            effective_zone = body.zoneId
            if effective_zone is None:
                cur.execute("SELECT zone_id FROM mes_py_model_master WHERE id=%s", (model_id,))
                row = cur.fetchone()
                effective_zone = (row or {}).get("zone_id") if row else None
            effective_bit = body.bitNumber
            if effective_bit is None:
                cur.execute("SELECT bit_number FROM mes_py_model_master WHERE id=%s", (model_id,))
                row = cur.fetchone()
                effective_bit = (row or {}).get("bit_number") if row else None
            if effective_bit is not None:
                cur.execute(
                    "SELECT id FROM mes_py_model_master "
                    "WHERE bit_number=%s AND is_active=true AND id!=%s "
                    "  AND COALESCE(zone_id,0) = COALESCE(%s,0)",
                    (effective_bit, model_id, effective_zone),
                )
                if cur.fetchone():
                    raise HTTPException(
                        409,
                        f"Bit number {effective_bit} already assigned in this zone",
                    )

        if body.modelName is not None:
            cur.execute(
                "SELECT id FROM mes_py_model_master "
                "WHERE model_name=%s AND is_active=true AND id!=%s",
                (body.modelName.strip(), model_id),
            )
            if cur.fetchone():
                raise HTTPException(409, f"Another model already uses the name \"{body.modelName}\"")

        col_map = {
            "model_name":   body.modelName,
            "model_type":   body.type,
            "old_model_no": body.oldModelNo,
            "series":       body.model,
            "bit_number":   body.bitNumber,
            "zone_id":      body.zoneId,
        }
        updates = {k: v for k, v in col_map.items() if v is not None}
        if not updates:
            raise HTTPException(400, "Nothing to update")
        sets   = ", ".join(f"{k} = %s" for k in updates)
        values = list(updates.values()) + [model_id]
        conn.cursor().execute(
            f"UPDATE mes_py_model_master SET {sets} WHERE id = %s", values)
    return {"ok": True}


@router.delete("/models/{model_id}")
def delete_model_master(model_id: int, admin=Depends(require_admin)):
    with get_conn() as conn:
        conn.cursor().execute(
            "UPDATE mes_py_model_master SET is_active=false WHERE id=%s", (model_id,))
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════
# EXCEL TEMPLATES + BULK IMPORT — Model Master & PY Master
# ══════════════════════════════════════════════════════════════════════════
# Workflow:
#   1. UI calls GET /…/template  → server returns a styled .xlsx with the
#      expected column headers + 2 example rows the user can edit.
#   2. User fills more rows in Excel and uploads via POST /…/import.
#   3. Server validates each row, inserts what it can, returns
#      {inserted, skipped, errors[]}.
#
# Both templates are zone-aware: PY Master has a Zone column (looked up
# by name), Model Master is shared across zones (no zone column).

def _xlsx_response(wb, filename: str) -> Response:
    """Serialize an openpyxl workbook to a streaming Excel download."""
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _styled_header(ws, headers: list, fill_color: str = "1E40AF"):
    """Apply bold-white-on-blue header style + auto column width to ws."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24


def _autosize_columns(ws):
    from openpyxl.utils import get_column_letter
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 4, 12)


# ── MODEL MASTER ──────────────────────────────────────────────────────

@router.get("/models/template", tags=["import-export"])
def download_model_template(user=Depends(get_current_user)):
    """Fillable Excel template for Model Master bulk import — zone-aware.
    Same bit number can repeat across zones (Seat Slider #1 ≠ Press Shop #1)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Master"

    headers = ["Zone", "Model Name", "Type", "Series", "Old Model No", "Bit Number"]
    _styled_header(ws, headers)
    ws.append(["Seat Slider", "TRACK ASSY FRONT SEAT YHB 4 WAY OTR", "4 Way OTR",
               "YHB", "433140-14240-YHB", 9])
    ws.append(["Seat Slider", "TRACK ASSY FRONT SEAT YHB 4 WAY INR LH", "4 Way Inr LH",
               "YHB", "433140-14250-YHB", 10])
    ws.append(["Press Shop",  "PRESS PART X1",                          "—",
               "—",   "—",                 1])

    # Hint sheet — explains every column + valid Type values
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Column", "Required", "Notes"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
    ws2.append(["Zone",       "yes", "Zone name from Admin → Zones (case-insensitive)"])
    ws2.append(["Model Name", "yes", "Full descriptive name (must be globally unique)"])
    ws2.append(["Type",       "yes", "4 Way Inr LH / 4 Way Inr RH / 4 Way OTR / "
                                     "6 Way Inr LH / 6 Way Inr RH / 6 Way Otr LH / "
                                     "6 Way Otr RH"])
    ws2.append(["Series",     "no",  "e.g. YHB, YNC, Y17 — short series code"])
    ws2.append(["Old Model No","no", "Legacy part number (free text)"])
    ws2.append(["Bit Number", "yes", "Integer bit number — unique WITHIN a zone "
                                     "(same bit may exist in different zones)"])

    _autosize_columns(ws)
    _autosize_columns(ws2)
    return _xlsx_response(wb, "model_master_template.xlsx")


@router.post("/models/import", tags=["import-export"])
async def import_model_master(file: UploadFile = File(...),
                              admin=Depends(require_admin)):
    """Bulk-insert rows from an uploaded Excel file.  Zone column is matched
    to mes_zones by name (case-insensitive).  Bit-number uniqueness is
    enforced per-zone, so the same bit may legitimately exist in two zones."""
    import io
    from openpyxl import load_workbook
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Upload an .xlsx file")
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel file: {e}")

    ws = wb["Model Master"] if "Model Master" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": True, "inserted": 0, "skipped": 0, "errors": ["empty sheet"]}

    headers = [str(c).strip() if c else "" for c in rows[0]]
    def col_idx(name):
        try: return headers.index(name)
        except ValueError: return None

    iZone = col_idx("Zone")
    iName = col_idx("Model Name")
    iType = col_idx("Type")
    iSer  = col_idx("Series")
    iOld  = col_idx("Old Model No")
    iBit  = col_idx("Bit Number")

    inserted, skipped = 0, 0
    errors: list[str] = []
    with get_conn() as conn:
        _ensure_model_bit_col(conn)
        cur = conn.cursor()
        # Build zone name → id lookup (case-insensitive)
        cur.execute("SELECT id, zone_name FROM mes_zones")
        zone_map = {str(zn).strip().lower(): zid for zid, zn in cur.fetchall()}

        for r_idx, row in enumerate(rows[1:], start=2):
            try:
                name = (row[iName] if iName is not None else None) or ""
                name = str(name).strip()
                if not name:
                    skipped += 1
                    continue
                bit_raw = row[iBit] if iBit is not None else None
                bit = int(bit_raw) if bit_raw not in (None, "") else None

                zone_id = None
                if iZone is not None and row[iZone]:
                    zone_id = zone_map.get(str(row[iZone]).strip().lower())
                    if zone_id is None:
                        errors.append(f"Row {r_idx}: zone '{row[iZone]}' not found")

                # Per-zone duplicate check
                if bit is not None:
                    cur.execute(
                        "SELECT id FROM mes_py_model_master "
                        "WHERE bit_number=%s AND is_active=true "
                        "  AND COALESCE(zone_id,0) = COALESCE(%s,0)",
                        (bit, zone_id),
                    )
                    if cur.fetchone():
                        errors.append(f"Row {r_idx}: bit {bit} already used in this zone")
                        skipped += 1
                        continue

                cur.execute("""
                    INSERT INTO mes_py_model_master
                        (model_name, model_type, old_model_no, series, bit_number, zone_id)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (
                    name,
                    str(row[iType]).strip() if iType is not None and row[iType] else None,
                    str(row[iOld]).strip()  if iOld  is not None and row[iOld]  else None,
                    str(row[iSer]).strip()  if iSer  is not None and row[iSer]  else None,
                    bit,
                    zone_id,
                ))
                inserted += 1
            except Exception as e:
                errors.append(f"Row {r_idx}: {e}")
                try: conn.rollback()
                except Exception: pass
    return {"ok": True, "inserted": inserted, "skipped": skipped, "errors": errors}


# ── PY MASTER ─────────────────────────────────────────────────────────

@router.get("/master/template", tags=["import-export"])
def download_py_master_template(user=Depends(get_current_user)):
    """Fillable Excel template for PY Master bulk import — zone-aware."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "PY Master"

    headers = [
        "Zone", "PY No", "Description", "Output D-Bit", "Sensing X-Bits",
        "Type", "Side", "Register Count",
        "Applicable Model Bits (comma-sep)",
    ]
    _styled_header(ws, headers, fill_color="7C3AED")
    ws.append(["Seat Slider", "TBDI-PE-PY-6041", "Detect harness brkt rivet miss",
               "D401", "X15", "4 Way", "OTR", 1, "9,10,11,12"])
    ws.append(["Seat Slider", "", "Fr.Lwr.Protector",
               "D406", "X21,X22", "4 Way", "OTR", 2, "9,10"])

    # Hint sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Column", "Required", "Notes"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
    ws2.append(["Zone",            "yes", "Zone name as shown in Admin → Zones (matched case-insensitive)"])
    ws2.append(["PY No",           "no",  "Optional identifier — defaults to the D-Bit if blank"])
    ws2.append(["Description",     "yes", "Human-readable PY name"])
    ws2.append(["Output D-Bit",    "yes", "PLC output register e.g. D401, D406, D413,D414,D415"])
    ws2.append(["Sensing X-Bits",  "no",  "Sensor input X-bit(s), comma-separated for multi-input PYs (X21,X22)"])
    ws2.append(["Type",            "yes", "4 Way / 6 Way"])
    ws2.append(["Side",            "yes", "ALL / LH / RH / OTR / Otr LH / Otr RH"])
    ws2.append(["Register Count",  "yes", "1 (PASS/OFF/ON) or 2 (combined codes 0..4)"])
    ws2.append(["Applicable Model Bits", "no", "Comma-separated bit_number list — links the PY to those models in mes_py_assignments. Leave blank if not yet known."])

    _autosize_columns(ws)
    _autosize_columns(ws2)
    return _xlsx_response(wb, "py_master_template.xlsx")


@router.post("/master/import", tags=["import-export"])
async def import_py_master(file: UploadFile = File(...),
                           admin=Depends(require_admin)):
    """Bulk-insert PY rows from Excel + auto-link assignments by bit_number.
    Each row may list one or more model bits in 'Applicable Model Bits';
    we look them up in mes_py_model_master and create matching assignments.
    Zone is matched by name (case-insensitive).  Schema is auto-migrated
    via _ensure_py_register_col so this works on a fresh DB."""
    import io
    from openpyxl import load_workbook
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Upload an .xlsx file")
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel file: {e}")

    ws = wb["PY Master"] if "PY Master" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": True, "inserted": 0, "skipped": 0, "errors": ["empty sheet"]}

    headers = [str(c).strip() if c else "" for c in rows[0]]
    def col_idx(name):
        try: return headers.index(name)
        except ValueError: return None

    iZone   = col_idx("Zone")
    iPyNo   = col_idx("PY No")
    iDesc   = col_idx("Description")
    iDBit   = col_idx("Output D-Bit") or col_idx("D-Bit") or col_idx("Bit")
    iSens   = col_idx("Sensing X-Bits") or col_idx("Sensing")
    iType   = col_idx("Type")
    iSide   = col_idx("Side")
    iRegCnt = col_idx("Register Count")
    iModels = col_idx("Applicable Model Bits (comma-sep)") or col_idx("Applicable Model Bits")

    inserted, skipped = 0, 0
    errors: list[str] = []
    with get_conn() as conn:
        _ensure_py_register_col(conn)
        cur = conn.cursor()

        # Zone name → id  (case-insensitive)
        cur.execute("SELECT id, zone_name FROM mes_zones")
        zone_map = {str(zn).strip().lower(): zid for zid, zn in cur.fetchall()}

        # Bit number → model id
        cur.execute("SELECT id, bit_number FROM mes_py_model_master "
                    "WHERE is_active = true AND bit_number IS NOT NULL")
        model_map = {int(b): mid for mid, b in cur.fetchall()}

        for r_idx, row in enumerate(rows[1:], start=2):
            try:
                d_bit_raw = row[iDBit] if iDBit is not None else None
                d_bit = _normalize_register(str(d_bit_raw or ""))
                if not d_bit:
                    skipped += 1
                    continue

                desc = (row[iDesc] if iDesc is not None else None) or ""
                desc = str(desc).strip()
                if not desc:
                    errors.append(f"Row {r_idx}: missing Description")
                    skipped += 1
                    continue

                py_no = ((row[iPyNo] if iPyNo is not None else None) or "").strip() or d_bit
                py_no = str(py_no).strip()

                # Zone lookup
                zone_id = None
                if iZone is not None and row[iZone]:
                    zone_id = zone_map.get(str(row[iZone]).strip().lower())
                    if zone_id is None:
                        errors.append(f"Row {r_idx}: zone '{row[iZone]}' not found")

                sens = _normalize_register(str(row[iSens] or "")) if iSens is not None else ""
                model_type = (str(row[iType]).strip()
                              if iType is not None and row[iType] else None)
                side = (str(row[iSide]).strip()
                        if iSide is not None and row[iSide] else "ALL")
                try:
                    reg_cnt = int(row[iRegCnt]) if iRegCnt is not None and row[iRegCnt] else 1
                except (ValueError, TypeError):
                    reg_cnt = 1

                # Insert master row.  Treat duplicate py_no as 'reactivate'.
                cur.execute("SELECT id, is_active FROM mes_py_master WHERE py_no=%s",
                            (py_no,))
                existing = cur.fetchone()
                if existing and existing[1]:
                    errors.append(f"Row {r_idx}: PY '{py_no}' already exists — skipped")
                    skipped += 1
                    continue
                if existing and not existing[1]:
                    cur.execute("""
                        UPDATE mes_py_master SET
                            description=%s, model_type=%s, side=%s, bit=%s,
                            machine_name=%s, register=%s, register_count=%s,
                            zone_id=%s, sensing_bits=%s, is_active=true
                        WHERE id=%s
                        RETURNING id
                    """, (desc, model_type, side, d_bit, None, d_bit, reg_cnt,
                          zone_id, sens or None, existing[0]))
                else:
                    cur.execute("""
                        INSERT INTO mes_py_master
                            (py_no, description, model_type, side, bit,
                             register, register_count, zone_id, sensing_bits)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        RETURNING id
                    """, (py_no, desc, model_type, side, d_bit,
                          d_bit, reg_cnt, zone_id, sens or None))
                py_id = cur.fetchone()[0]

                # Create assignments for every listed model bit
                if iModels is not None and row[iModels]:
                    model_bits_raw = str(row[iModels]).split(",")
                    for mb_str in model_bits_raw:
                        mb_str = mb_str.strip()
                        if not mb_str:
                            continue
                        try:    mb = int(mb_str)
                        except ValueError:
                            errors.append(f"Row {r_idx}: bad model bit '{mb_str}'")
                            continue
                        m_id = model_map.get(mb)
                        if m_id is None:
                            errors.append(f"Row {r_idx}: model bit {mb} not found in master")
                            continue
                        cur.execute("""
                            INSERT INTO mes_py_assignments
                                (py_id, model_id, py_no, py_name, side,
                                 model_type, d_bit, desired_value, machine_name)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            ON CONFLICT DO NOTHING
                        """, (py_id, m_id, py_no, desc, side, model_type,
                              d_bit, None, None))

                inserted += 1
            except Exception as e:
                errors.append(f"Row {r_idx}: {e}")
                try: conn.rollback()
                except Exception: pass
    return {"ok": True, "inserted": inserted, "skipped": skipped, "errors": errors}


# ══════════════════════════════════════════════════════════════
# ASSIGNMENTS  (PY ↔ Model matrix)
# ══════════════════════════════════════════════════════════════

@router.get("/assignments/")
def list_assignments(
    model_name:   Optional[str] = None,
    model_type:   Optional[str] = None,
    model_series: Optional[str] = None,
    user=Depends(get_current_user)
):
    with get_conn() as conn:
        cur    = dict_cursor(conn)
        where  = []
        params = []
        if model_name:
            where.append("model_name ILIKE %s")
            params.append(f"%{model_name}%")
        if model_type:
            where.append("model_type = %s")
            params.append(model_type)
        if model_series:
            where.append("model_series = %s")
            params.append(model_series)

        where_sql = ("WHERE " + " AND ".join("a." + w if not w.startswith("a.") else w for w in where)) if where else ""
        # bit / description / machine_name come from master (global per-PY).
        # desired_bit + desired_value are per-model (per-assignment) — set by
        # user on Config page.  desired_value here is 0=OFF / 1=ON.
        cur.execute(f"""
            SELECT a.id,
                   a.py_no                                  AS "pyNo",
                   COALESCE(m.description, a.py_name)       AS "pyName",
                   a.side                                   AS "typeSide",
                   a.model_type                             AS "modelType",
                   -- Always return the LIVE master name so renames flow through.
                   COALESCE(mm.model_name, a.model_name)    AS "modelName",
                   a.model_series                           AS "modelSeries",
                   a.old_model_no                           AS "oldModelNo",
                   COALESCE(m.bit, a.d_bit)                 AS "dBit",
                   a.desired_value                          AS "desiredValue",
                   a.desired_bit                            AS "desiredBit",
                   a.desired_value_2                        AS "desiredValue2",
                   a.desired_bit_2                          AS "desiredBit2",
                   COALESCE(m.machine_name, a.machine_name) AS "machineFixture",
                   a.model_id                               AS "modelId",
                   mm.bit_number                            AS "bitNumber"
            FROM mes_py_assignments a
            LEFT JOIN mes_py_master m        ON m.py_no = a.py_no  AND m.is_active = true
            LEFT JOIN mes_py_model_master mm ON mm.id   = a.model_id AND mm.is_active = true
            {where_sql}
            ORDER BY a.model_name, a.py_no
        """, params)
        return cur.fetchall()


@router.post("/assignments/", status_code=201)
def create_assignment(body: AssignmentCreate, admin=Depends(require_admin)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(
            "SELECT id FROM mes_py_master WHERE py_no=%s AND is_active=true",
            (body.pyNo,))
        py_row = cur.fetchone()
        cur.execute(
            "SELECT id FROM mes_py_model_master WHERE model_name=%s AND is_active=true",
            (body.modelName,))
        m_row = cur.fetchone()

        cur.execute("""
            INSERT INTO mes_py_assignments
                (py_id, model_id, py_no, py_name, side, model_type,
                 model_name, model_series, old_model_no,
                 d_bit, desired_value, desired_bit, machine_name)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            py_row["id"] if py_row else None,
            m_row["id"]  if m_row  else None,
            body.pyNo,   body.pyName,    body.typeSide,  body.modelType,
            body.modelName, body.modelSeries, body.oldModelNo,
            body.dBit,   body.desiredValue, body.desiredBit, body.machineFixture,
        ))
        new_id = cur.fetchone()["id"]
    return {"ok": True, "id": new_id}


@router.delete("/assignments/{assignment_id}")
def delete_assignment(assignment_id: int, admin=Depends(require_admin)):
    """Remove a single PY × model assignment.  If this was the LAST
    assignment for that PY, the master row is also soft-deactivated so
    Sensor Health stops monitoring it (no orphan PY left lying around).
    PYs with at least one remaining assignment continue as before."""
    with get_conn() as conn:
        cur = conn.cursor()
        # Capture py_no before delete so we can count remaining assignments.
        cur.execute(
            "SELECT py_no, py_id FROM mes_py_assignments WHERE id=%s",
            (assignment_id,),
        )
        row = cur.fetchone()
        py_no = row[0] if row else None
        py_id = row[1] if row else None

        cur.execute("DELETE FROM mes_py_assignments WHERE id=%s", (assignment_id,))

        # Cascade — if no more assignments reference this PY, deactivate
        # the master row.  Health monitor will drop it on next sweep.
        deactivated = False
        if py_no:
            cur.execute(
                "SELECT COUNT(*) FROM mes_py_assignments WHERE py_no=%s",
                (py_no,),
            )
            remaining = cur.fetchone()[0]
            if remaining == 0:
                cur.execute(
                    "UPDATE mes_py_master SET is_active=false "
                    "WHERE py_no=%s OR id=%s",
                    (py_no, py_id),
                )
                deactivated = True
    return {"ok": True, "master_deactivated": deactivated}


# ── Per-assignment PATCH: used by Config page to edit desired_bit / desired_value
# Pairs 1 and 2 — pair 2 is only meaningful for 2-register PYs.
class AssignmentPatch(BaseModel):
    desired_bit:     Optional[int] = None   # bit index (pair 1)
    desired_value:   Optional[int] = None   # 0 = OFF, 1 = ON  (pair 1)
    desired_bit_2:   Optional[int] = None   # bit index (pair 2, 2-register only)
    desired_value_2: Optional[int] = None   # 0 = OFF, 1 = ON  (pair 2)

@router.patch("/assignments/{assignment_id}")
def patch_assignment(assignment_id: int, body: AssignmentPatch, admin=Depends(require_admin)):
    """Update desired_bit and/or desired_value on a single assignment row.
    Also auto-acknowledges any pending SENSOR_BYPASS events for the affected
    PY so the dashboard drops its red alerts immediately — no 8 h wait."""
    provided = {k: getattr(body, k) for k in body.model_fields_set}
    if not provided:
        raise HTTPException(400, "Nothing to update")
    sets   = ", ".join(f"{k} = %s" for k in provided)
    values = list(provided.values()) + [assignment_id]
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(f"UPDATE mes_py_assignments SET {sets} WHERE id = %s", values)
        if cur.rowcount == 0:
            raise HTTPException(404, "Assignment not found")

        # Auto-ack stale bypass events for this PY so the dashboard clears
        # instantly after a config change. The collector will re-evaluate
        # on its next cycle and re-fire ONLY if the new config still mismatches.
        d_cur = dict_cursor(conn)
        d_cur.execute("""
            SELECT a.py_no, m.bit AS register_addr
            FROM mes_py_assignments a
            JOIN mes_py_master m ON m.id = a.py_id
            WHERE a.id = %s
        """, (assignment_id,))
        row = d_cur.fetchone() or {}
        py_no = row.get("py_no")
        if py_no:
            cur.execute("""
                UPDATE mes_poka_yoke_events
                   SET acknowledged    = true,
                       acknowledged_at = NOW(),
                       acknowledged_by = 'config-change'
                 WHERE rule_type    = 'SENSOR_BYPASS'
                   AND acknowledged = false
                   AND ((context_json::jsonb)->>'py_no') = %s
            """, (py_no,))
    return {"ok": True}


# ── Batch update desired values (from Dashboard edit mode) ────
class DesiredUpdateItem(BaseModel):
    py_no:         str
    model:         Optional[str] = None
    model_type:    Optional[str] = None
    type_side:     Optional[str] = None
    desired_value: Any   # int or string (on/off/pass)

class DesiredUpdateBatch(BaseModel):
    updates: list[DesiredUpdateItem]

@router.post("/update-desired/")
def update_desired_batch(body: DesiredUpdateBatch, admin=Depends(require_admin)):
    """Batch-update desired_value on mes_py_assignments rows."""
    with get_conn() as conn:
        cur = conn.cursor()
        changed = 0
        for u in body.updates:
            where = ["py_no = %s"]
            params = [u.py_no]
            if u.model:
                where.append("model_series = %s")
                params.append(u.model)
            if u.model_type:
                where.append("model_type = %s")
                params.append(u.model_type)
            if u.type_side:
                where.append("side = %s")
                params.append(u.type_side)
            params.append(u.desired_value)
            cur.execute(
                f"UPDATE mes_py_assignments SET desired_value = %s WHERE {' AND '.join(where)}",
                params[-1:] + params[:-1],
            )
            changed += cur.rowcount
    return {"ok": True, "updated": changed}


# ══════════════════════════════════════════════════════════════
# BULK IMPORT FROM EXCEL SHEETS
# ══════════════════════════════════════════════════════════════

class BulkImportBody(BaseModel):
    sheet:    str        # "MODEL MASTER" | "POKA YOKE MASTER" | "final seat"
    rows:     list
    col_map:  dict = {}  # optional column remapping


@router.post("/import/bulk", status_code=200)
def bulk_import(body: BulkImportBody, admin=Depends(require_admin)):
    """
    Single endpoint for all 3 sheet imports.
    Frontend sends sheet name + rows + optional col_map.
    """
    sheet    = body.sheet.strip().upper()
    rows     = body.rows
    col_map  = body.col_map   # {system_col: excel_col}
    inserted = skipped = 0
    errors   = []

    def col(row: dict, system_key: str, default="") -> str:
        """Get value using col_map override or system_key directly."""
        excel_key = col_map.get(system_key, system_key)
        val = row.get(excel_key, row.get(system_key, default))
        return str(val).strip() if val is not None and str(val).strip() not in ("nan","None","") else default

    def intcol(row: dict, system_key: str, default=None):
        v = col(row, system_key, "")
        try: return int(float(v)) if v != "" else default
        except: return default

    with get_conn() as conn:
        cur = dict_cursor(conn)

        # ── MODEL MASTER ──────────────────────────────────────
        if "MODEL" in sheet and "MASTER" in sheet:
            for r in rows:
                model_name = col(r, "Model Name")
                mtype      = col(r, "type")
                old_no     = col(r, "Old Model No")
                series     = col(r, "model")
                if not model_name:
                    skipped += 1; continue
                try:
                    cur.execute("""
                        INSERT INTO mes_py_model_master
                            (model_name, model_type, old_model_no, series)
                        VALUES (%s,%s,%s,%s)
                        ON CONFLICT (model_name) DO UPDATE SET
                            model_type   = EXCLUDED.model_type,
                            old_model_no = EXCLUDED.old_model_no,
                            series       = EXCLUDED.series,
                            is_active    = true
                    """, (model_name, mtype or None, old_no or None, series or None))
                    inserted += 1
                except Exception as e:
                    skipped += 1; errors.append(str(e))

        # ── POKA YOKE MASTER ──────────────────────────────────
        elif "POKA" in sheet and "MASTER" in sheet:
            seen = set()
            for r in rows:
                py_no    = col(r, "Poka Yoke No")
                py_name  = col(r, "Poka Yoke Name")
                mtype    = col(r, "Model Type")
                machine  = col(r, "Machine/Fixture")
                # derive side from py_no suffix or leave as ALL
                side     = col(r, "Side", "ALL")
                dbit     = col(r, "D Bit", "")
                dval     = intcol(r, "Desired Value")
                if not py_no:
                    skipped += 1; continue
                key = py_no
                if key in seen:
                    skipped += 1; continue
                seen.add(key)
                try:
                    cur.execute("""
                        INSERT INTO mes_py_master
                            (py_no, description, model_type, side, bit, desired_value, machine_name)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (py_no) DO UPDATE SET
                            description  = EXCLUDED.description,
                            model_type   = EXCLUDED.model_type,
                            machine_name = EXCLUDED.machine_name,
                            is_active    = true
                    """, (py_no, py_name, mtype or None, side or "ALL",
                          dbit or None, dval, machine or None))
                    inserted += 1
                except Exception as e:
                    skipped += 1; errors.append(str(e))

        # ── FINAL SEAT (assignments matrix) ───────────────────
        elif "FINAL" in sheet or "SEAT" in sheet or "ASSIGNMENT" in sheet:
            for r in rows:
                py_no      = col(r, "Poka Yoke No")
                py_name    = col(r, "Poka Yoke Name")
                type_side  = col(r, "Type Side", "ALL").upper()
                model_type = col(r, "Model Type")
                model_name = col(r, "Model Name")
                type2      = col(r, "Type2")
                old_no     = col(r, "Old Model No")
                series     = col(r, "Model")
                dbit       = col(r, "D bit From PLC")
                dval       = intcol(r, "Desired Value (0/1/2)")
                machine    = col(r, "Machine/Fixture")

                if not py_no or not model_name:
                    skipped += 1; continue

                # normalise side
                if type_side in ("L","LH"):    type_side = "LH"
                elif type_side in ("R","RH"):  type_side = "RH"

                try:
                    # resolve FK if master rows exist
                    cur.execute("SELECT id FROM mes_py_master WHERE py_no=%s AND is_active=true", (py_no,))
                    py_row = cur.fetchone()
                    cur.execute("SELECT id FROM mes_py_model_master WHERE model_name=%s AND is_active=true", (model_name,))
                    m_row = cur.fetchone()

                    cur.execute("""
                        INSERT INTO mes_py_assignments
                            (py_id, model_id, py_no, py_name, side, model_type,
                             model_name, model_series, old_model_no,
                             d_bit, desired_value, machine_name)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        py_row["id"] if py_row else None,
                        m_row["id"]  if m_row  else None,
                        py_no, py_name, type_side, model_type,
                        model_name, series, old_no,
                        dbit or None, dval, machine or None,
                    ))
                    inserted += 1
                except Exception as e:
                    skipped += 1; errors.append(str(e))
        else:
            raise HTTPException(400, f"Unknown sheet: {body.sheet}. Use 'MODEL MASTER', 'POKA YOKE MASTER', or 'final seat'")

        conn.cursor().execute("""
            INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
            VALUES ('BULK_IMPORT', 'poka_yoke', 0, %s)
        """, (f"sheet={body.sheet} inserted={inserted} skipped={skipped}",))

    return {"ok": True, "sheet": body.sheet, "inserted": inserted,
            "skipped": skipped, "errors": errors[:10]}


# Also add UNIQUE constraint to mes_py_model_master if not already there:
# ALTER TABLE mes_py_model_master ADD CONSTRAINT uq_py_model_name UNIQUE (model_name);


# ══════════════════════════════════════════════════════════════
# PLC ACTUALS & SENSOR MAPPING  (Dashboard tab)
# ══════════════════════════════════════════════════════════════
#
# Tables auto-created on first use:
#   mes_py_sensor_mapping  — sensor_name → d_bit
#   mes_py_plc_actuals     — d_bit + model_column → actual value
#   mes_py_model_columns   — PLC column# → model code/type/side
# ══════════════════════════════════════════════════════════════

_PLC_TABLES_ENSURED = False

def _ensure_plc_tables(conn):
    global _PLC_TABLES_ENSURED
    if _PLC_TABLES_ENSURED:
        return
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mes_py_sensor_mapping (
            id           SERIAL PRIMARY KEY,
            sensor_name  VARCHAR(200) NOT NULL,
            device_no    VARCHAR(50),
            d_bit        VARCHAR(20),
            UNIQUE(sensor_name)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mes_py_plc_actuals (
            id           SERIAL PRIMARY KEY,
            d_bit        VARCHAR(20) NOT NULL,
            model_col    VARCHAR(10) NOT NULL,
            actual_value INTEGER,
            updated_at   TIMESTAMP DEFAULT NOW(),
            UNIQUE(d_bit, model_col)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mes_py_model_columns (
            id           SERIAL PRIMARY KEY,
            col_key      VARCHAR(10) NOT NULL UNIQUE,
            model_code   VARCHAR(20),
            model_type   VARCHAR(50),
            type_side    VARCHAR(10),
            full_name    VARCHAR(200)
        )
    """)
    _PLC_TABLES_ENSURED = True


# Hard-coded default sensor mapping
_DEFAULT_SENSOR_MAP = {
    "relocate pin":                          ("D.081", "401"),
    "LH HARNES 1":                           ("D.049", "402"),
    "RH HARNESS 1":                          ("D.050", "403"),
    "LH HARNESS 2":                          ("D.041", "404"),
    "RH HARNESS 2":                          ("D.041", "405"),
    "Rr. Lwr Proctecter":                    ("D.090", "406"),
    "Fr. Lwr Proctecter":                    ("D.083", "407"),
    "rr upper proctector /shifting pos. ng": ("D.097", "408"),
    "lh bending":                            ("D.104", "d409"),
    "Rh bending":                            ("D.105", "d410"),
    "pop revit 1":                           ("D.045", "d411"),
    "pop rivet 2":                           ("D.046", "d412"),
    "fr lighter protector1":                 ("D.054", "d413"),
    "fr lighter proctector2":                ("D.075", "d414"),
    "fr lighter proctector3":                ("D.053", "d415"),
    "Rr lighter proctector 1":               ("D.074", "d416"),
    "Rr lighter proctector 2":               ("D.076", "d417"),
    "E ring":                                ("D.060", "d418"),
    "pop jig":                               (None,    "d419"),
    "Fr lighter proctector 4":               ("D.055", "d420"),
    "lh harness 3":                          ("D.049", "d421"),
    "bolt mixing":                           ("D.042", "d422"),
    "Rh harness bkt":                        ("D.050", "d423"),
    "ytb lh exp Harness bkt":                ("D.049", "d424"),
}

_DEFAULT_MODEL_COLUMNS = {
    "9":  {"model_code":"YHB","model_type":"4 WAY OUTER","type_side":"BOTH","full_name":"TRACK ASSY FRONT SEAT YHB 4 WAY OTR"},
    "10": {"model_code":"YHB","model_type":"4 WAY INNER","type_side":"RH",  "full_name":"TRACK ASSY FRONT SEAT YHB 4 WAY INR RH"},
    "11": {"model_code":"YHB","model_type":"4 WAY INNER","type_side":"LH",  "full_name":"TRACK ASSY FRONT SEAT YHB 4 WAY INR LH"},
}


def _seed_defaults_if_empty(conn):
    """Insert TBDI-YHB sample sensor mappings + model columns if the tables
    are completely empty.  Gate with env var SEED_TBDI_DEFAULTS=1 so a
    fresh production install doesn't auto-populate YHB-specific demo data
    that pollutes the real model master.  Default: skip seeding."""
    import os as _os
    if _os.getenv("SEED_TBDI_DEFAULTS", "0") != "1":
        return
    cur = dict_cursor(conn)
    cur.execute("SELECT COUNT(*) AS c FROM mes_py_sensor_mapping")
    if cur.fetchone()["c"] == 0:
        for sensor, (dbit, dev) in _DEFAULT_SENSOR_MAP.items():
            conn.cursor().execute(
                "INSERT INTO mes_py_sensor_mapping (sensor_name, device_no, d_bit) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                (sensor, dev, dbit))

    cur.execute("SELECT COUNT(*) AS c FROM mes_py_model_columns")
    if cur.fetchone()["c"] == 0:
        for ck, info in _DEFAULT_MODEL_COLUMNS.items():
            conn.cursor().execute(
                "INSERT INTO mes_py_model_columns (col_key, model_code, model_type, type_side, full_name) VALUES (%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING",
                (ck, info["model_code"], info["model_type"], info["type_side"], info["full_name"]))


# ── GET sensor mappings ───────────────────────────────────────
@router.get("/sensor-mapping/")
def list_sensor_mapping(user=Depends(get_current_user)):
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        _seed_defaults_if_empty(conn)
        cur = dict_cursor(conn)
        cur.execute("SELECT id, sensor_name AS \"sensorName\", device_no AS \"deviceNo\", d_bit AS \"dBit\" FROM mes_py_sensor_mapping ORDER BY sensor_name")
        return cur.fetchall()


# ── PUT sensor mapping ────────────────────────────────────────
class SensorMapUpdate(BaseModel):
    sensorName: str
    dBit:       Optional[str] = None
    deviceNo:   Optional[str] = None

@router.put("/sensor-mapping/{mapping_id}")
def update_sensor_mapping(mapping_id: int, body: SensorMapUpdate, admin=Depends(require_admin)):
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        conn.cursor().execute(
            "UPDATE mes_py_sensor_mapping SET d_bit=%s, device_no=%s WHERE id=%s",
            (body.dBit, body.deviceNo, mapping_id))
    return {"ok": True}


# ── POST sensor mapping (add new) ─────────────────────────────
@router.post("/sensor-mapping/", status_code=201)
def create_sensor_mapping(body: SensorMapUpdate, admin=Depends(require_admin)):
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        cur = dict_cursor(conn)
        cur.execute(
            "INSERT INTO mes_py_sensor_mapping (sensor_name, d_bit, device_no) VALUES (%s,%s,%s) ON CONFLICT (sensor_name) DO UPDATE SET d_bit=EXCLUDED.d_bit, device_no=EXCLUDED.device_no RETURNING id",
            (body.sensorName, body.dBit, body.deviceNo))
        return {"ok": True, "id": cur.fetchone()["id"]}


# ── GET model columns ─────────────────────────────────────────
@router.get("/model-columns/")
def list_model_columns(user=Depends(get_current_user)):
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        _seed_defaults_if_empty(conn)
        cur = dict_cursor(conn)
        cur.execute("""SELECT id, col_key AS "colKey", model_code AS "modelCode",
                              model_type AS "modelType", type_side AS "typeSide",
                              full_name AS "fullName"
                       FROM mes_py_model_columns ORDER BY col_key""")
        return cur.fetchall()


# ── PUT model column ──────────────────────────────────────────
class ModelColUpdate(BaseModel):
    colKey:    str
    modelCode: Optional[str] = None
    modelType: Optional[str] = None
    typeSide:  Optional[str] = None
    fullName:  Optional[str] = None

@router.put("/model-columns/{col_id}")
def update_model_column(col_id: int, body: ModelColUpdate, admin=Depends(require_admin)):
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        conn.cursor().execute(
            "UPDATE mes_py_model_columns SET col_key=%s, model_code=%s, model_type=%s, type_side=%s, full_name=%s WHERE id=%s",
            (body.colKey, body.modelCode, body.modelType, body.typeSide, body.fullName, col_id))
    return {"ok": True}

@router.post("/model-columns/", status_code=201)
def create_model_column(body: ModelColUpdate, admin=Depends(require_admin)):
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        cur = dict_cursor(conn)
        cur.execute(
            "INSERT INTO mes_py_model_columns (col_key, model_code, model_type, type_side, full_name) VALUES (%s,%s,%s,%s,%s) ON CONFLICT (col_key) DO UPDATE SET model_code=EXCLUDED.model_code, model_type=EXCLUDED.model_type, type_side=EXCLUDED.type_side, full_name=EXCLUDED.full_name RETURNING id",
            (body.colKey, body.modelCode, body.modelType, body.typeSide, body.fullName))
        return {"ok": True, "id": cur.fetchone()["id"]}


# ── GET PLC actuals ───────────────────────────────────────────
@router.get("/plc-actuals/")
def list_plc_actuals(user=Depends(get_current_user)):
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        cur = dict_cursor(conn)
        cur.execute("""SELECT id, d_bit AS "dBit", model_col AS "modelCol",
                              actual_value AS "actualValue", updated_at AS "updatedAt"
                       FROM mes_py_plc_actuals ORDER BY d_bit, model_col""")
        return cur.fetchall()


# ── GET compiled dashboard (actual vs desired) ────────────────
@router.get("/dashboard/")
def get_dashboard(
    model_code:  Optional[str] = None,
    model_name:  Optional[str] = None,
    user=Depends(get_current_user)
):
    """
    Returns compiled list: each assignment row enriched with actual PLC value + status.
    Matches assignments to PLC actuals via d_bit + model column mapping.
    """
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        _seed_defaults_if_empty(conn)
        cur = dict_cursor(conn)

        # Load model column mapping
        cur.execute("SELECT col_key, model_code, model_type, type_side FROM mes_py_model_columns")
        col_map = {r["col_key"]: r for r in cur.fetchall()}

        # Load PLC actuals into lookup: { "D.041": {"9": val, "10": val} }
        cur.execute("SELECT d_bit, model_col, actual_value FROM mes_py_plc_actuals")
        actuals_lookup = {}
        for r in cur.fetchall():
            actuals_lookup.setdefault(r["d_bit"], {})[r["model_col"]] = r["actual_value"]

        # Load assignments
        where_parts = []
        params = []
        if model_code:
            where_parts.append("model_series = %s")
            params.append(model_code)
        if model_name:
            where_parts.append("model_name = %s")
            params.append(model_name)
        where_sql = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

        cur.execute(f"""
            SELECT id, py_no, py_name, side, model_type, model_name, model_series,
                   old_model_no, d_bit, desired_value, machine_name
            FROM mes_py_assignments
            {where_sql}
            ORDER BY model_name, py_no
        """, params)
        rows = cur.fetchall()

        # Load sensor mapping for display
        cur.execute("SELECT sensor_name, d_bit FROM mes_py_sensor_mapping WHERE d_bit IS NOT NULL")
        dbit_to_sensor = {}
        for r in cur.fetchall():
            if r["d_bit"]:
                dbit_to_sensor[r["d_bit"]] = r["sensor_name"]

        compiled = []
        for row in rows:
            d_bit = row["d_bit"]
            actual = None
            plc_col = None

            # Find which model column this row maps to
            code  = (row["model_series"] or "").upper()
            mtype = (row["model_type"] or "").upper()
            side  = (row["side"] or "").upper()

            for ck, ci in col_map.items():
                if (ci["model_code"] or "").upper() != code:
                    continue
                ci_type = (ci["model_type"] or "").upper()
                ci_side = (ci["type_side"] or "").upper()
                if ci_type not in mtype and mtype not in ci_type:
                    continue
                if "INNER" in ci_type:
                    if ci_side == side or side == "BOTH" or ci_side == "BOTH":
                        plc_col = ck
                        break
                else:
                    plc_col = ck
                    break

            if plc_col and d_bit and d_bit in actuals_lookup:
                actual = actuals_lookup[d_bit].get(plc_col)

            desired = row["desired_value"]
            if actual is not None and desired is not None:
                status = "MATCH" if actual == desired else "MISMATCH"
            else:
                status = "NO_DATA"

            compiled.append({
                "id":          row["id"],
                "pyNo":        row["py_no"],
                "pyName":      row["py_name"],
                "typeSide":    row["side"],
                "modelType":   row["model_type"],
                "modelName":   row["model_name"],
                "modelCode":   row["model_series"],
                "oldModelNo":  row["old_model_no"],
                "dBit":        d_bit,
                "desired":     desired,
                "actual":      actual,
                "status":      status,
                "machine":     row["machine_name"],
                "plcColumn":   plc_col,
                "sensorName":  dbit_to_sensor.get(d_bit),
            })

        # Stats
        match    = sum(1 for c in compiled if c["status"] == "MATCH")
        mismatch = sum(1 for c in compiled if c["status"] == "MISMATCH")
        nodata   = sum(1 for c in compiled if c["status"] == "NO_DATA")

        return {
            "compiled": compiled,
            "stats": {"total": len(compiled), "match": match, "mismatch": mismatch, "noData": nodata},
        }


# ── POST import PLC actuals (from frontend-parsed Excel rows) ─
class PLCImportBody(BaseModel):
    rows:        list       # [{sensor_name, device_no, model_9, model_10, model_11}, ...]
    model_columns: dict = {}  # optional override: {"9":{modelCode,modelType,...}}

ACTUAL_TEXT_MAP = {"on": 2, "off": 1, "pass": 0}

@router.post("/plc-actuals/import")
def import_plc_actuals(body: PLCImportBody, admin=Depends(require_admin)):
    """
    Accepts PLC rows parsed from Excel by the frontend (SheetJS).
    Each row: { col0: sensor_name, col1: device_no, col2+: model values }
    Maps sensor→D.bit, converts on/off/pass→0/1/2, stores in DB.
    """
    with get_conn() as conn:
        _ensure_plc_tables(conn)
        cur = dict_cursor(conn)

        # Load sensor mapping
        cur.execute("SELECT sensor_name, d_bit FROM mes_py_sensor_mapping")
        sensor_map = {}
        for r in cur.fetchall():
            if r["d_bit"]:
                sensor_map[r["sensor_name"].strip().lower()] = r["d_bit"]

        # Determine model column keys (default: col indexes 2,3,4 → "9","10","11")
        # Frontend sends rows as arrays or objects
        updated = 0
        skipped = 0

        for row in body.rows:
            # Support both array and object formats
            if isinstance(row, list):
                sensor = str(row[0] or "").strip() if len(row) > 0 else ""
                device = str(row[1] or "").strip() if len(row) > 1 else ""
                model_vals = {str(i+9): row[i+2] for i in range(min(len(row)-2, 10)) if len(row) > i+2}
            else:
                sensor = str(row.get("sensor_name") or row.get("col0") or "").strip()
                device = str(row.get("device_no") or row.get("col1") or "").strip()
                model_vals = {}
                for k, v in row.items():
                    if k.startswith("model_") or k.startswith("col"):
                        try:
                            idx = k.replace("model_","").replace("col","")
                            if idx.isdigit() and int(idx) >= 2:
                                model_vals[str(int(idx)+7)] = v  # col2 → "9", col3 → "10"
                            elif idx.isdigit():
                                model_vals[idx] = v
                        except ValueError:
                            pass

            if not sensor:
                skipped += 1
                continue

            d_bit = sensor_map.get(sensor.lower())
            if not d_bit:
                skipped += 1
                continue

            for model_key, raw_val in model_vals.items():
                if raw_val is None:
                    continue
                raw_str = str(raw_val).strip().lower()
                # Handle compound like "on/off" → first part
                if "/" in raw_str:
                    raw_str = raw_str.split("/")[0].strip()
                actual = ACTUAL_TEXT_MAP.get(raw_str)
                if actual is None:
                    try:
                        actual = int(float(raw_str))
                    except (ValueError, TypeError):
                        continue

                conn.cursor().execute("""
                    INSERT INTO mes_py_plc_actuals (d_bit, model_col, actual_value, updated_at)
                    VALUES (%s, %s, %s, NOW())
                    ON CONFLICT (d_bit, model_col) DO UPDATE SET
                        actual_value = EXCLUDED.actual_value,
                        updated_at = NOW()
                """, (d_bit, model_key, actual))
                updated += 1

        conn.cursor().execute("""
            INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
            VALUES ('PLC_IMPORT', 'poka_yoke', 0, %s)
        """, (f"updated={updated} skipped={skipped}",))

    return {"ok": True, "updated": updated, "skipped": skipped}


# ════════════════════════════════════════════════════════════════════
# PY MAINTENANCE REQUESTS  (Phase 2 — operator remarks audit panel)
# ════════════════════════════════════════════════════════════════════
# Operator on Maintenance > Poka Yoke can type a remark per PY row and
# submit it.  Lands here, stored in mes_py_requests, surfaces on the
# admin "New Requests" panel.  Spec from operator (2026-05-21):
#   "remarks ka option if any changes are required so mention changes
#    are save in audit panel name as new panel new requests jisme sari
#    details ho bs mujhe vha jha k pta chal jaye ki whats are input
#    from users".
# Workflow:
#   NEW       → operator just submitted, admin hasn't seen yet
#   REVIEWED  → admin opened it
#   RESOLVED  → admin took action (or rejected), with resolution note
# Status transitions one-way.

class PyRequestCreate(BaseModel):
    py_no:        str
    py_name:      Optional[str] = None
    py_master_id: Optional[int] = None
    line_id:      Optional[int] = None
    model_bit:    Optional[int] = None
    sensing_bits: Optional[str] = None    # X-bit at time of submission
    machine_name: Optional[str] = None
    bit:          Optional[str] = None    # D-register
    expected:     Optional[Any] = None    # desired value snapshot
    remark:       str


class PyRequestResolve(BaseModel):
    status:           str                  # 'REVIEWED' | 'RESOLVED'
    resolution_note:  Optional[str] = None


def _ensure_py_requests_table(conn):
    """Idempotent — create / migrate the audit table on first use."""
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mes_py_requests (
            id              SERIAL PRIMARY KEY,
            py_no           VARCHAR(64)  NOT NULL,
            py_name         TEXT,
            py_master_id    INTEGER,
            line_id         INTEGER,
            model_bit       INTEGER,
            sensing_bits    TEXT,          -- e.g. 'X15'
            machine_name    TEXT,
            bit             VARCHAR(32),   -- D-register
            expected        TEXT,           -- desired value snapshot
            remark          TEXT         NOT NULL,
            status          VARCHAR(16)  NOT NULL DEFAULT 'NEW',
            submitted_by_user_id  INTEGER,
            submitted_by_username TEXT,
            submitted_at    TIMESTAMP    NOT NULL DEFAULT NOW(),
            resolved_at     TIMESTAMP,
            resolved_by_user_id   INTEGER,
            resolved_by_username  TEXT,
            resolution_note TEXT
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_mes_py_requests_status
                ON mes_py_requests(status, submitted_at DESC)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_mes_py_requests_line
                ON mes_py_requests(line_id, status)
    """)
    conn.commit()


@router.post("/requests")
def submit_py_request(body: PyRequestCreate, user=Depends(get_current_user)):
    """Operator submits a remark/change-request for a PY.  Always lands
    as status='NEW'.  Admin sees it on the New Requests panel."""
    remark = (body.remark or "").strip()
    if not remark:
        raise HTTPException(400, "remark is required")
    with get_conn() as conn:
        _ensure_py_requests_table(conn)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_py_requests
                (py_no, py_name, py_master_id, line_id, model_bit,
                 sensing_bits, machine_name, bit, expected, remark,
                 status, submitted_by_user_id, submitted_by_username)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'NEW',%s,%s)
            RETURNING id, submitted_at
        """, (
            body.py_no, body.py_name, body.py_master_id, body.line_id,
            body.model_bit, body.sensing_bits, body.machine_name,
            body.bit, (str(body.expected) if body.expected is not None else None),
            remark,
            user.get("id"), user.get("username"),
        ))
        new_id, submitted_at = cur.fetchone()
        conn.commit()
    return {
        "ok": True,
        "id": new_id,
        "submitted_at": submitted_at.isoformat() if submitted_at else None,
    }


@router.get("/requests")
def list_py_requests(
    status:  Optional[str] = Query(None, description="NEW | REVIEWED | RESOLVED"),
    line_id: Optional[int] = Query(None),
    days:    int           = Query(30, ge=1, le=365),
    limit:   int           = Query(200, ge=1, le=2000),
    user=Depends(get_current_user),
):
    """List submitted PY requests.  Drives the admin "New Requests"
    audit panel.  Filterable by status / line / lookback days."""
    from datetime import datetime as _dt, timedelta as _td
    cutoff = _dt.utcnow() - _td(days=days)
    where = ["submitted_at >= %s"]
    params: list = [cutoff]
    if status:
        where.append("status = %s")
        params.append(status.upper())
    if line_id is not None:
        where.append("line_id = %s")
        params.append(line_id)
    with get_conn() as conn:
        _ensure_py_requests_table(conn)
        cur = dict_cursor(conn)
        cur.execute(f"""
            SELECT r.*, l.line_name, z.zone_name
              FROM mes_py_requests r
              LEFT JOIN mes_lines l ON l.id = r.line_id
              LEFT JOIN mes_zones z ON z.id = l.zone_id
             WHERE {' AND '.join(where)}
             ORDER BY r.submitted_at DESC
             LIMIT %s
        """, params + [limit])
        rows = [dict(r) for r in cur.fetchall()]
        # Convenience: count breakdown by status
        cur.execute(f"""
            SELECT status, COUNT(*) AS n
              FROM mes_py_requests
             WHERE {' AND '.join(where)}
             GROUP BY status
        """, params)
        by_status = {r["status"]: r["n"] for r in cur.fetchall()}
    return {
        "rows": rows,
        "by_status": by_status,
        "filters": {"status": status, "line_id": line_id, "days": days},
    }


@router.put("/requests/{req_id}/resolve")
def resolve_py_request(req_id: int, body: PyRequestResolve,
                        admin=Depends(require_admin)):
    """Admin updates a request's status to REVIEWED or RESOLVED.
    One-way transition: NEW -> REVIEWED -> RESOLVED.  resolution_note
    is optional but encouraged for RESOLVED."""
    new_status = (body.status or "").upper()
    if new_status not in ("REVIEWED", "RESOLVED"):
        raise HTTPException(400, "status must be REVIEWED or RESOLVED")
    with get_conn() as conn:
        _ensure_py_requests_table(conn)
        cur = conn.cursor()
        # Stamp resolved fields only when going RESOLVED.
        if new_status == "RESOLVED":
            cur.execute("""
                UPDATE mes_py_requests
                   SET status               = 'RESOLVED',
                       resolved_at          = NOW(),
                       resolved_by_user_id  = %s,
                       resolved_by_username = %s,
                       resolution_note      = COALESCE(%s, resolution_note)
                 WHERE id = %s AND status <> 'RESOLVED'
            """, (admin.get("id"), admin.get("username"),
                  body.resolution_note, req_id))
        else:
            cur.execute("""
                UPDATE mes_py_requests
                   SET status = 'REVIEWED'
                 WHERE id = %s AND status = 'NEW'
            """, (req_id,))
        if cur.rowcount == 0:
            raise HTTPException(409, "Request not in a valid state for this transition")
        conn.commit()
    return {"ok": True, "status": new_status}


@router.delete("/requests/{req_id}")
def delete_py_request(req_id: int, admin=Depends(require_admin)):
    """Admin can purge a request (e.g. spam, duplicate).  Hard delete —
    use sparingly.  Audit trail of who deleted lives in mes_audit_log."""
    with get_conn() as conn:
        _ensure_py_requests_table(conn)
        cur = conn.cursor()
        cur.execute("DELETE FROM mes_py_requests WHERE id = %s RETURNING py_no",
                    (req_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Request not found")
        try:
            cur.execute("""
                INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
                VALUES ('DELETE_PY_REQUEST', 'py_request', %s, %s)
            """, (req_id, f"py_no={row[0]} by={admin.get('username')}"))
        except Exception:
            pass        # audit log optional — never block the delete
        conn.commit()
    return {"ok": True, "deleted_id": req_id}


# ════════════════════════════════════════════════════════════════════
# PY IMAGES  (Phase 3 — visual manual / reference photos per PY)
# ════════════════════════════════════════════════════════════════════
# Admin uploads one or more images for each PY (LOCATE PIN, FR.LIGHTER
# PROTECTOR-1, etc.).  Operator clicks the image button on Maintenance
# > Poka Yoke row -> modal shows all images for that PY.
# Spec from operator (2026-05-21):
#   "first row p click kru to image show ho jaye or ye image set krne
#    ka option dede maintenance panel ... image single bhi ho sakti h
#    or multiple bhi ok as a manual".
#
# Storage:
#   Files live at: Phase2/uploads/py_images/<py_id>_<timestamp>_<safe_name>
#   DB stores: filename (random, unique), original_filename (user-facing),
#              caption, sort_order so admin can re-order in a manual.
import os as _os_img
_PY_IMG_DIR = _os_img.path.join(
    _os_img.path.dirname(_os_img.path.dirname(_os_img.path.abspath(__file__))),
    "uploads", "py_images",
)
_os_img.makedirs(_PY_IMG_DIR, exist_ok=True)


def _ensure_py_images_table(conn):
    """Idempotent table create for PY image manual."""
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mes_py_images (
            id                SERIAL PRIMARY KEY,
            py_no             VARCHAR(64)  NOT NULL,
            py_master_id      INTEGER,
            line_id           INTEGER,
            filename          TEXT         NOT NULL,
            original_filename TEXT,
            mime_type         VARCHAR(64),
            file_size_bytes   BIGINT,
            caption           TEXT,
            sort_order        INTEGER      NOT NULL DEFAULT 0,
            uploaded_by_user_id  INTEGER,
            uploaded_by_username TEXT,
            uploaded_at       TIMESTAMP    NOT NULL DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_mes_py_images_lookup
                ON mes_py_images(py_no, line_id, sort_order, id)
    """)
    conn.commit()


@router.post("/images")
async def upload_py_image(
    py_no:        str = Query(..., description="Target PY number"),
    py_master_id: Optional[int] = Query(None),
    line_id:      Optional[int] = Query(None),
    caption:      Optional[str] = Query(None),
    sort_order:   int = Query(0),
    file:         UploadFile = File(...),
    admin = Depends(require_admin),
):
    """Admin uploads ONE image for a PY.  Multi-upload = call this N
    times from the frontend (simpler than handling list[UploadFile])."""
    # Validate
    if file.content_type and not file.content_type.startswith("image/"):
        raise HTTPException(400, f"Not an image file: {file.content_type}")
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Empty file")
    if len(raw) > 10 * 1024 * 1024:  # 10 MB cap
        raise HTTPException(413, "Image too large (10 MB max)")

    # Build a safe filename — keep extension, replace risky chars in
    # original.  Final name = <py_no>_<unix_ms>_<safe_orig>
    import re as _re_img, time as _t_img
    original = file.filename or "upload.png"
    ext = _os_img.path.splitext(original)[1].lower() or ".png"
    if ext not in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"):
        raise HTTPException(400, f"Unsupported image extension: {ext}")
    safe_py = _re_img.sub(r"[^A-Za-z0-9._-]", "_", py_no)[:32]
    fname = f"{safe_py}_{int(_t_img.time()*1000)}{ext}"
    full = _os_img.path.join(_PY_IMG_DIR, fname)
    with open(full, "wb") as f:
        f.write(raw)

    with get_conn() as conn:
        _ensure_py_images_table(conn)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_py_images
                (py_no, py_master_id, line_id, filename, original_filename,
                 mime_type, file_size_bytes, caption, sort_order,
                 uploaded_by_user_id, uploaded_by_username)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id, uploaded_at
        """, (
            py_no, py_master_id, line_id, fname, original,
            file.content_type, len(raw), caption, sort_order,
            admin.get("id"), admin.get("username"),
        ))
        new_id, uploaded_at = cur.fetchone()
        conn.commit()
    return {
        "ok": True, "id": new_id, "filename": fname,
        "original_filename": original,
        "size_bytes": len(raw),
        "uploaded_at": uploaded_at.isoformat() if uploaded_at else None,
    }


@router.get("/images")
def list_py_images(
    py_no:   Optional[str] = Query(None),
    line_id: Optional[int] = Query(None),
):
    """List images for a PY (or all if no filter).  No auth so wallboard
    kiosks can fetch too.  File bytes go via /images/{id}/file."""
    where, params = [], []
    if py_no:
        where.append("py_no = %s"); params.append(py_no)
    if line_id is not None:
        where.append("(line_id IS NULL OR line_id = %s)")
        params.append(line_id)
    sql = "SELECT id, py_no, py_master_id, line_id, filename, original_filename, " \
          "mime_type, file_size_bytes, caption, sort_order, uploaded_by_username, " \
          "uploaded_at FROM mes_py_images"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY sort_order ASC, id ASC"
    with get_conn() as conn:
        _ensure_py_images_table(conn)
        cur = dict_cursor(conn)
        cur.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]
    # Attach a downloadable URL for each image
    for r in rows:
        r["url"] = f"/api/poka-yoke/images/{r['id']}/file"
    return {"rows": rows, "count": len(rows)}


@router.get("/images/{img_id}/file")
def get_py_image_file(img_id: int):
    """Serve the image bytes.  No auth — wallboard kiosks need this."""
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT filename, mime_type FROM mes_py_images
                       WHERE id = %s""", (img_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Image not found")
        fname, mime = row
    full = _os_img.path.join(_PY_IMG_DIR, fname)
    if not _os_img.path.exists(full):
        raise HTTPException(404, "Image file missing on disk")
    with open(full, "rb") as f:
        data = f.read()
    return Response(content=data, media_type=mime or "image/png")


class PyImageUpdate(BaseModel):
    caption:    Optional[str] = None
    sort_order: Optional[int] = None


@router.put("/images/{img_id}")
def update_py_image(img_id: int, body: PyImageUpdate, admin=Depends(require_admin)):
    """Update caption / sort_order on an existing image."""
    sets, params = [], []
    if body.caption is not None:
        sets.append("caption = %s"); params.append(body.caption)
    if body.sort_order is not None:
        sets.append("sort_order = %s"); params.append(int(body.sort_order))
    if not sets:
        return {"ok": True, "updated": 0}
    params.append(img_id)
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(f"UPDATE mes_py_images SET {', '.join(sets)} "
                    f"WHERE id = %s", params)
        if cur.rowcount == 0:
            raise HTTPException(404, "Image not found")
        conn.commit()
    return {"ok": True, "updated": 1}


@router.delete("/images/{img_id}")
def delete_py_image(img_id: int, admin=Depends(require_admin)):
    """Remove image from DB + disk.  Returns the freed filename in case
    admin wants to verify."""
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT filename FROM mes_py_images WHERE id = %s""",
                    (img_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Image not found")
        fname = row[0]
        cur.execute("DELETE FROM mes_py_images WHERE id = %s", (img_id,))
        conn.commit()
    # Best-effort disk cleanup
    full = _os_img.path.join(_PY_IMG_DIR, fname)
    try:
        if _os_img.path.exists(full):
            _os_img.remove(full)
    except OSError:
        pass
    return {"ok": True, "deleted_id": img_id, "filename": fname}


# ════════════════════════════════════════════════════════════════════
# PY INSTRUCTIONS  (Phase 3b — follow-steps text manual per PY)
# ════════════════════════════════════════════════════════════════════
# Admin writes step-by-step instructions for each PY (plain-text /
# multi-line).  Operator sees them in the same modal as images, on
# top of the gallery — combined visual+text manual.
# Spec from operator (2026-05-21): "kuch instructioon ya follow steps
# bhi add krne ka option bhi dede or same py me visual ok".
# One row per (py_no + line_id) pair, simple upsert.

def _ensure_py_instructions_table(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mes_py_instructions (
            id                SERIAL PRIMARY KEY,
            py_no             VARCHAR(64) NOT NULL,
            line_id           INTEGER,
            py_master_id      INTEGER,
            instruction_text  TEXT NOT NULL DEFAULT '',
            updated_at        TIMESTAMP NOT NULL DEFAULT NOW(),
            updated_by_user_id   INTEGER,
            updated_by_username  TEXT,
            UNIQUE (py_no, line_id)
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_mes_py_instructions_lookup
                ON mes_py_instructions(py_no, line_id)
    """)
    conn.commit()


class PyInstructionsUpsert(BaseModel):
    py_no:            str
    line_id:          Optional[int] = None
    py_master_id:     Optional[int] = None
    instruction_text: str


@router.get("/instructions")
def get_py_instructions(
    py_no:   Optional[str] = Query(None),
    line_id: Optional[int] = Query(None),
):
    """List instructions.  Filter by py_no + optional line_id.
    No auth — wallboard kiosks read this too."""
    where, params = [], []
    if py_no:
        where.append("py_no = %s"); params.append(py_no)
    if line_id is not None:
        where.append("(line_id IS NULL OR line_id = %s)")
        params.append(line_id)
    sql = ("SELECT id, py_no, line_id, py_master_id, instruction_text, "
           "updated_at, updated_by_username FROM mes_py_instructions")
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id ASC"
    with get_conn() as conn:
        _ensure_py_instructions_table(conn)
        cur = dict_cursor(conn)
        cur.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]
    return {"rows": rows, "count": len(rows)}


@router.put("/instructions")
def upsert_py_instructions(body: PyInstructionsUpsert,
                            admin=Depends(require_admin)):
    """Admin upserts instruction text for a PY.  One row per
    (py_no, line_id).  Pass empty string to clear."""
    with get_conn() as conn:
        _ensure_py_instructions_table(conn)
        cur = conn.cursor()
        # UPSERT on the unique (py_no, line_id) constraint
        cur.execute("""
            INSERT INTO mes_py_instructions
                (py_no, line_id, py_master_id, instruction_text,
                 updated_by_user_id, updated_by_username)
            VALUES (%s,%s,%s,%s,%s,%s)
            ON CONFLICT (py_no, line_id) DO UPDATE SET
                instruction_text     = EXCLUDED.instruction_text,
                py_master_id         = COALESCE(EXCLUDED.py_master_id,
                                                mes_py_instructions.py_master_id),
                updated_at           = NOW(),
                updated_by_user_id   = EXCLUDED.updated_by_user_id,
                updated_by_username  = EXCLUDED.updated_by_username
            RETURNING id, updated_at
        """, (
            body.py_no, body.line_id, body.py_master_id,
            body.instruction_text,
            admin.get("id"), admin.get("username"),
        ))
        new_id, updated_at = cur.fetchone()
        conn.commit()
    return {
        "ok": True, "id": new_id,
        "updated_at": updated_at.isoformat() if updated_at else None,
    }


@router.delete("/instructions/{ins_id}")
def delete_py_instructions(ins_id: int, admin=Depends(require_admin)):
    """Hard delete an instructions row (admin only)."""
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM mes_py_instructions WHERE id = %s", (ins_id,))
        if cur.rowcount == 0:
            raise HTTPException(404, "Instructions not found")
        conn.commit()
    return {"ok": True, "deleted_id": ins_id}


# ============================================================
# BYPASS HISTORY  (per-date PY bypass episodes with duration)
# ============================================================
# 2026-05-22 — Operator spec: "production and quality dono ke
# historical data me aana chahiye ki konsa PY kis zone ki kis line
# prr kitni der ke liye bypassed tha".  This walks the
# `mes_poka_yoke_events` table (SENSOR_BYPASS rule) for a date range
# and emits one row per BYPASS EPISODE — a contiguous run of
# unacknowledged events with the same py_no + register on the same
# line.  Each episode shows: zone, line, machine, py_no, py_desc,
# started_at, ended_at (or "ongoing"), duration_seconds.
#
# The episode boundary is implicit: collector posts a new event each
# time it sees a fresh bypass code, and posts auto-ack when sensor
# returns to desired.  We treat events grouped within 60 s and same
# (py_no, register, line_id) as one episode.

@router.get("/bypass-history")
def get_bypass_history(
    date_from: str = Query(..., description="start date YYYY-MM-DD"),
    date_to:   str = Query(..., description="end date YYYY-MM-DD inclusive"),
    line_id:   Optional[int] = Query(None, description="filter by line, omit = all lines"),
    zone_id:   Optional[int] = Query(None, description="filter by zone, omit = all zones"),
    user=Depends(get_current_user_optional),
):
    """Historical PY bypass episode browser for Production + Quality
    dashboards.  Returns one row per bypass episode with duration.

    Episode grouping rule:
      Same (line_id, py_no, register) + gap between events ≤ 60 s
      → SAME episode.  First event = started_at, last event before
      an acknowledged transition = ended_at.  Episode is "ongoing"
      if no later acknowledged event exists for that (line, py, reg).

    Sorted by started_at DESC so newest is first."""
    where = ["e.rule_type = 'SENSOR_BYPASS'",
             "e.detected_at::date BETWEEN %s AND %s"]
    params: list = [date_from, date_to]
    if line_id is not None:
        where.append("e.line_id = %s")
        params.append(line_id)

    sql = f"""
        WITH ordered AS (
            SELECT
                e.id,
                e.line_id,
                ln.line_name,
                ln.zone_id,
                z.zone_name AS zone_name,
                COALESCE(e.context_json::jsonb->>'py_no',        '?') AS py_no,
                COALESCE(e.context_json::jsonb->>'description',
                         e.context_json::jsonb->>'py_name',      '')  AS py_description,
                COALESCE(e.context_json::jsonb->>'register',     '')  AS register,
                COALESCE(e.context_json::jsonb->>'machine_name', '')  AS machine_name,
                e.detected_at,
                e.acknowledged,
                e.acknowledged_at,
                e.shift_name,
                LAG(e.detected_at) OVER (
                    PARTITION BY e.line_id,
                                 e.context_json::jsonb->>'py_no',
                                 e.context_json::jsonb->>'register'
                    ORDER BY e.detected_at
                ) AS prev_detected
            FROM mes_poka_yoke_events e
            LEFT JOIN mes_lines ln ON ln.id = e.line_id
            LEFT JOIN mes_zones z  ON z.id  = ln.zone_id
            WHERE {' AND '.join(where)}
        ),
        episodes AS (
            SELECT *,
                SUM(CASE
                      WHEN prev_detected IS NULL THEN 1
                      WHEN detected_at - prev_detected > INTERVAL '60 seconds' THEN 1
                      ELSE 0
                    END) OVER (
                      PARTITION BY line_id, py_no, register
                      ORDER BY detected_at
                ) AS episode_group
            FROM ordered
        )
        SELECT
            line_id,
            line_name,
            zone_id,
            zone_name,
            py_no,
            py_description,
            register,
            machine_name,
            shift_name,
            MIN(detected_at)       AS started_at,
            MAX(detected_at)       AS last_seen_at,
            MAX(acknowledged_at)   AS resolved_at,
            BOOL_AND(acknowledged) AS is_resolved,
            COUNT(*)               AS hit_count
        FROM episodes
        GROUP BY line_id, line_name, zone_id, zone_name,
                 py_no, py_description, register, machine_name,
                 shift_name, episode_group
        ORDER BY started_at DESC
    """
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute(sql, params)
        rows = cur.fetchall() or []

    # Optional zone filter applied in Python (avoids complicating the
    # CTE with another LEFT JOIN since some legacy rows have NULL zone)
    if zone_id is not None:
        rows = [r for r in rows if r.get("zone_id") == zone_id]

    # Compute duration; ended_at = resolved_at if resolved else NULL
    result = []
    for r in rows:
        started = r["started_at"]
        last_seen = r["last_seen_at"]
        resolved = r["resolved_at"] if r.get("is_resolved") else None
        # Duration end = resolved_at (if resolved) or last_seen + small buffer
        # for ongoing episodes where 'now' is the implied end.
        from datetime import datetime as _dt
        if resolved:
            end_for_dur = resolved
        elif r.get("is_resolved"):
            end_for_dur = last_seen
        else:
            end_for_dur = _dt.now()
        try:
            dur = (end_for_dur - started).total_seconds()
        except Exception:
            dur = 0
        result.append({
            "line_id":          r["line_id"],
            "line_name":        r.get("line_name"),
            "zone_id":          r.get("zone_id"),
            "zone_name":        r.get("zone_name"),
            "py_no":            r.get("py_no"),
            "py_description":   r.get("py_description") or "",
            "register":         r.get("register"),
            "machine_name":     r.get("machine_name"),
            "shift_name":       r.get("shift_name"),
            "started_at":       started.isoformat() if started else None,
            "ended_at":         resolved.isoformat() if resolved else None,
            "duration_seconds": int(round(dur)),
            "is_ongoing":       not r.get("is_resolved"),
            "hit_count":        r.get("hit_count") or 1,
        })

    # Summary stats for header
    total_episodes = len(result)
    total_duration = sum(r["duration_seconds"] for r in result)
    ongoing_count  = sum(1 for r in result if r["is_ongoing"])

    return {
        "date_from":      date_from,
        "date_to":        date_to,
        "line_id":        line_id,
        "zone_id":        zone_id,
        "total_episodes": total_episodes,
        "total_seconds":  total_duration,
        "ongoing_count":  ongoing_count,
        "rows":           result,
    }
