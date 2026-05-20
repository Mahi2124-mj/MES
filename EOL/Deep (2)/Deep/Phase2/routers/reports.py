"""
routers/reports.py
==================
Shift-level reporting:
  - Excel download (formatted, one row per hourly slot + KPI footer)
  - PDF download (matplotlib-rendered, charts + KPI grid)
  - Background scheduler that mails the PDF to a configured list at
    every shift-end transition

Endpoints
---------
GET  /api/reports/shift-excel?line_id=&date=&shift=
GET  /api/reports/shift-pdf?line_id=&date=&shift=
GET  /api/reports/email-config         (admin)
PUT  /api/reports/email-config         (admin)
POST /api/reports/email-now            (admin — manual fire)

The auto-email scheduler is started by main.py at app boot; it lives in
the same module so the import graph stays simple.
"""
from __future__ import annotations

import io
import os
import threading
import time
import traceback
from datetime import datetime, date, timedelta
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from database import get_conn, dict_cursor
from auth import require_admin, get_current_user
from routers.breakdown_mail import _send_email   # reuse the same SMTP helper

router = APIRouter(prefix="/api/reports", tags=["reports"])


# ════════════════════════════════════════════════════════════════════
#  Email-config table — list of recipients per line / report kind
# ════════════════════════════════════════════════════════════════════

def _ensure_email_config_table() -> None:
    """Idempotent create.  Runs once at first endpoint hit."""
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mes_report_email_config (
                id           SERIAL PRIMARY KEY,
                line_id      INTEGER NOT NULL,
                report_kind  VARCHAR(20) NOT NULL DEFAULT 'shift_end',
                to_addresses TEXT NOT NULL DEFAULT '',
                cc_addresses TEXT NOT NULL DEFAULT '',
                is_active    BOOLEAN NOT NULL DEFAULT TRUE,
                updated_at   TIMESTAMP DEFAULT NOW(),
                UNIQUE (line_id, report_kind)
            )
        """)
        conn.commit()


# ════════════════════════════════════════════════════════════════════
#  Data loader — pulls one shift's full row out of the dashboard table
# ════════════════════════════════════════════════════════════════════

def _load_shift_row(line_id: int, record_date: date, shift_name: str) -> dict:
    """Return the dashboard row for this shift (latest if multiple).
    Raises 404 if nothing found."""
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT l.line_name, l.db_table_name
              FROM mes_lines l WHERE l.id = %s
        """, (line_id,))
        line = cur.fetchone()
        if not line:
            raise HTTPException(404, f"line_id {line_id} not found")
        table = line["db_table_name"]
        cur.execute(f"""
            SELECT *
              FROM {table}
             WHERE record_date = %s AND shift_name = %s
             ORDER BY id DESC
             LIMIT 1
        """, (record_date, shift_name))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404,
                f"No data for line {line_id} on {record_date} shift {shift_name}")
        row["line_name"]      = line["line_name"]
        row["table_name"]     = table
        row["_report_date"]   = record_date
        row["_report_shift"]  = shift_name
        return dict(row)


def _hourly_slots_from_row(row: dict) -> List[dict]:
    """Walk hour_*_plan / actual / ok / ng columns into a list of
    {label, plan, actual, ok, ng, variance} dicts in slot order."""
    out = []
    seen = set()
    for k in row.keys():
        if not k.startswith("hour_") or not k.endswith("_plan"):
            continue
        prefix = k[:-5]          # strip "_plan"
        label_raw = prefix[5:]   # strip "hour_"
        if label_raw in seen:
            continue
        seen.add(label_raw)
        # "0830_0930" -> "08:30-09:30"
        try:
            a, b = label_raw.split("_")
            label = f"{a[:2]}:{a[2:]}-{b[:2]}:{b[2:]}"
        except Exception:
            label = label_raw
        out.append({
            "label":    label,
            "plan":     row.get(f"{prefix}_plan")     or 0,
            "actual":   row.get(f"{prefix}_actual")   or 0,
            "ok":       row.get(f"{prefix}_ok")       or 0,
            "ng":       row.get(f"{prefix}_ng")       or 0,
            "variance": row.get(f"{prefix}_variance") or 0,
        })
    return out


# ════════════════════════════════════════════════════════════════════
#  EXCEL render
# ════════════════════════════════════════════════════════════════════

def build_excel_report(row: dict) -> bytes:
    """Return xlsx bytes for one shift."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Report"

    BOLD  = Font(bold=True, color="FFFFFF")
    HDR   = PatternFill("solid", fgColor="1F4E79")
    SUB   = PatternFill("solid", fgColor="DCE6F1")
    OK    = PatternFill("solid", fgColor="C6EFCE")
    NG    = PatternFill("solid", fgColor="FFC7CE")
    CENTER = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="888888")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title band
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"{row.get('line_name','')} — Shift {row.get('shift_name','')} — {row.get('_report_date','')}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = HDR
    c.alignment = CENTER

    # KPI block
    kpis = [
        ("Plan",         row.get("shift_plan", 0)),
        ("Actual",       (row.get("ok_count", 0) or 0) + (row.get("ng_count", 0) or 0)),
        ("OK",           row.get("ok_count", 0)),
        ("NG",           row.get("ng_count", 0)),
        ("Availability", f"{float(row.get('availability') or 0):.1f}%"),
        ("Performance",  f"{float(row.get('performance') or 0):.1f}%"),
        ("Quality",      f"{float(row.get('quality_oee') or 0):.1f}%"),
        ("Overall OEE",  f"{float(row.get('overall_oee') or 0):.1f}%"),
        ("OEE Grade",    row.get("oee_grade", "—")),
        ("Model",        row.get("current_model_name") or "—"),
        ("Ideal CT (s)", row.get("cycle_time_plan", "—")),
        ("Actual CT (s)",row.get("cycle_time_actual", "—")),
    ]
    r = 3
    for i, (lab, val) in enumerate(kpis):
        col = 1 + (i % 3) * 2
        ws.cell(row=r + i // 3, column=col,     value=lab).font = Font(bold=True)
        ws.cell(row=r + i // 3, column=col + 1, value=val)
        ws.cell(row=r + i // 3, column=col).fill = SUB
        ws.cell(row=r + i // 3, column=col).border     = BORDER
        ws.cell(row=r + i // 3, column=col + 1).border = BORDER

    # Hourly slot grid
    slots = _hourly_slots_from_row(row)
    hdr_row = r + (len(kpis) + 2) // 3 + 2
    headers = ["Slot", "Plan", "Actual", "OK", "NG", "Variance"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=j, value=h)
        c.font = BOLD; c.fill = HDR; c.alignment = CENTER; c.border = BORDER
    for i, s in enumerate(slots, start=1):
        ws.cell(row=hdr_row + i, column=1, value=s["label"]).border = BORDER
        ws.cell(row=hdr_row + i, column=2, value=s["plan"]).border = BORDER
        cell_actual = ws.cell(row=hdr_row + i, column=3, value=s["actual"])
        cell_actual.border = BORDER
        cell_actual.fill   = OK if s["actual"] >= s["plan"] else NG
        ws.cell(row=hdr_row + i, column=4, value=s["ok"]).border = BORDER
        ws.cell(row=hdr_row + i, column=5, value=s["ng"]).border = BORDER
        ws.cell(row=hdr_row + i, column=6, value=s["variance"]).border = BORDER

    # Column widths
    for col_letter, w in zip("ABCDEF", (16, 12, 12, 10, 10, 12)):
        ws.column_dimensions[col_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════
#  PDF render (matplotlib — no external dependency)
# ════════════════════════════════════════════════════════════════════

def build_pdf_report(row: dict) -> bytes:
    """Return PDF bytes for one shift.  Single-page A4 landscape with
    KPI tiles, hourly bar chart, and OEE gauges drawn in matplotlib."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.patches import FancyBboxPatch, Rectangle, Circle

    fig = plt.figure(figsize=(11.69, 8.27))
    fig.patch.set_facecolor("white")
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, 100); ax.set_ylim(0, 70)
    ax.set_xticks([]); ax.set_yticks([])
    for s in ax.spines.values(): s.set_visible(False)

    # ── Title bar
    ax.add_patch(Rectangle((0, 65), 100, 5, color="#1F4E79"))
    ax.text(2, 67.5,
            f"{row.get('line_name','')}  ·  Shift {row.get('shift_name','')}  ·  {row.get('_report_date','')}",
            color="white", weight="bold", fontsize=15, va="center")
    ax.text(99, 67.5, "End-of-Shift Report",
            color="white", fontsize=10, va="center", ha="right")

    # ── KPI tiles (top row)
    overall = float(row.get("overall_oee") or 0)
    grade   = row.get("oee_grade", "—") or "—"
    avail   = float(row.get("availability") or 0)
    perf    = float(row.get("performance") or 0)
    qual    = float(row.get("quality_oee") or 0)
    plan    = int(row.get("shift_plan") or 0)
    ok      = int(row.get("ok_count") or 0)
    ng      = int(row.get("ng_count") or 0)
    actual  = ok + ng

    def tile(x, y, w, h, label, value, color):
        ax.add_patch(FancyBboxPatch((x, y), w, h,
                    boxstyle="round,pad=0.1,rounding_size=0.4",
                    facecolor=color, edgecolor=color))
        ax.text(x + w/2, y + h*0.62, value, ha="center", va="center",
                color="white", weight="bold", fontsize=18)
        ax.text(x + w/2, y + h*0.22, label, ha="center", va="center",
                color="white", fontsize=9.5, alpha=0.95)

    tile( 2, 53,  18, 10, "OVERALL OEE", f"{overall:.1f}%",
          "#06A77D" if overall >= 75 else ("#F4A261" if overall >= 50 else "#D62828"))
    tile(22, 53,  18, 10, "AVAILABILITY",  f"{avail:.1f}%",  "#3A86FF")
    tile(42, 53,  18, 10, "PERFORMANCE",   f"{perf:.1f}%",   "#19376D")
    tile(62, 53,  18, 10, "QUALITY",       f"{qual:.1f}%",   "#5E548E")
    tile(82, 53,  16, 10, "GRADE",         grade,            "#475569")

    # ── Production block
    tile( 2, 41,  22, 10, "PLAN",    f"{plan:,}",   "#0B2447")
    tile(26, 41,  22, 10, "ACTUAL",  f"{actual:,}", "#06A77D")
    tile(50, 41,  22, 10, "OK",      f"{ok:,}",     "#06A77D")
    tile(74, 41,  22, 10, "NG",      f"{ng:,}",     "#D62828")

    # ── Hourly bar chart inset
    slots = _hourly_slots_from_row(row)
    chart = fig.add_axes([0.06, 0.07, 0.65, 0.32])
    if slots:
        idx = list(range(len(slots)))
        chart.bar([i - 0.20 for i in idx], [s["plan"] for s in slots],
                  width=0.4, label="Plan",   color="#3A86FF")
        chart.bar([i + 0.20 for i in idx], [s["actual"] for s in slots],
                  width=0.4, label="Actual", color="#06A77D")
        chart.set_xticks(idx)
        chart.set_xticklabels([s["label"] for s in slots], rotation=30, ha="right", fontsize=8)
        chart.set_title("Hourly Plan vs Actual", fontsize=10, weight="bold", color="#1F4E79")
        chart.legend(loc="upper right", fontsize=8)
        chart.grid(axis="y", linestyle=":", alpha=0.5)
        for spine in ("top", "right"): chart.spines[spine].set_visible(False)

    # ── Model / CT side block
    side = fig.add_axes([0.74, 0.07, 0.24, 0.32])
    side.set_xticks([]); side.set_yticks([])
    for s in side.spines.values(): s.set_visible(False)
    side.text(0.04, 0.92, "MODEL", color="#1F4E79", weight="bold", fontsize=10)
    side.text(0.04, 0.85, str(row.get("current_model_name") or "—"), fontsize=9, wrap=True)
    side.text(0.04, 0.72, "IDEAL CT (s)", color="#1F4E79", weight="bold", fontsize=10)
    side.text(0.04, 0.65, str(row.get("cycle_time_plan") or "—"), fontsize=11)
    side.text(0.04, 0.55, "ACTUAL CT (s)", color="#1F4E79", weight="bold", fontsize=10)
    side.text(0.04, 0.48, str(row.get("cycle_time_actual") or "—"), fontsize=11)
    side.text(0.04, 0.36, "STARTED",  color="#1F4E79", weight="bold", fontsize=10)
    side.text(0.04, 0.29, str(row.get("shift_start_time") or "—"), fontsize=9)
    side.text(0.04, 0.18, "ENDED", color="#1F4E79", weight="bold", fontsize=10)
    side.text(0.04, 0.11, str(row.get("shift_end_time") or "—"), fontsize=9)

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════
#  Endpoints
# ════════════════════════════════════════════════════════════════════

@router.get("/shift-excel")
def shift_excel(line_id: int = Query(...),
                date: str = Query(...),
                shift: str = Query(...),
                user=Depends(get_current_user)):
    """Stream a formatted Excel report for one shift.  Anyone with a
    valid token can download (intentionally permissive — reports flow
    up the org chart, no PII)."""
    try:
        d = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "date must be YYYY-MM-DD")
    row = _load_shift_row(line_id, d, shift)
    data = build_excel_report(row)
    fname = f"shift_{row.get('line_name','line')}_{d}_{shift}.xlsx".replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/shift-pdf")
def shift_pdf(line_id: int = Query(...),
              date: str = Query(...),
              shift: str = Query(...),
              user=Depends(get_current_user)):
    """Stream a formatted PDF report for one shift."""
    try:
        d = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "date must be YYYY-MM-DD")
    row = _load_shift_row(line_id, d, shift)
    data = build_pdf_report(row)
    fname = f"shift_{row.get('line_name','line')}_{d}_{shift}.pdf".replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Email-config CRUD ────────────────────────────────────────────────

class EmailConfigUpsert(BaseModel):
    line_id:      int
    report_kind:  str = "shift_end"
    to_addresses: str = ""
    cc_addresses: str = ""
    is_active:    bool = True


@router.get("/email-config")
def list_email_config(line_id: Optional[int] = None,
                       user=Depends(get_current_user)):
    _ensure_email_config_table()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        if line_id is not None:
            cur.execute("""SELECT * FROM mes_report_email_config
                            WHERE line_id = %s ORDER BY report_kind""", (line_id,))
        else:
            cur.execute("SELECT * FROM mes_report_email_config ORDER BY line_id, report_kind")
        return cur.fetchall()


@router.put("/email-config")
def upsert_email_config(body: EmailConfigUpsert,
                         admin=Depends(require_admin)):
    _ensure_email_config_table()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_report_email_config
                (line_id, report_kind, to_addresses, cc_addresses, is_active)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (line_id, report_kind) DO UPDATE
                SET to_addresses = EXCLUDED.to_addresses,
                    cc_addresses = EXCLUDED.cc_addresses,
                    is_active    = EXCLUDED.is_active,
                    updated_at   = NOW()
        """, (body.line_id, body.report_kind,
              body.to_addresses, body.cc_addresses, body.is_active))
        conn.commit()
    return {"ok": True}


# ── Manual fire (admin) ──────────────────────────────────────────────

class EmailNowBody(BaseModel):
    line_id: int
    date:    str   # YYYY-MM-DD
    shift:   str
    kinds:   List[str] = ["excel", "pdf"]   # which attachments


@router.post("/email-now")
def email_now(body: EmailNowBody, admin=Depends(require_admin)):
    """Generate the shift report(s) and email them to the configured
    recipients NOW.  Useful for testing the chain or replaying a missed
    auto-send."""
    _ensure_email_config_table()
    try:
        d = datetime.strptime(body.date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "date must be YYYY-MM-DD")
    row = _load_shift_row(body.line_id, d, body.shift)

    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT to_addresses, cc_addresses FROM mes_report_email_config
                        WHERE line_id=%s AND report_kind='shift_end' AND is_active=TRUE""",
                    (body.line_id,))
        cfg = cur.fetchone()
    if not cfg or not cfg.get("to_addresses"):
        raise HTTPException(400, "No active email-config for this line — add recipients first.")

    to_list = [a.strip() for a in (cfg["to_addresses"] or "").split(",") if a.strip()]
    cc_list = [a.strip() for a in (cfg.get("cc_addresses") or "").split(",") if a.strip()]
    subject = f"[MES · End-of-Shift] {row.get('line_name','')} — Shift {body.shift} {body.date}"
    overall = float(row.get("overall_oee") or 0)
    grade   = row.get("oee_grade", "—") or "—"
    html = f"""
    <p>End-of-shift summary attached.</p>
    <ul>
      <li>OEE: <b>{overall:.1f}%</b> ({grade})</li>
      <li>Plan: {row.get('shift_plan','—')}  ·  Actual: {(row.get('ok_count',0) or 0)+(row.get('ng_count',0) or 0)}</li>
      <li>OK: {row.get('ok_count','—')}  ·  NG: {row.get('ng_count','—')}</li>
      <li>Model: {row.get('current_model_name','—')}</li>
    </ul>"""

    # We attach files via multipart — extend _send_email with our own
    # attachment-aware send (the breakdown_mail._send_email is HTML-only).
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587") or 587)
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    if not (smtp_user and smtp_pass):
        raise HTTPException(500, "SMTP credentials not configured (.env SMTP_USER / SMTP_PASS)")

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(to_list)
    if cc_list: msg["Cc"] = ", ".join(cc_list)
    msg.attach(MIMEText(html, "html"))

    if "excel" in body.kinds:
        xlsx = build_excel_report(row)
        part = MIMEApplication(xlsx, _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.add_header("Content-Disposition", "attachment",
                        filename=f"shift_{body.date}_{body.shift}.xlsx")
        msg.attach(part)
    if "pdf" in body.kinds:
        pdf = build_pdf_report(row)
        part = MIMEApplication(pdf, _subtype="pdf")
        part.add_header("Content-Disposition", "attachment",
                        filename=f"shift_{body.date}_{body.shift}.pdf")
        msg.attach(part)

    with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
        server.ehlo(); server.starttls(); server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, to_list + cc_list, msg.as_string())
    return {"ok": True, "to": to_list, "cc": cc_list}


# ════════════════════════════════════════════════════════════════════
#  Auto-mail scheduler  (run as a daemon thread from main.py startup)
# ════════════════════════════════════════════════════════════════════

_AUTO_REPORT_THREAD: Optional[threading.Thread] = None
_AUTO_REPORT_STOP   = threading.Event()
_LAST_SENT_KEY: dict = {}   # {(line_id, date, shift): True}


def _shift_end_today(line_id: int, shift_name: str) -> Optional[datetime]:
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT start_time, end_time FROM mes_shift_configs
                        WHERE line_id=%s AND shift_name=%s""", (line_id, shift_name))
        r = cur.fetchone()
        if not r or not r.get("end_time"): return None
        return datetime.combine(date.today(), r["end_time"])


def _scheduler_tick() -> None:
    """One pass through every line × every active shift config.  When
    the wall-clock has just crossed `end_time + 90s` and we haven't
    sent today, fire the email and remember we did."""
    try:
        with get_conn() as conn:
            cur = dict_cursor(conn)
            cur.execute("""SELECT DISTINCT c.line_id, sc.shift_name
                             FROM mes_report_email_config c
                             JOIN mes_shift_configs sc ON sc.line_id = c.line_id
                            WHERE c.is_active = TRUE
                              AND c.report_kind = 'shift_end'
                              AND NOT sc.shift_name LIKE 'GAP%'""")
            jobs = cur.fetchall()
    except Exception as exc:
        print(f"[REPORT-SCHED] DB error: {exc}")
        return

    now = datetime.now()
    for j in jobs:
        line_id, shift = j["line_id"], j["shift_name"]
        end_dt = _shift_end_today(line_id, shift)
        if not end_dt:
            continue
        # Send window: 60-240 s after shift end (90 s grace for collector
        # to flush its final row).  Single-shot per (line, date, shift).
        delta = (now - end_dt).total_seconds()
        if not (60 <= delta <= 240):
            continue
        key = (line_id, date.today(), shift)
        if _LAST_SENT_KEY.get(key):
            continue
        try:
            from fastapi.testclient import TestClient   # not used; reuse via direct call
        except Exception:
            pass
        # Direct in-process call — mimics the email_now() body without HTTP.
        try:
            row = _load_shift_row(line_id, date.today(), shift)
            with get_conn() as conn:
                cur = dict_cursor(conn)
                cur.execute("""SELECT to_addresses, cc_addresses
                                 FROM mes_report_email_config
                                WHERE line_id=%s AND report_kind='shift_end' AND is_active=TRUE""",
                            (line_id,))
                cfg = cur.fetchone()
            if not cfg or not cfg.get("to_addresses"):
                continue
            to_list = [a.strip() for a in (cfg["to_addresses"] or "").split(",") if a.strip()]
            cc_list = [a.strip() for a in (cfg.get("cc_addresses") or "").split(",") if a.strip()]
            if not to_list:
                continue
            xlsx = build_excel_report(row)
            pdf  = build_pdf_report(row)
            import smtplib
            from email.mime.text        import MIMEText
            from email.mime.multipart   import MIMEMultipart
            from email.mime.application import MIMEApplication
            smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
            smtp_port = int(os.getenv("SMTP_PORT", "587") or 587)
            smtp_user = os.getenv("SMTP_USER", "")
            smtp_pass = os.getenv("SMTP_PASS", "")
            if not (smtp_user and smtp_pass):
                print("[REPORT-SCHED] SMTP_USER/PASS not set, skipping send.")
                _LAST_SENT_KEY[key] = True   # don't retry every minute
                continue

            overall = float(row.get("overall_oee") or 0)
            grade   = row.get("oee_grade", "—") or "—"
            msg = MIMEMultipart()
            msg["Subject"] = f"[MES · End-of-Shift] {row.get('line_name','')} — Shift {shift} {date.today()}"
            msg["From"] = smtp_user
            msg["To"]   = ", ".join(to_list)
            if cc_list: msg["Cc"] = ", ".join(cc_list)
            html = f"""<p>End-of-shift summary attached.</p>
                       <ul><li>OEE: <b>{overall:.1f}%</b> ({grade})</li>
                       <li>Plan: {row.get('shift_plan','—')}  ·  Actual: {(row.get('ok_count',0) or 0)+(row.get('ng_count',0) or 0)}</li>
                       <li>OK: {row.get('ok_count','—')}  ·  NG: {row.get('ng_count','—')}</li>
                       <li>Model: {row.get('current_model_name','—')}</li></ul>"""
            msg.attach(MIMEText(html, "html"))
            p1 = MIMEApplication(xlsx, _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            p1.add_header("Content-Disposition", "attachment",
                          filename=f"shift_{date.today()}_{shift}.xlsx")
            msg.attach(p1)
            p2 = MIMEApplication(pdf, _subtype="pdf")
            p2.add_header("Content-Disposition", "attachment",
                          filename=f"shift_{date.today()}_{shift}.pdf")
            msg.attach(p2)
            with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as srv:
                srv.ehlo(); srv.starttls(); srv.login(smtp_user, smtp_pass)
                srv.sendmail(smtp_user, to_list + cc_list, msg.as_string())
            _LAST_SENT_KEY[key] = True
            print(f"[REPORT-SCHED] Sent shift-end report for line {line_id} shift {shift} to {to_list}")
        except Exception as exc:
            print(f"[REPORT-SCHED] send failed for line {line_id} shift {shift}: {exc}")
            traceback.print_exc()


def _scheduler_loop() -> None:
    while not _AUTO_REPORT_STOP.is_set():
        try:
            _scheduler_tick()
        except Exception as exc:
            print(f"[REPORT-SCHED] tick error: {exc}")
        # Reset the "sent today" memory at midnight so tomorrow's shifts
        # can fire fresh.
        if datetime.now().hour == 0 and datetime.now().minute < 2:
            _LAST_SENT_KEY.clear()
        _AUTO_REPORT_STOP.wait(30)


def start_scheduler() -> None:
    global _AUTO_REPORT_THREAD
    if _AUTO_REPORT_THREAD and _AUTO_REPORT_THREAD.is_alive():
        return
    _AUTO_REPORT_STOP.clear()
    _AUTO_REPORT_THREAD = threading.Thread(target=_scheduler_loop, daemon=True,
                                            name="report-scheduler")
    _AUTO_REPORT_THREAD.start()
    print("[REPORT-SCHED] Worker started — checks every 30 s for end-of-shift sends")
