"""
main.py
=======
Platform — FastAPI Backend
Toyota Boshoku Device India Pvt. Ltd., Bawal, Haryana
"""

import os
import io
import csv
from datetime import datetime, timedelta
from typing import List, Any

# IMPORTANT: load .env BEFORE importing any router. Routers read env vars
# (e.g. CYCLE_VIDEO_BASE_URL) at import time — if dotenv loads after, they
# silently fall back to defaults and proxy calls hit the wrong port.
from dotenv import load_dotenv
# override=True so .env values WIN over any pre-set system env vars.
# Windows sometimes has empty ANTHROPIC_API_KEY / OPENAI_API_KEY set
# at user level; without override the .env value gets ignored and
# downstream API clients fail with auth errors.
load_dotenv(override=True)

import uvicorn
from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from database import get_conn, dict_cursor
from auth import auth_router, get_current_user, require_admin
from routers.plants          import router as plants_router
from routers.lines           import router as lines_router
from routers.config          import router as config_router
from routers.poka_yoke       import router as poka_router
from routers.status_schema   import router as status_router
from routers.users           import router as users_router
from routers.zones           import router as zones_router
from routers.non_production  import router as npd_router
from routers.submachines     import router as submachines_router
from routers.departments     import router as departments_router
from routers.breakdowns      import router as breakdowns_router
from routers.machines        import router as machines_router
from routers.breakdown_mail  import router as breakdown_mail_router
from routers.reports         import router as reports_router, start_scheduler as start_reports_scheduler
from routers.operators       import router as operators_router
from routers.manpower        import router as manpower_router, start_watcher as start_manpower_watcher
from routers.store_dispatch  import router as store_router, dispatch_router
from routers.shift_calc      import router as shift_calc_router
from routers.kanban          import router as kanban_router, start_watcher as start_kanban_watcher
from routers.anything_wrong  import router as anything_wrong_router
from routers.heijunka        import router as heijunka_router
from routers.five_s          import router as five_s_router
from routers.pdca            import router as pdca_router
from routers.cms_sync        import router as cms_sync_router
from routers.maintenance_kpi import router as maintenance_kpi_router
from routers.capa            import router as capa_router
from routers.quality         import router as quality_router
from routers.wallboard       import router as wallboard_router

# ── App ────────────────────────────────────────────────────────
app = FastAPI(
    title       = "Platform — Toyota Boshoku Device India",
    description = "Manufacturing Execution System API",
    version     = "2.0.0",
    docs_url    = "/docs",
    redoc_url   = "/redoc",
)

# ── Static files ───────────────────────────────────────────────
try:
    app.mount("/static", StaticFiles(directory=os.path.dirname(__file__)), name="static")
except Exception:
    pass

# ── CORS ───────────────────────────────────────────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins     = ["*"],
    allow_credentials = True,
    allow_methods     = ["*"],
    allow_headers     = ["*"],
)

# ── Gzip ───────────────────────────────────────────────────────
# 2026-05-18 — LAN-access perf fix.  Realtime/submachines/history
# responses are pure JSON and compress 5-10x.  minimum_size=500 skips
# tiny replies (ok/health) where compression overhead doesn't pay.
# Removes the buffering symptom on remote-PC dashboard access.
app.add_middleware(GZipMiddleware, minimum_size=500)

# ── Routers ────────────────────────────────────────────────────
app.include_router(auth_router)
app.include_router(plants_router)
app.include_router(lines_router)
app.include_router(config_router)
app.include_router(poka_router)
app.include_router(status_router)
app.include_router(users_router)
app.include_router(zones_router)
app.include_router(npd_router)
app.include_router(submachines_router)
app.include_router(departments_router)
app.include_router(breakdowns_router)
app.include_router(machines_router)
app.include_router(breakdown_mail_router)
app.include_router(maintenance_kpi_router)
app.include_router(capa_router)
app.include_router(quality_router)
app.include_router(reports_router)
app.include_router(operators_router)
app.include_router(manpower_router)
app.include_router(store_router)
app.include_router(dispatch_router)
app.include_router(shift_calc_router)
app.include_router(kanban_router)
app.include_router(anything_wrong_router)
app.include_router(heijunka_router)
app.include_router(five_s_router)
app.include_router(pdca_router)
app.include_router(cms_sync_router)   # NF2/CMS bidirectional sync (loopback-only)
app.include_router(wallboard_router)  # 65" portrait wall-display feeders


@app.on_event("startup")
def _start_manpower_watcher():
    """Background worker for manpower alerts (unallocated + escalation)."""
    try:
        start_manpower_watcher()
    except Exception as exc:
        print(f"[MANPOWER-ALERT] failed to start: {exc}")


@app.on_event("startup")
def _start_kanban_watcher():
    """Background worker that auto-fires kanban window logs (12 PM,
    Shift A end, Shift B end) per line based on cycles produced."""
    try:
        start_kanban_watcher()
    except Exception as exc:
        print(f"[KANBAN] failed to start watcher: {exc}")


@app.on_event("startup")
def _start_reports_scheduler():
    """Background worker that mails the end-of-shift report 90 s after
    each shift end_time, to recipients configured per line in
    mes_report_email_config.  Defined in routers/reports.py."""
    try:
        start_reports_scheduler()
    except Exception as exc:
        print(f"[REPORT-SCHED] failed to start: {exc}")


# ── OEE drop alarm background worker ─────────────────────────────
@app.on_event("startup")
def _start_oee_alarm():
    """Email watchdog that fires when overall_oee stays below
    `threshold_pct` for `sustain_minutes` consecutive minutes per line."""
    try:
        from oee_alarm import start as start_oee_alarm
        start_oee_alarm()
    except Exception as exc:
        print(f"[OEE-ALARM] failed to start: {exc}")


# ── OEE alarm config endpoints (admin only) ──────────────────────
from pydantic import BaseModel as _BM_oee


class _OEEAlarmCfg(_BM_oee):
    line_id:          int
    threshold_pct:    float = 60.0
    sustain_minutes:  int   = 10
    cooldown_minutes: int   = 60
    to_addresses:     str   = ""
    cc_addresses:     str   = ""
    is_active:        bool  = True


@app.get("/api/oee-alarm")
def list_oee_alarms(user=Depends(get_current_user)):
    from oee_alarm import _ensure_table
    _ensure_table()
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT * FROM mes_oee_alarm_config ORDER BY line_id")
        return cur.fetchall()


@app.put("/api/oee-alarm")
def upsert_oee_alarm(body: _OEEAlarmCfg, admin=Depends(require_admin)):
    from oee_alarm import _ensure_table
    _ensure_table()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mes_oee_alarm_config
                (line_id, threshold_pct, sustain_minutes, cooldown_minutes,
                 to_addresses, cc_addresses, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (line_id) DO UPDATE
                SET threshold_pct    = EXCLUDED.threshold_pct,
                    sustain_minutes  = EXCLUDED.sustain_minutes,
                    cooldown_minutes = EXCLUDED.cooldown_minutes,
                    to_addresses     = EXCLUDED.to_addresses,
                    cc_addresses     = EXCLUDED.cc_addresses,
                    is_active        = EXCLUDED.is_active
        """, (body.line_id, body.threshold_pct, body.sustain_minutes,
              body.cooldown_minutes, body.to_addresses, body.cc_addresses,
              body.is_active))
        conn.commit()
    return {"ok": True}


# ── Startup migrations ─────────────────────────────────────────
@app.on_event("startup")
def run_migrations():
    """Apply any pending schema changes that are safe to run on every startup."""
    migrations = [
        "ALTER TABLE mes_lines ADD COLUMN IF NOT EXISTS current_shift_row_id INTEGER",
        "ALTER TABLE mes_lines ADD COLUMN IF NOT EXISTS ot_active_shift VARCHAR(10)",
        "ALTER TABLE mes_lines ADD COLUMN IF NOT EXISTS ot_start_a TIME",
        "ALTER TABLE mes_lines ADD COLUMN IF NOT EXISTS ot_end_a   TIME",
        "ALTER TABLE mes_lines ADD COLUMN IF NOT EXISTS ot_start_b TIME",
        "ALTER TABLE mes_lines ADD COLUMN IF NOT EXISTS ot_end_b   TIME",
        # NOTE: The three ALTER COLUMN TYPE statements for desired_value/value
        # were historical one-time migrations.  They are already applied in
        # production (columns are already VARCHAR).  Re-running on every
        # startup caused the whole backend to hang for 30+ s waiting for an
        # AccessExclusive lock that the running collector keeps holding, so
        # they were removed.  If a fresh DB needs them, run once manually.
        "ALTER TABLE mes_shift_configs ADD COLUMN IF NOT EXISTS ot_start_time TIME",
        "ALTER TABLE mes_shift_configs ADD COLUMN IF NOT EXISTS ot_end_time   TIME",
        # ── Machine display sequence (M-1, M-2, ...) — admin-assigned per
        # machine row.  Renders as the big "M-N" label on the Dashboard's
        # sub-machine tiles.  NULL means "no preference" → UI falls back to id.
        "ALTER TABLE mes_plc_configs ADD COLUMN IF NOT EXISTS machine_seq INTEGER",
        """
        CREATE TABLE IF NOT EXISTS mes_status_log (
            id          BIGSERIAL PRIMARY KEY,
            line_id     INTEGER        NOT NULL,
            record_date DATE           NOT NULL,
            shift_name  VARCHAR(20)    NOT NULL,
            status      VARCHAR(50)    NOT NULL,
            ts          TIMESTAMPTZ    NOT NULL,
            nowminfrac  DOUBLE PRECISION NOT NULL
        )
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_mes_status_log_line_date
            ON mes_status_log (line_id, record_date)
        """,

        # ── Departments (admin-managed list of departments) ───────────────
        # Used to label the "{DeptName} Panel" entry in the slide-nav for
        # department users.  Admin can add new departments anytime (e.g.
        # "Tool Room", "Stores") from Admin Panel → Departments.  `slug`
        # is the URL-safe identifier (lowercase, underscores) auto-derived
        # from `name` on create.
        """
        CREATE TABLE IF NOT EXISTS mes_departments (
            id          SERIAL PRIMARY KEY,
            name        VARCHAR(60) NOT NULL UNIQUE,
            slug        VARCHAR(40) NOT NULL UNIQUE,
            description TEXT,
            created_at  TIMESTAMP DEFAULT NOW(),
            updated_at  TIMESTAMP DEFAULT NOW()
        )
        """,
        # Seed Maintenance + Quality (only ones with a known interaction flow
        # right now — admin can add Tool Room / others later from the UI).
        """
        INSERT INTO mes_departments (name, slug, description) VALUES
            ('Maintenance', 'maintenance', 'Maintenance team — handles PY fail fixes and deviations.'),
            ('Quality',     'quality',     'Quality team — verifies fixes, approves deviations.')
        ON CONFLICT (slug) DO NOTHING
        """,
        # Each department user (role='department') belongs to exactly one
        # department.  Other roles ignore this column.  ON DELETE SET NULL
        # so deleting a department doesn't cascade-delete the users.
        "ALTER TABLE mes_admin ADD COLUMN IF NOT EXISTS department_id INTEGER REFERENCES mes_departments(id) ON DELETE SET NULL",

        # ── Maintenance Breakdown tracking ────────────────────────────────
        # Powers the Maintenance Dashboard:
        #   • ANDON live view = rows where ended_at IS NULL  (state='OPEN')
        #   • History table   = rows ended_at within last 2 days
        #   • Closure form    = JSON payload filled when ticket is closed
        #
        # `serial_in_shift` resets to 1 at every shift change for that line
        # (computed at insert-time from existing OPEN+RESOLVED rows whose
        # started_at falls in the same shift window).
        """
        CREATE TABLE IF NOT EXISTS mes_breakdowns (
            id                SERIAL PRIMARY KEY,
            line_id           INTEGER NOT NULL,
            zone_id           INTEGER,
            shift_name        VARCHAR(20),
            serial_in_shift   INTEGER,
            started_at        TIMESTAMP NOT NULL DEFAULT NOW(),
            ended_at          TIMESTAMP,
            -- 'OPEN'      → currently broken (shows on ANDON)
            -- 'RESOLVED'  → line back to running, closure form pending
            -- 'CLOSED'    → closure form submitted, ticket archived
            state             VARCHAR(20) NOT NULL DEFAULT 'OPEN',
            reason            TEXT,
            opened_by_user_id INTEGER,
            closed_by_user_id INTEGER,
            closure_data      JSONB,
            closed_at         TIMESTAMP,
            created_at        TIMESTAMP DEFAULT NOW(),
            updated_at        TIMESTAMP DEFAULT NOW()
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_breakdowns_state    ON mes_breakdowns (state) WHERE state IN ('OPEN','RESOLVED')",
        "CREATE INDEX IF NOT EXISTS idx_breakdowns_line     ON mes_breakdowns (line_id, started_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_breakdowns_zone     ON mes_breakdowns (zone_id, started_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_breakdowns_recent   ON mes_breakdowns (ended_at DESC) WHERE ended_at IS NOT NULL",

        # ── Split closure form into Production + Maintenance halves ────────
        # Production fills the upper half of the BREAK DOWN SLIP (line / zone /
        # machine / shift / category / received-time / problem reported).
        # Maintenance fills the lower half (problem observed / action / spares
        # / attended-by / signatures).  B/D Start/End Time + Date come from
        # the collector and are immutable for both sides.  closure_data stays
        # for backward-compat reads of existing rows.
        "ALTER TABLE mes_breakdowns ADD COLUMN IF NOT EXISTS production_data              JSONB",
        "ALTER TABLE mes_breakdowns ADD COLUMN IF NOT EXISTS production_filled_at         TIMESTAMP",
        "ALTER TABLE mes_breakdowns ADD COLUMN IF NOT EXISTS production_filled_by_user_id INTEGER",
        "ALTER TABLE mes_breakdowns ADD COLUMN IF NOT EXISTS maintenance_data             JSONB",
        "ALTER TABLE mes_breakdowns ADD COLUMN IF NOT EXISTS maintenance_filled_at        TIMESTAMP",
        "ALTER TABLE mes_breakdowns ADD COLUMN IF NOT EXISTS maintenance_filled_by_user_id INTEGER",

        # ── Machine master list (zone × line × machine_no → machine_name) ─
        # Powers the auto-fill in the Maintenance closure form: when the
        # user types a Machine No., the Machine Name auto-fetches from
        # this list.  Source-of-truth is NF2's zones.json (admin imports
        # it once via scripts/import_nf2_machines.py); the table is
        # maintainable from Admin Panel later if the user wants edits.
        """
        CREATE TABLE IF NOT EXISTS mes_machines (
            id           SERIAL PRIMARY KEY,
            source_id    VARCHAR(40),     -- 'machine_<id>' from NF2 (traceability)
            zone_name    VARCHAR(60)  NOT NULL,
            line_name    VARCHAR(60)  NOT NULL,
            machine_no   INTEGER      NOT NULL,
            machine_name VARCHAR(120) NOT NULL,
            is_active    BOOLEAN      NOT NULL DEFAULT TRUE,
            created_at   TIMESTAMP    DEFAULT NOW(),
            updated_at   TIMESTAMP    DEFAULT NOW(),
            UNIQUE (zone_name, line_name, machine_no)
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_mes_machines_lookup ON mes_machines (LOWER(zone_name), LOWER(line_name), machine_no)",
        "CREATE INDEX IF NOT EXISTS idx_mes_machines_line   ON mes_machines (LOWER(zone_name), LOWER(line_name))",

        # Optional override on mes_lines so admin can explicitly map a
        # MES line to its NF2 line name when fuzzy matching can't pick it
        # (e.g. when two NF2 lines share the same prefix).  Falls back
        # to fuzzy lookup if NULL.
        "ALTER TABLE mes_lines ADD COLUMN IF NOT EXISTS nf2_line_name VARCHAR(80)",

        # ── Breakdown mail escalation chain (Maintenance dept) ─────────────
        # Admin-managed levels of escalation emails for OPEN breakdowns.
        # Each level has its own delay_minutes (relative to the breakdown's
        # started_at), label, To, and Cc.  A background worker polls all
        # OPEN breakdowns every 30 s and fires each level's mail once when
        # its delay has elapsed — provided the breakdown is still OPEN.
        # If the line transitions back to RUNNING (collector stamps
        # ended_at), no further levels fire for that ticket.
        """
        CREATE TABLE IF NOT EXISTS mes_breakdown_mail_levels (
            id              SERIAL PRIMARY KEY,
            level_no        INTEGER NOT NULL,
            label           VARCHAR(120),
            delay_minutes   INTEGER NOT NULL DEFAULT 0,
            to_addresses    TEXT NOT NULL DEFAULT '',
            cc_addresses    TEXT NOT NULL DEFAULT '',
            is_active       BOOLEAN NOT NULL DEFAULT TRUE,
            created_at      TIMESTAMP DEFAULT NOW(),
            updated_at      TIMESTAMP DEFAULT NOW(),
            UNIQUE (level_no)
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_bdmail_levels_active ON mes_breakdown_mail_levels (is_active, level_no)",

        # Per-(breakdown, level) send log — UNIQUE constraint = idempotency.
        # Worker INSERTs after a successful send; existence of a row tells
        # the worker 'already sent, skip'.  Failed sends still get a row
        # with status='FAIL' so we don't retry forever — admin retries
        # manually via the test-send button if needed.
        """
        CREATE TABLE IF NOT EXISTS mes_breakdown_mail_log (
            id              SERIAL PRIMARY KEY,
            breakdown_id    INTEGER NOT NULL REFERENCES mes_breakdowns(id)               ON DELETE CASCADE,
            level_id        INTEGER NOT NULL REFERENCES mes_breakdown_mail_levels(id)    ON DELETE CASCADE,
            sent_at         TIMESTAMP NOT NULL DEFAULT NOW(),
            status          VARCHAR(20) NOT NULL DEFAULT 'OK',
            to_addresses    TEXT,
            cc_addresses    TEXT,
            error           TEXT,
            UNIQUE (breakdown_id, level_id)
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_bdmail_log_breakdown ON mes_breakdown_mail_log (breakdown_id, level_id)",

        # Seed 4 default levels (all inactive until admin fills addresses).
        """
        INSERT INTO mes_breakdown_mail_levels (level_no, label, delay_minutes, is_active)
        VALUES
          (1, 'Immediate alert (Maintenance Engineer)',  0,  FALSE),
          (2, 'Section Head escalation',                 10, FALSE),
          (3, 'HOD escalation',                          20, FALSE),
          (4, 'Plant Head escalation',                   30, FALSE)
        ON CONFLICT (level_no) DO NOTHING
        """,

        # ── Maintenance KPI targets ────────────────────────────────────────
        # Admin-set thresholds for each KPI.  line_id NULL = plant-wide
        # default; per-line overrides stored as separate rows.  direction
        # tells the UI whether higher (e.g. MTBF, Availability) or lower
        # (e.g. MTTR, downtime) is the goal — drives the pass/fail badge
        # colour on the KPI dashboard.
        """
        CREATE TABLE IF NOT EXISTS mes_kpi_targets (
            id           SERIAL PRIMARY KEY,
            kpi_key      VARCHAR(40)       NOT NULL,
            line_id      INTEGER,
            target_value DOUBLE PRECISION  NOT NULL,
            unit         VARCHAR(20),
            direction    VARCHAR(10) NOT NULL DEFAULT 'higher',
            is_active    BOOLEAN NOT NULL DEFAULT TRUE,
            created_at   TIMESTAMP DEFAULT NOW(),
            updated_at   TIMESTAMP DEFAULT NOW(),
            UNIQUE (kpi_key, line_id)
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_kpi_targets_lookup ON mes_kpi_targets (kpi_key, line_id)",

        # Seed sensible plant-wide defaults.
        """
        INSERT INTO mes_kpi_targets (kpi_key, line_id, target_value, unit, direction, is_active) VALUES
            ('mtbf_hours',         NULL, 100.0, 'hours',   'higher', TRUE),
            ('mttr_minutes',       NULL,  30.0, 'minutes', 'lower',  TRUE),
            ('availability_pct',   NULL,  95.0, '%',       'higher', TRUE),
            ('breakdowns_count',   NULL,  10.0, 'count',   'lower',  TRUE),
            ('total_downtime_min', NULL, 120.0, 'minutes', 'lower',  TRUE),
            ('pending_closures',   NULL,   0.0, 'count',   'lower',  TRUE)
        ON CONFLICT (kpi_key, line_id) DO NOTHING
        """,

        # ── CAPA (Corrective Action / Preventive Action) ───────────────────
        # Threshold table: admin sets monthly-sum + single-breakdown limits
        # at three scopes — global, per-line, per-(line, machine_no).  The
        # backend resolver walks Machine → Line → Global to pick the most
        # specific match.
        """
        CREATE TABLE IF NOT EXISTS mes_capa_thresholds (
            id                              SERIAL PRIMARY KEY,
            scope                           VARCHAR(10) NOT NULL DEFAULT 'GLOBAL',
            line_id                         INTEGER,
            machine_no                      VARCHAR(40),
            label                           VARCHAR(120),
            monthly_sum_minutes_limit       INTEGER NOT NULL DEFAULT 120,
            single_breakdown_minutes_limit  INTEGER NOT NULL DEFAULT 60,
            is_active                       BOOLEAN NOT NULL DEFAULT TRUE,
            created_at                      TIMESTAMP DEFAULT NOW(),
            updated_at                      TIMESTAMP DEFAULT NOW(),
            CHECK (
                (scope = 'GLOBAL'  AND line_id IS NULL AND machine_no IS NULL) OR
                (scope = 'LINE'    AND line_id IS NOT NULL AND machine_no IS NULL) OR
                (scope = 'MACHINE' AND line_id IS NOT NULL AND machine_no IS NOT NULL)
            )
        )
        """,
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_capa_thresh_global  ON mes_capa_thresholds (scope) WHERE scope='GLOBAL'",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_capa_thresh_line    ON mes_capa_thresholds (line_id) WHERE scope='LINE'",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_capa_thresh_machine ON mes_capa_thresholds (line_id, machine_no) WHERE scope='MACHINE'",

        # Pareto-CAPA cutoff: of the machines that breached the threshold,
        # the top-N % (by cumulative breakdown minutes) are MANDATED to
        # file a CAPA.  Configurable from the Admin Maintenance Panel —
        # admin sets one number (e.g. 80) and every Pareto chart in the
        # Maintenance dashboard re-cuts itself.  Stored on the GLOBAL
        # threshold row for simplicity (per-line override possible later).
        "ALTER TABLE mes_capa_thresholds ADD COLUMN IF NOT EXISTS pareto_pct INTEGER NOT NULL DEFAULT 80",
        # Seed the GLOBAL row if it doesn't exist (so /pareto-config GET
        # returns a sane default the first time the admin opens the page).
        """
        INSERT INTO mes_capa_thresholds (scope, monthly_sum_minutes_limit,
                                          single_breakdown_minutes_limit, pareto_pct)
        VALUES ('GLOBAL', 120, 60, 80)
        ON CONFLICT DO NOTHING
        """,

        # CAPA filings — one row per (machine, trigger).  Maintenance fills
        # the 8D-style payload into capa_data JSONB.  trigger_kind is
        # 'SINGLE_LIMIT' (a single breakdown crossed the per-event limit)
        # or 'MONTHLY_LIMIT' (the month-to-date sum crossed the monthly
        # limit).  status walks OPEN → IN_PROGRESS → CLOSED.
        """
        CREATE TABLE IF NOT EXISTS mes_capa (
            id                       SERIAL PRIMARY KEY,
            breakdown_id             INTEGER REFERENCES mes_breakdowns(id) ON DELETE SET NULL,
            trigger_kind             VARCHAR(20) NOT NULL,
            trigger_value_minutes    INTEGER,
            threshold_minutes        INTEGER,
            line_id                  INTEGER,
            line_name                VARCHAR(120),
            zone_id                  INTEGER,
            zone_name                VARCHAR(120),
            machine_no               VARCHAR(40),
            machine_name             VARCHAR(120),
            month_year               CHAR(7),
            status                   VARCHAR(20) NOT NULL DEFAULT 'OPEN',
            capa_data                JSONB,
            opened_by_user_id        INTEGER,
            opened_at                TIMESTAMP DEFAULT NOW(),
            closed_by_user_id        INTEGER,
            closed_at                TIMESTAMP,
            updated_at               TIMESTAMP DEFAULT NOW()
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_capa_status   ON mes_capa (status, opened_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_capa_machine  ON mes_capa (line_id, machine_no, opened_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_capa_month    ON mes_capa (month_year)",

        # Seed default GLOBAL threshold (120 min/month, 60 min single).
        """
        INSERT INTO mes_capa_thresholds
            (scope, label, monthly_sum_minutes_limit, single_breakdown_minutes_limit, is_active)
        VALUES ('GLOBAL', 'Default plant-wide thresholds', 120, 60, TRUE)
        ON CONFLICT DO NOTHING
        """,

        # ── Quality module ─────────────────────────────────────────────
        # Three workflows wired here:
        #
        #   1. mes_quality_verifications — when Maintenance closes a BD
        #      online, the ticket is auto-routed to a Quality verification
        #      queue.  Quality user clicks Approve / Deny.  No new form;
        #      Quality is just signing off on Maintenance's repair.
        #
        #   2. mes_quality_deviations — when Maintenance can't fix a PY
        #      within 24h they raise an Online Deviation.  Maintenance
        #      fills the upper half (Non-conformance / Root cause /
        #      Containment + Permanent CA), Quality Sec Head approves /
        #      rejects.  QA Head can grant extensions.  Mirrors the
        #      paper Deviation Form layout.
        #
        #   3. mes_quality_4m_changes — 4M Change Intimation Note.
        #      Production raises Part A, Quality fills Part B.
        #
        # PY Fail email escalation (L1/L2/L3 immediate/24h/48h) is
        # delivered through the existing mes_breakdown_mail_levels
        # pipeline — no new table needed there.
        """
        CREATE TABLE IF NOT EXISTS mes_quality_verifications (
            id                 SERIAL PRIMARY KEY,
            breakdown_id       INTEGER NOT NULL REFERENCES mes_breakdowns(id) ON DELETE CASCADE,
            line_id            INTEGER,
            line_name          VARCHAR(120),
            zone_id            INTEGER,
            zone_name          VARCHAR(120),
            machine_no         VARCHAR(40),
            machine_name       VARCHAR(120),
            -- PENDING → APPROVED | DENIED
            status             VARCHAR(20) NOT NULL DEFAULT 'PENDING',
            routed_at          TIMESTAMP DEFAULT NOW(),
            decided_at         TIMESTAMP,
            decided_by_user_id INTEGER,
            remarks            TEXT,
            created_at         TIMESTAMP DEFAULT NOW(),
            updated_at         TIMESTAMP DEFAULT NOW()
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_q_verif_status  ON mes_quality_verifications (status, routed_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_q_verif_bd      ON mes_quality_verifications (breakdown_id)",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_q_verif_bd ON mes_quality_verifications (breakdown_id)",

        """
        CREATE TABLE IF NOT EXISTS mes_quality_deviations (
            id                  SERIAL PRIMARY KEY,
            dev_no              VARCHAR(40) UNIQUE,            -- DEV-2026-0001
            breakdown_id        INTEGER REFERENCES mes_breakdowns(id) ON DELETE SET NULL,
            line_id             INTEGER,
            line_name           VARCHAR(120),
            zone_id             INTEGER,
            zone_name           VARCHAR(120),
            machine_no          VARCHAR(40),
            machine_name        VARCHAR(120),
            -- Header fields (mirror paper Deviation Form)
            category            VARCHAR(60),                   -- "Process" | "In-House" | etc.
            process_name        VARCHAR(160),
            process_no          VARCHAR(60),
            srv_no              VARCHAR(60),
            deviation_qty       INTEGER,
            deviation_upto_qty  INTEGER,
            deviation_upto_date DATE,
            initiated_by        VARCHAR(120),
            initiated_at        TIMESTAMP,
            reason              TEXT,
            -- Non-Conformance
            requirement         TEXT,
            observation         TEXT,
            -- Root Cause
            root_cause_occurrence TEXT,
            root_cause_detection  TEXT,
            potential_consequences TEXT,
            -- Sign-offs (Production HOD + Quality HOD)
            hod_production      VARCHAR(120),
            hod_production_note TEXT,
            hod_quality         VARCHAR(120),
            hod_quality_note    TEXT,
            -- Action plans (each row: {action, resp, deptt, tgt_date, approver, remarks})
            containment_actions JSONB DEFAULT '[]'::JSONB,
            permanent_actions   JSONB DEFAULT '[]'::JSONB,
            -- Extensions list (each row: {from_qty_date, to_qty_date, reason, hod_concerned, sign, hod_quality, hod_operation, status})
            extensions          JSONB DEFAULT '[]'::JSONB,
            closure_remarks     TEXT,
            hod_concerned_close VARCHAR(120),
            hod_quality_close   VARCHAR(120),
            -- Workflow: PENDING_QA → APPROVED → CLOSED  /  REJECTED  /  EXTENDED
            status              VARCHAR(20) NOT NULL DEFAULT 'PENDING_QA',
            raised_by_user_id   INTEGER,
            approved_by_user_id INTEGER,
            approved_at         TIMESTAMP,
            closed_at           TIMESTAMP,
            created_at          TIMESTAMP DEFAULT NOW(),
            updated_at          TIMESTAMP DEFAULT NOW()
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_q_dev_status  ON mes_quality_deviations (status, created_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_q_dev_line    ON mes_quality_deviations (line_id, created_at DESC)",

        """
        CREATE TABLE IF NOT EXISTS mes_quality_4m_changes (
            id                  SERIAL PRIMARY KEY,
            note_no             VARCHAR(40) UNIQUE,            -- 4M-2026-0001
            zone_id             INTEGER,
            zone_name           VARCHAR(120),
            line_id             INTEGER,
            line_name           VARCHAR(120),
            part_name           VARCHAR(160),
            model               VARCHAR(80),
            shift_name          VARCHAR(20),
            issue_date          DATE,
            start_batch_code    VARCHAR(80),
            end_batch_code      VARCHAR(80),
            originator_name     VARCHAR(120),
            originator_user_id  INTEGER,
            -- Changing point flags + details (Man / Machine / Material / Method / Tool / Others)
            changing_points     JSONB DEFAULT '{}'::JSONB,     -- {man:true, machine:false, ...}
            change_details      TEXT,
            -- Quality feedback (Part B)
            qa_engineer         VARCHAR(120),
            change_acceptance   VARCHAR(10),                   -- YES | NO
            ifm_required        VARCHAR(10),
            confirmation_marking VARCHAR(10),
            control_next_station VARCHAR(10),
            -- Date-wise OK/NG (each row: {sr, date, shift, ok_qty, ng_qty})
            qty_status_log      JSONB DEFAULT '[]'::JSONB,
            qty_produced        INTEGER,
            ng_qty              INTEGER,
            retroactive_check_status VARCHAR(10),              -- OK | NG
            comments            TEXT,
            termination_date    DATE,
            qa_sign             VARCHAR(120),
            -- OPEN → QUALITY_REVIEW → CLOSED
            status              VARCHAR(20) NOT NULL DEFAULT 'OPEN',
            created_by_user_id  INTEGER,
            closed_by_user_id   INTEGER,
            created_at          TIMESTAMP DEFAULT NOW(),
            updated_at          TIMESTAMP DEFAULT NOW()
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_q_4m_status   ON mes_quality_4m_changes (status, created_at DESC)",
    ]
    # Smart-skip: pre-check existing columns/tables so we don't even acquire
    # a lock on tables where the schema is already in the desired state.
    # This keeps backend startup fast even when the collector is holding
    # read locks on those tables.
    import re as _re
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT table_name, column_name FROM information_schema.columns
                WHERE table_schema='public'
            """)
            existing_cols = {(r[0], r[1]) for r in cur.fetchall()}
            cur.execute("""
                SELECT table_name FROM information_schema.tables WHERE table_schema='public'
            """)
            existing_tables = {r[0] for r in cur.fetchall()}
            cur.close()
    except Exception as e:
        print(f"[STARTUP] Schema introspection failed: {e}"); existing_cols = set(); existing_tables = set()

    def _is_needed(sql: str) -> bool:
        s = sql.strip().upper()
        # ALTER TABLE <t> ADD COLUMN IF NOT EXISTS <c>
        m = _re.match(r"ALTER TABLE\s+(\w+)\s+ADD COLUMN IF NOT EXISTS\s+(\w+)", s, _re.IGNORECASE)
        if m:
            return (m.group(1).lower(), m.group(2).lower()) not in existing_cols
        # CREATE TABLE IF NOT EXISTS <t>
        m = _re.match(r"CREATE TABLE IF NOT EXISTS\s+(\w+)", s, _re.IGNORECASE)
        if m:
            return m.group(1).lower() not in existing_tables
        # CREATE INDEX IF NOT EXISTS — cheap, let it run
        return True

    skipped = 0
    already  = 0
    for sql in migrations:
        if not _is_needed(sql):
            already += 1
            continue
        try:
            with get_conn() as conn:
                cur = conn.cursor()
                # CRITICAL: use SET LOCAL so the timeout vanishes with the
                # transaction. Plain SET would persist on the pooled connection
                # and poison every later query on that same handle.
                cur.execute("BEGIN")
                cur.execute("SET LOCAL lock_timeout = '3s'")
                cur.execute("SET LOCAL statement_timeout = '10s'")
                cur.execute(sql)
                cur.execute("COMMIT")   # LOCAL settings reset automatically
                cur.close()
        except Exception as m_exc:
            skipped += 1
            # Roll back so the connection is returned to the pool in a clean state.
            try:
                with get_conn() as _c:
                    _c.rollback()
            except Exception:
                pass
            print(f"[MIGRATION] skipped ({type(m_exc).__name__}): {str(m_exc).strip()[:120]}")
    print(f"[STARTUP] migrations: {already} already-applied, {skipped} skipped")

    # Belt-and-suspenders: explicitly reset timeouts on any already-pooled
    # connection that may have been poisoned by a previous instance.
    try:
        with get_conn() as conn:
            c = conn.cursor()
            c.execute("RESET lock_timeout")
            c.execute("RESET statement_timeout")
            conn.commit()
            c.close()
    except Exception:
        pass


# ── Health ─────────────────────────────────────────────────────
@app.get("/api/health")
def health():
    try:
        with get_conn() as conn:
            cur = dict_cursor(conn)
            cur.execute("SELECT COUNT(*) AS plants FROM mes_plants")
            row = cur.fetchone()
            return {"status": "ok", "plants": row["plants"], "version": "2.0.0"}
    except Exception as e:
        return {"status": "error", "detail": str(e)}


# ── Audit ──────────────────────────────────────────────────────

def write_audit(conn, *, action, entity_type=None, entity_id=None,
                details=None, user=None):
    """Idempotent helper: append one row to mes_audit_log.

    2026-05-18 — Centralised so every endpoint that wants an audit
    trail can call this with a single line.  `user` is the dict
    returned by get_current_user() (has id + username); if omitted,
    user_id/username land NULL (e.g. system-driven events).
    Never raises — audit must never block business logic.
    """
    try:
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO mes_audit_log
                   (action, entity_type, entity_id, details,
                    user_id, username)
               VALUES (%s, %s, %s, %s, %s, %s)""",
            (action,
             entity_type,
             int(entity_id) if entity_id is not None else None,
             details,
             (user or {}).get("id"),
             (user or {}).get("username"))
        )
    except Exception as exc:
        try: conn.rollback()
        except Exception: pass
        print(f"[AUDIT] write failed action={action}: {exc}")


@app.get("/api/audit")
def audit_log(
    limit:     int = 50,
    offset:    int = 0,
    date_from: str = None,
    date_to:   str = None,
    action:    str = None,
    username:  str = None,
    user=Depends(get_current_user)
):
    """Paged audit-log read.  Optional filters:
      • date_from / date_to (inclusive)
      • action  — exact match
      • username — filter to one user (NEW 2026-05-18)
    """
    with get_conn() as conn:
        cur    = dict_cursor(conn)
        where  = []
        params = []
        if date_from:
            where.append("created_at >= %s")
            params.append(date_from + " 00:00:00")
        if date_to:
            where.append("created_at <= %s")
            params.append(date_to + " 23:59:59")
        if action:
            where.append("action = %s")
            params.append(action)
        if username:
            where.append("username = %s")
            params.append(username)

        where_sql = ("WHERE " + " AND ".join(where)) if where else ""
        cur.execute(f"SELECT COUNT(*) as total FROM mes_audit_log {where_sql}", params)
        total = cur.fetchone()["total"]
        cur.execute(f"""
            SELECT * FROM mes_audit_log {where_sql}
            ORDER BY created_at DESC LIMIT %s OFFSET %s
        """, params + [limit, offset])
        return {
            "logs":     cur.fetchall(),
            "total":    total,
            "offset":   offset,
            "limit":    limit,
            "has_more": (offset + limit) < total,
        }


@app.get("/api/audit/actions")
def audit_actions(user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT DISTINCT action FROM mes_audit_log ORDER BY action")
        return [r["action"] for r in cur.fetchall()]


@app.get("/api/audit/users")
def audit_users(user=Depends(get_current_user)):
    """Return every user with their last login + 24-h activity count.

    2026-05-18 — Backs the "Users · Last Login" top card on the Audit
    page so admin can see at-a-glance who's actively using the system.
    Joins mes_admin (canonical user list) with a lateral aggregate of
    mes_audit_log for last-action time + 24-h count.
    """
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT u.id, u.username, u.role, u.last_login,
                   d.name AS department_name,
                   (SELECT COUNT(*) FROM mes_audit_log a
                     WHERE a.username = u.username
                       AND a.created_at >= NOW() - INTERVAL '24 hours') AS actions_24h,
                   (SELECT MAX(created_at) FROM mes_audit_log a
                     WHERE a.username = u.username) AS last_action_at
            FROM mes_admin u
            LEFT JOIN mes_departments d ON d.id = u.department_id
            ORDER BY (u.last_login IS NULL), u.last_login DESC, u.username
        """)
        return cur.fetchall()


# ── CMS camera-grid proxy ─────────────────────────────────────
# The CMS portal (New folder 2) requires its own JWT auth, so the Deep
# frontend can't call it directly. This endpoint logs in once, caches
# the token, and proxies /api/camera-grid responses.
import requests as _req

_CMS_BASE   = os.environ.get("CMS_BASE_URL", "http://127.0.0.1:5000").rstrip("/")
_CMS_USER   = os.environ.get("CMS_USER", "admin")
_CMS_PASS   = os.environ.get("CMS_PASS", "admin123")
_cms_token   = {"jwt": None, "ts": 0}

def _cms_auth_header():
    """Get a valid CMS JWT, logging in again if expired (every 7h)."""
    import time as _t
    if _cms_token["jwt"] and (_t.time() - _cms_token["ts"]) < 7 * 3600:
        return {"Authorization": f"Bearer {_cms_token['jwt']}"}
    try:
        r = _req.post(f"{_CMS_BASE}/api/auth/login",
                      json={"username": _CMS_USER, "password": _CMS_PASS}, timeout=5)
        if r.status_code == 200:
            data = r.json()
            tok = data.get("data", {}).get("token") or data.get("token")
            if tok:
                _cms_token["jwt"] = tok
                _cms_token["ts"]  = _t.time()
                return {"Authorization": f"Bearer {tok}"}
    except Exception as e:
        print(f"[CMS] Login failed: {e}")
    return {}

@app.get("/api/cms/camera-grid")
def cms_camera_grid(user=Depends(get_current_user)):
    """Proxy the CMS portal's /api/camera-grid with cached auth."""
    hdr = _cms_auth_header()
    if not hdr:
        raise HTTPException(502, "Cannot authenticate with CMS portal")
    try:
        r = _req.get(f"{_CMS_BASE}/api/camera-grid", headers=hdr, timeout=10)
        if r.status_code == 200:
            body = r.json()
            return body.get("data") if isinstance(body.get("data"), list) else body
        raise HTTPException(r.status_code, f"CMS error: {r.text[:200]}")
    except _req.RequestException as e:
        raise HTTPException(502, f"CMS unreachable: {e}")


@app.get("/api/cms/cameras")
def cms_cameras_list(user=Depends(get_current_user)):
    """Proxy the CMS portal's /api/masters/cameras list — returns all
    registered cameras so the admin UI can show them in an assign dropdown."""
    hdr = _cms_auth_header()
    if not hdr:
        raise HTTPException(502, "Cannot authenticate with CMS portal")
    try:
        r = _req.get(f"{_CMS_BASE}/api/masters/cameras", headers=hdr, timeout=10)
        if r.status_code == 200:
            body = r.json()
            return body.get("data") if isinstance(body.get("data"), list) else body
        raise HTTPException(r.status_code, f"CMS error: {r.text[:200]}")
    except _req.RequestException as e:
        raise HTTPException(502, f"CMS unreachable: {e}")


@app.patch("/api/cms/machines/{zone_id}/{line_id}/{machine_id}/camera")
async def cms_assign_camera(zone_id: str, line_id: str, machine_id: str,
                            request: Request, user=Depends(get_current_user)):
    """Proxy admin's 'Assign Camera' click to CMS portal. Frontend sends
    { camera_id: '<id>' } — we forward with the cached CMS JWT. This is the
    ONE endpoint that links a machine to a camera; after this the recorder
    and Fullscreen video playback all work automatically."""
    if user.get("role") not in ("admin", "zone"):
        raise HTTPException(403, "Only admins can assign cameras")
    hdr = _cms_auth_header()
    if not hdr:
        raise HTTPException(502, "Cannot authenticate with CMS portal")
    try:
        body = await request.json()
    except Exception:
        body = {}
    try:
        r = _req.patch(
            f"{_CMS_BASE}/api/masters/machines/{zone_id}/{line_id}/{machine_id}/camera",
            headers={**hdr, "Content-Type": "application/json"},
            json={"camera_id": str(body.get("camera_id") or "").strip()},
            timeout=10,
        )
        if 200 <= r.status_code < 300:
            return r.json() if r.content else {"ok": True}
        raise HTTPException(r.status_code, f"CMS error: {r.text[:200]}")
    except _req.RequestException as e:
        raise HTTPException(502, f"CMS unreachable: {e}")


# ── Ping check (TCP connect test for camera/device IPs) ──────
@app.get("/api/ping")
def ping_host(ip: str, port: int = 554):
    """TCP connect test. Returns {ok: true/false, ms: latency}.
    Used by admin Camera List page to show online/offline status."""
    import socket, time as _t
    try:
        t0 = _t.time()
        s = socket.create_connection((ip, port), timeout=3)
        ms = round((_t.time() - t0) * 1000)
        s.close()
        return {"ok": True, "ms": ms}
    except Exception:
        return {"ok": False, "ms": 0}


# ── Export production data ─────────────────────────────────────
@app.get("/api/export/data")
def export_data(
    line_id:   int,
    date_from: str,
    date_to:   str,
    format:    str = "xlsx",
    user=Depends(get_current_user)
):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT l.*, p.plant_name, z.zone_name
            FROM mes_lines l
            JOIN mes_plants p ON p.id = l.plant_id
            LEFT JOIN mes_zones z ON z.id = l.zone_id
            WHERE l.id = %s
        """, (line_id,))
        line = cur.fetchone()
        if not line:
            raise HTTPException(404, "Line not found")
        table = line["db_table_name"]
        cur.execute(f"""
            SELECT * FROM {table}
            WHERE record_date BETWEEN %s AND %s
            ORDER BY record_date, shift_name
        """, (date_from, date_to))
        records = cur.fetchall()
        if not records:
            raise HTTPException(404, "No data found for selected date range")
        cur.execute("""
            SELECT slot_label, db_column_prefix, shift_name, slot_order
            FROM mes_hourly_slots
            WHERE line_id = %s
            ORDER BY shift_name, slot_order
        """, (line_id,))
        slots = cur.fetchall()

    slot_labels   = [s["slot_label"]        for s in slots]
    slot_prefixes = {s["slot_label"]: s["db_column_prefix"] for s in slots}

    summary_cols = [
        "record_date", "shift_name", "ok_count", "ng_count",
        "shift_plan", "shift_plan_completed",
        "overall_oee", "availability", "performance", "quality_oee",
        "oee_grade", "operating_status",
        "loss_breakdown_seconds", "loss_quality_seconds",
        "loss_setup_seconds", "loss_material_seconds",
        "loss_others_seconds", "loss_speed_seconds",
        "loss_change_over_seconds",
        "ct_avg_20", "min_ct", "max_ct",
    ]
    summary_headers = [
        "Date", "Shift", "OK Count", "NG Count",
        "Shift Plan", "Plan Completed",
        "Overall OEE (%)", "Availability (%)", "Performance (%)", "Quality (%)",
        "OEE Grade", "Operating Status",
        "Loss Breakdown (s)", "Loss Quality (s)",
        "Loss Setup (s)", "Loss Material (s)",
        "Loss Others (s)", "Loss Speed (s)", "Loss Change Over (s)",
        "Avg Cycle Time (s)", "Min CT (s)", "Max CT (s)",
    ]
    hourly_plan_headers   = [f"{lbl} Plan"          for lbl in slot_labels]
    hourly_actual_headers = [f"{lbl} Actual"         for lbl in slot_labels]
    hourly_var_headers    = [f"{lbl} Variance"       for lbl in slot_labels]
    hourly_eff_headers    = [f"{lbl} Efficiency (%)" for lbl in slot_labels]
    all_headers = (summary_headers + hourly_plan_headers +
                   hourly_actual_headers + hourly_var_headers + hourly_eff_headers)

    rows = []
    for r in records:
        row = [r.get(col, "") for col in summary_cols]
        for lbl in slot_labels:
            pfx = slot_prefixes[lbl]
            row.append(r.get(f"{pfx}_plan", 0) or 0)
        for lbl in slot_labels:
            pfx = slot_prefixes[lbl]
            row.append(r.get(f"{pfx}_actual", 0) or 0)
        for lbl in slot_labels:
            pfx = slot_prefixes[lbl]
            row.append(r.get(f"{pfx}_variance", 0) or 0)
        for lbl in slot_labels:
            pfx    = slot_prefixes[lbl]
            plan   = r.get(f"{pfx}_plan",   0) or 0
            actual = r.get(f"{pfx}_actual", 0) or 0
            row.append(round((actual / plan * 100), 1) if plan > 0 else 0)
        rows.append(row)

    filename = f"{line['line_code']}_{date_from}_to_{date_to}"

    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(500, "openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Production Data"
        ws.merge_cells(f"A1:{get_column_letter(len(all_headers))}1")
        c = ws["A1"]
        c.value     = f"Production Report — {line['line_name']} ({line['plant_name']}) | {date_from} to {date_to}"
        c.font      = Font(bold=True, size=12, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1E40AF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A2:{get_column_letter(len(all_headers))}2")
        c = ws["A2"]
        c.value     = f"Zone: {line.get('zone_name','—')} | Line: {line['line_code']} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        c.font      = Font(size=10, color="475569")
        c.fill      = PatternFill("solid", fgColor="EFF6FF")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20
        summary_end = len(summary_headers)
        thin_white  = Side(style="thin", color="FFFFFF")
        thin_gray   = Side(style="thin", color="E2E8F0")
        for ci, header in enumerate(all_headers, 1):
            c           = ws.cell(row=3, column=ci, value=header)
            c.font      = Font(bold=True, size=10, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill      = PatternFill("solid", fgColor="1E40AF" if ci <= summary_end else "1E3A8A")
            c.border    = Border(left=thin_white, right=thin_white, top=thin_white, bottom=thin_white)
        ws.row_dimensions[3].height = 36
        for ri, row in enumerate(rows, 4):
            for ci, val in enumerate(row, 1):
                c           = ws.cell(row=ri, column=ci, value=val)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font      = Font(size=10)
                c.border    = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
                if ri % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="F8FAFC")
                if ci == 7 and isinstance(val, (int, float)):
                    color = "16A34A" if val >= 85 else ("D97706" if val >= 65 else "DC2626")
                    c.font = Font(size=10, color=color, bold=True)
        for ci in range(1, len(all_headers) + 1):
            col_letter = get_column_letter(ci)
            max_len = max(
                len(str(ws.cell(row=r, column=ci).value or ""))
                for r in range(1, len(rows) + 4)
            )
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 28)
        ws.freeze_panes = "A4"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )

    elif format == "csv":
        buf = io.StringIO()
        w   = csv.writer(buf)
        w.writerow([f"Production Report — {line['line_name']} ({line['plant_name']})"])
        w.writerow([f"Zone: {line.get('zone_name','—')} | Date: {date_from} to {date_to} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        w.writerow([])
        w.writerow(all_headers)
        for row in rows:
            w.writerow(row)
        buf.seek(0)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    raise HTTPException(400, "Invalid format — use xlsx or csv")


# ── HTML pages ─────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def serve_admin():
    path = os.path.join(os.path.dirname(__file__), "admin.html")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    return HTMLResponse("<h2>Backend running — <a href='/docs'>API docs</a></h2>")


@app.get("/fullscreen.html", response_class=HTMLResponse)
def serve_fullscreen():
    path = os.path.join(os.path.dirname(__file__), "fullscreen.html")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    return HTMLResponse("Fullscreen page not found", status_code=404)


# ── AI Chat ────────────────────────────────────────────────────
# ── AI schema cache (loaded once, refreshed every 5 min) ──────────────────
_AI_SCHEMA_CACHE = {"ts": 0.0, "prompt": "", "lines": []}

def _get_ai_schema_info():
    """Build the schema block for the AI system prompt. Cached for 5 minutes
    so we don't hammer the DB on every chat message."""
    import time as _t
    if _AI_SCHEMA_CACHE["prompt"] and (_t.time() - _AI_SCHEMA_CACHE["ts"]) < 300:
        return _AI_SCHEMA_CACHE["prompt"], _AI_SCHEMA_CACHE["lines"]
    try:
        with get_conn() as conn:
            cur = dict_cursor(conn)
            cur.execute("""
                SELECT l.id, l.line_name, l.db_table_name,
                       p.plant_name, z.zone_name, l.collector_status
                FROM mes_lines l
                JOIN mes_plants p ON p.id = l.plant_id
                LEFT JOIN mes_zones z ON z.id = l.zone_id
                ORDER BY l.line_name
            """)
            lines = cur.fetchall()
            schema_info = "LINES:\n"
            for l in lines:
                schema_info += f"  {l['line_name']} -> {l['db_table_name']}\n"
            schema_info += (
                "\nDASHBOARD COLS: record_date, shift_name, ok_count, ng_count, "
                "shift_plan, shift_plan_completed, overall_oee, availability, "
                "performance, quality_oee, operating_status, cycle_time_actual, "
                "ct_avg_20, loss_breakdown_seconds, loss_quality_seconds, "
                "loss_setup_seconds, loss_material_seconds, loss_others_seconds, "
                "loss_speed_seconds, loss_change_over_seconds, total_loss, "
                "current_model_name, current_model_number\n"
                "POKA: mes_poka_yoke_events(line_id, rule_type, alert_level, "
                "detected_at, acknowledged, plc_value, context_json); "
                "mes_poka_yoke_rules(line_id, poka_yoke_name, bit, value)\n"
                "OTHER: mes_lines, mes_plants, mes_zones, mes_model_mappings, "
                "mes_audit_log, mes_py_master, mes_py_model_master, mes_py_assignments"
            )
            _AI_SCHEMA_CACHE["prompt"] = schema_info
            _AI_SCHEMA_CACHE["lines"]  = lines
            _AI_SCHEMA_CACHE["ts"]     = _t.time()
            return schema_info, lines
    except Exception as e:
        return f"(schema load error: {e})", []


@app.post("/api/ai/chat")
async def ai_chat(request: Request, user=Depends(get_current_user)):
    """Fast AI chat — Haiku model, cached schema, smaller token budget, lower
    tool-iteration cap.  Typical response 2-10 s (was 1-3 min)."""
    import anthropic

    body    = await request.json()
    message = body.get("message", "").strip()
    context = body.get("context", {})
    history = body.get("history", [])

    if not message:
        raise HTTPException(400, "Message required")

    today     = datetime.now().date()
    yesterday = today - timedelta(days=1)
    schema_info, _ = _get_ai_schema_info()

    # Compact prompt — fewer tokens = much faster response.
    system_prompt = (
        f"You are the production-data assistant for Toyota Boshoku Device India, Bawal.\n"
        f"Today={today}  Yesterday={yesterday}  Page={context.get('page','Dashboard')}\n\n"
        f"{schema_info}\n\n"
        "RULES:\n"
        "- ALWAYS use run_query for data; never guess numbers.\n"
        "- SELECT only.  Use record_date='YYYY-MM-DD' and shift_name='A'/'B'.\n"
        "- Be brief: one-line answer + short table or bullets. No preamble.\n"
        "- Format OEE as %, times as HH:MM:SS.\n"
        "- Prefer ONE well-written query over many small ones."
    )

    tools = [{
        "name": "run_query",
        "description": "Execute a SQL SELECT on the PostgreSQL DB.",
        "input_schema": {
            "type": "object",
            "properties": {
                "sql":         {"type": "string"},
                "description": {"type": "string"},
            },
            "required": ["sql"],
        },
    }]

    def execute_query(sql: str) -> str:
        sql = sql.strip()
        if not sql.upper().startswith("SELECT"):
            return "Error: only SELECT allowed."
        try:
            with get_conn() as conn:
                cur  = dict_cursor(conn)
                cur.execute("SET LOCAL statement_timeout = '8s'")
                cur.execute(sql)
                rows = cur.fetchall()
                if not rows:
                    return "No rows."
                if len(rows) == 1 and len(rows[0]) == 1:
                    return str(list(rows[0].values())[0])
                headers = list(rows[0].keys())
                out     = [" | ".join(str(h) for h in headers)]
                for row in rows[:20]:     # was 50 — tighter context
                    out.append(" | ".join(
                        str(v) if v is not None else "-" for v in row.values()))
                if len(rows) > 20:
                    out.append(f"... +{len(rows)-20} more rows")
                return "\n".join(out)
        except Exception as e:
            return f"Query error: {str(e)[:200]}"

    try:
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        msgs   = []
        # Shorter history — only last 4 turns instead of 8. Each history
        # message can be big, so this drops a lot of input tokens.
        for h in history[-4:]:
            msgs.append({"role": h["role"], "content": h["content"]})
        msgs.append({"role": "user", "content": message})

        # Model choice:
        #  - AI_MODEL env override wins (pick any Anthropic model ID).
        #  - Default = Haiku 4.5 — fastest tier available on this account
        #    (this account doesn't have Haiku 3.5 access). Set
        #    AI_MODEL=claude-sonnet-4-6 for heavier reasoning at +latency.
        model = os.getenv("AI_MODEL", "claude-haiku-4-5")

        # Cap tool-use loops to 3 instead of 5. Each iteration = full round-trip
        # to Anthropic + DB, which is the biggest source of latency.
        max_iterations = 3
        iteration      = 0
        while iteration < max_iterations:
            iteration += 1
            resp = client.messages.create(
                model      = model,
                max_tokens = 1024,         # was 2048 — shorter outputs
                system     = system_prompt,
                tools      = tools,
                messages   = msgs,
            )
            if resp.stop_reason == "end_turn":
                reply = ""
                for block in resp.content:
                    if hasattr(block, "text"):
                        reply += block.text
                return {"reply": reply.strip(), "provider": model.split("-")[1] if "-" in model else "claude"}
            if resp.stop_reason == "tool_use":
                msgs.append({"role": "assistant", "content": resp.content})
                tool_results = []
                for block in resp.content:
                    if block.type == "tool_use":
                        sql         = block.input.get("sql", "")
                        description = block.input.get("description", "")
                        print(f"[AI] {description or sql[:60]}")
                        result = execute_query(sql)
                        tool_results.append({
                            "type":        "tool_result",
                            "tool_use_id": block.id,
                            "content":     result,
                        })
                msgs.append({"role": "user", "content": tool_results})
                continue
            break

        return {"reply": "Unable to complete analysis. Please rephrase.", "provider": "claude"}

    except Exception as e:
        print(f"[AI] Error: {e}")
        raise HTTPException(500, f"AI error: {str(e)}")


# ── Import: Poka Yoke rules (line-scoped, from Excel) ──────────
@app.post("/api/import/poka-yoke")
async def import_poka_yoke(request: Request, user=Depends(get_current_user)):
    body       = await request.json()
    line_id    = body.get("line_id")
    rows       = body.get("rows", [])
    sheet_name = body.get("sheet_name", "")

    if not line_id:
        raise HTTPException(400, "line_id required")

    with get_conn() as conn:
        cur      = dict_cursor(conn)
        inserted = 0
        skipped  = 0
        errors   = []

        for r in rows:
            py_no   = str(r.get("POKA-YOKE No",  "") or "").strip()
            side    = str(r.get("SIDE",           "") or "all").strip().upper()
            name    = str(r.get("POKA-YOKE NAME", "") or "").strip()
            model   = str(r.get("MODEL",          "") or "all").strip()
            bit     = str(r.get("BIT",            "") or "").strip().upper()
            value   = r.get("VALUE", 1)
            machine = str(r.get("MACHINE NAME",   "") or "").strip()

            if not name or name.upper() in ("POKA-YOKE NAME", ""):
                skipped += 1
                continue

            if bit and not bit.startswith("D"):
                bit = f"D{bit}"

            if side in ("L", "LH"):   side = "LH"
            elif side in ("R", "RH"): side = "RH"
            else:                      side = "ALL"

            try:
                trig = int(value) if value not in (None, "") else 1
            except (ValueError, TypeError):
                trig = 1

            try:
                cur.execute("""
                    INSERT INTO mes_poka_yoke_rules (
                        line_id, poka_yoke_no, side, poka_yoke_name,
                        model, bit, value, machine_name,
                        sheet_name, alert_level, is_active
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'WARNING',true)
                """, (
                    line_id, py_no or None, side, name,
                    model or "all", bit or None, trig,
                    machine or None, sheet_name or None,
                ))
                inserted += 1
            except Exception as e:
                skipped += 1
                errors.append(str(e))

        cur.execute("""
            INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
            VALUES ('POKA_YOKE_BULK_IMPORT', 'line', %s, %s)
        """, (line_id, f"Sheet '{sheet_name}': inserted={inserted} skipped={skipped}"))

    return {"ok": True, "inserted": inserted, "skipped": skipped,
            "errors": errors[:5], "sheet": sheet_name}


# ── Import: PY assignments (matrix) ───────────────────────────
@app.post("/api/import/py-assignments")
async def import_py_assignments(request: Request, user=Depends(get_current_user)):
    body     = await request.json()
    rows     = body.get("rows", [])
    inserted = 0
    skipped  = 0

    with get_conn() as conn:
        cur = dict_cursor(conn)
        for r in rows:
            py_no      = str(r.get("pyNo",      r.get("Poka Yoke No",  ""))).strip()
            model_name = str(r.get("modelName", r.get("Model Name",   ""))).strip()
            if not py_no or not model_name:
                skipped += 1
                continue
            try:
                cur.execute("""
                    INSERT INTO mes_py_assignments
                        (py_no, py_name, side, model_type, model_name,
                         model_series, old_model_no, d_bit, desired_value, machine_name)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT DO NOTHING
                """, (
                    py_no,
                    str(r.get("pyName",        r.get("Poka Yoke Name", ""))).strip(),
                    str(r.get("typeSide",       r.get("Type Side",      "ALL"))).strip().upper(),
                    str(r.get("modelType",      r.get("Model Type",     ""))).strip(),
                    model_name,
                    str(r.get("modelSeries",    r.get("Model",          ""))).strip(),
                    str(r.get("oldModelNo",     r.get("Old Model No",   ""))).strip(),
                    str(r.get("dBit",           r.get("D bit From PLC", ""))).strip() or None,
                    r.get("desiredValue",       r.get("Desired Value")),
                    str(r.get("machineFixture", r.get("Machine/Fixture",""))).strip() or None,
                ))
                inserted += 1
            except Exception:
                skipped += 1

    return {"ok": True, "inserted": inserted, "skipped": skipped}


# ── Import: PLC model mappings ─────────────────────────────────
@app.post("/api/import/models")
async def import_models(request: Request, user=Depends(get_current_user)):
    body    = await request.json()
    line_id = body.get("line_id")
    rows    = body.get("rows", [])

    if not line_id:
        raise HTTPException(400, "line_id required")

    with get_conn() as conn:
        cur      = dict_cursor(conn)
        inserted = 0
        skipped  = 0
        for r in rows:
            try:
                model_number = int(r.get("Model Number", 0) or 0)
                model_name   = str(r.get("Model Name", "")).strip()
            except (ValueError, TypeError):
                skipped += 1
                continue
            if not model_number or not model_name:
                skipped += 1
                continue
            try:
                cur.execute("""
                    INSERT INTO mes_model_mappings (line_id, model_number, model_name)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (line_id, model_number)
                    DO UPDATE SET model_name = EXCLUDED.model_name
                """, (line_id, model_number, model_name))
                inserted += 1
            except Exception:
                skipped += 1
        cur.execute("""
            INSERT INTO mes_audit_log (action, entity_type, entity_id, details)
            VALUES ('MODEL_BULK_IMPORT', 'line', %s, %s)
        """, (line_id, f"Imported {inserted} models, skipped {skipped}"))

    return {"ok": True, "inserted": inserted, "skipped": skipped}


# ── Poka Yoke lookup ───────────────────────────────────────────
@app.get("/api/lines/{line_id}/poka-yoke-lookup")
def poka_yoke_lookup(line_id: int, user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT id, poka_yoke_no, side, poka_yoke_name,
                   model, bit, value, machine_name,
                   sheet_name, alert_level
            FROM mes_poka_yoke_rules
            WHERE line_id = %s AND is_active = true AND bit IS NOT NULL
            ORDER BY bit
        """, (line_id,))
        return cur.fetchall()


# ── Poka Yoke template export ──────────────────────────────────
@app.get("/api/export/template/poka-yoke")
def poka_yoke_template(user=Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Poka Yoke Rules"

    headers  = ["POKA-YOKE No", "SIDE", "POKA-YOKE NAME", "MODEL", "BIT", "VALUE", "MACHINE NAME"]
    examples = [
        ["TBDI-PE-PY-6041", "LH",  "detect harness brkt pop rivet miss", "all", "D401", 1, "HARNESS BRKT & POP RIVETING"],
        ["TBDI-PE-PY-6042", "RH",  "detect data matrix miss",             "all", "D402", 1, "FINAL INSPECTION MACHINE"],
        ["TBDI-PE-PY-6043", "LH",  "detect circlip miss LH",              "all", "D403", 2, "FINAL INSPECTION MACHINE"],
        ["TBDI-PE-PY-6044", "RH",  "detect circlip miss RH",              "all", "D404", 1, "FINAL INSPECTION MACHINE"],
    ]
    notes = [
        ["SIDE values →",    "LH / RH / ALL"],
        ["VALUE meanings →", "0 = Pass  |  1 = Off/Fault  |  2 = On/Fault"],
        ["BIT format →",     "D401, D402 ... (must start with D)"],
        ["MODEL examples →", "all / except yjc / yjc only"],
    ]

    for ci, h in enumerate(headers, 1):
        c           = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1E40AF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = 26
    ws.row_dimensions[1].height = 28

    for ri, row in enumerate(examples, 2):
        for ci, val in enumerate(row, 1):
            c           = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="EFF6FF")

    note_start = len(examples) + 3
    ws.cell(row=note_start - 1, column=1, value="── REFERENCE NOTES ──").font = Font(
        bold=True, size=10, color="64748B")
    for ni, (label, note) in enumerate(notes, note_start):
        ws.cell(row=ni, column=1, value=label).font = Font(bold=True, size=9, color="DC2626")
        ws.cell(row=ni, column=2, value=note).font  = Font(size=9, color="475569")
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="poka_yoke_template.xlsx"'},
    )


# ── Models template export ─────────────────────────────────────
@app.get("/api/export/template/models")
def models_template(user=Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model Mappings"

    headers  = ["Model Number", "Model Name"]
    examples = [[1, "YNC-SS Type A"], [2, "YNC-SS Type B"], [3, "YNC-SS Type C"]]

    for ci, h in enumerate(headers, 1):
        c           = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1E40AF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = 24

    for ri, row in enumerate(examples, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    ws.cell(row=6, column=1,
            value="Model Number must be an integer matching PLC word value").font = Font(
        color="DC2626", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="models_template.xlsx"'},
    )


# ── NG Events ─────────────────────────────────────────────────
@app.get("/api/lines/{line_id}/ng-events")
def ng_events(
    line_id:     int,
    record_date: str,
    shift_name:  str,
    user=Depends(get_current_user)
):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT db_table_name, plc_config FROM mes_lines WHERE id = %s", (line_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Line not found")
        table    = row["db_table_name"]
        plc_cfg  = row.get("plc_config") or {}
        ideal_ct = float(plc_cfg.get("ideal_cycle_time", 15))

        cur.execute(f"""
            SELECT
                current_model_number, current_model_name,
                ct1,ct2,ct3,ct4,ct5,ct6,ct7,ct8,ct9,ct10,
                ct11,ct12,ct13,ct14,ct15,ct16,ct17,ct18,ct19,ct20,
                ct_avg_20, min_ct, max_ct, std_dev_ct,
                ng_count, ok_count
            FROM {table}
            WHERE record_date = %s AND shift_name = %s
            ORDER BY updated_at DESC LIMIT 1
        """, (record_date, shift_name))
        rec = cur.fetchone()
        if not rec:
            raise HTTPException(404, "No data found")

    cts   = [float(rec[f"ct{i}"] or 0) for i in range(1, 21)]
    valid = [ct for ct in cts if ct > 0]
    above = [{"cycle": i+1, "ct": ct} for i, ct in enumerate(cts)
             if ct > ideal_ct and ct > 0]

    return {
        "events": [{
            "model_number":           rec["current_model_number"] or 0,
            "model_name":             rec["current_model_name"]   or f"Model #{rec['current_model_number']}",
            "ct_avg":                 float(rec["ct_avg_20"] or 0),
            "min_ct":                 float(rec["min_ct"]    or 0),
            "max_ct":                 float(rec["max_ct"]    or 0),
            "std_dev":                float(rec["std_dev_ct"] or 0),
            "cycles_above_threshold": above,
            "total_above":            len(above),
            "all_cts":                valid,
            "ng_count":               rec["ng_count"] or 0,
            "ok_count":               rec["ok_count"] or 0,
        }],
        "ideal_ct":    ideal_ct,
        "total_above": len(above),
    }


# ══════════════════════════════════════════════════════════════
# POKA YOKE MATRIX — Master, Models, Assignments, Bulk Import
# ══════════════════════════════════════════════════════════════

class BulkImportBody(BaseModel):
    sheet:   str
    rows:    List[Any]
    col_map: dict = {}


def _col(row, system_key, col_map, default=""):
    excel_key = col_map.get(system_key, system_key)
    val = row.get(excel_key, row.get(system_key, default))
    s = str(val).strip() if val is not None else ""
    return "" if s in ("nan", "None", "") else s


def _int(row, system_key, col_map, default=None):
    v = _col(row, system_key, col_map)
    try:
        return int(float(v)) if v else default
    except:
        return default


# ── PY MODEL MASTER ───────────────────────────────────────────
@app.get("/api/poka-yoke/models/")
def py_get_models(user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT * FROM mes_py_model_master WHERE is_active=true ORDER BY model_name")
        rows = [dict(r) for r in cur.fetchall()]
        for r in rows:
            r["type"]       = r.get("model_type", "")
            r["oldModelNo"] = r.get("old_model_no", "")
            r["model"]      = r.get("series", "")
            r["modelName"]  = r.get("model_name", "")
        return rows


@app.post("/api/poka-yoke/models/")
def py_add_model(body: dict, user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            INSERT INTO mes_py_model_master (model_name, model_type, old_model_no, series)
            VALUES (%s,%s,%s,%s) RETURNING *
        """, (body["modelName"], body.get("type"), body.get("oldModelNo"), body.get("model")))
        return dict(cur.fetchone())


@app.put("/api/poka-yoke/models/{mid}")
def py_update_model(mid: int, body: dict, user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            UPDATE mes_py_model_master
            SET model_name=%s, model_type=%s, old_model_no=%s, series=%s
            WHERE id=%s RETURNING *
        """, (body["modelName"], body.get("type"), body.get("oldModelNo"), body.get("model"), mid))
        return dict(cur.fetchone())


@app.delete("/api/poka-yoke/models/{mid}")
def py_delete_model(mid: int, user=Depends(get_current_user)):
    with get_conn() as conn:
        dict_cursor(conn).execute(
            "UPDATE mes_py_model_master SET is_active=false WHERE id=%s", (mid,))
    return {"ok": True}


# ── PY MASTER ─────────────────────────────────────────────────
@app.get("/api/poka-yoke/master/")
def py_get_master(user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT * FROM mes_py_master WHERE is_active=true ORDER BY py_no")
        rows = [dict(r) for r in cur.fetchall()]
        for r in rows:
            r["typeSide"]       = r.get("side", "ALL")
            r["dBit"]           = r.get("bit", "")
            r["desiredValue"]   = r.get("desired_value")
            r["machineFixture"] = r.get("machine_name", "")
            r["modelType"]      = r.get("model_type", "")
        return rows


@app.post("/api/poka-yoke/master/")
def py_add_master(body: dict, user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            INSERT INTO mes_py_master
                (py_no, description, model_type, side, bit, desired_value, machine_name)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (
            body["pyNo"], body.get("description"), body.get("modelType"),
            body.get("typeSide", "ALL"), body.get("dBit"),
            body.get("desiredValue") if body.get("desiredValue") not in ("", None) else None,
            body.get("machineFixture")
        ))
        return dict(cur.fetchone())


@app.put("/api/poka-yoke/master/{pid}")
def py_update_master(pid: int, body: dict, user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            UPDATE mes_py_master
            SET description=%s, model_type=%s, side=%s,
                bit=%s, desired_value=%s, machine_name=%s
            WHERE id=%s RETURNING *
        """, (
            body.get("description"), body.get("modelType"),
            body.get("typeSide", "ALL"), body.get("dBit"),
            body.get("desiredValue") if body.get("desiredValue") not in ("", None) else None,
            body.get("machineFixture"), pid
        ))
        return dict(cur.fetchone())


@app.delete("/api/poka-yoke/master/{pid}")
def py_delete_master(pid: int, user=Depends(get_current_user)):
    with get_conn() as conn:
        dict_cursor(conn).execute(
            "UPDATE mes_py_master SET is_active=false WHERE id=%s", (pid,))
    return {"ok": True}


# ── ASSIGNMENTS ───────────────────────────────────────────────
@app.get("/api/poka-yoke/assignments/")
def py_get_assignments(user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT * FROM mes_py_assignments ORDER BY model_name, py_no")
        rows = [dict(r) for r in cur.fetchall()]
        for r in rows:
            r["typeSide"]       = r.get("side", "")
            r["dBit"]           = r.get("d_bit", "")
            r["desiredValue"]   = r.get("desired_value")
            r["machineFixture"] = r.get("machine_name", "")
            r["modelSeries"]    = r.get("model_series", "")
            r["oldModelNo"]     = r.get("old_model_no", "")
            r["pyName"]         = r.get("py_name", "")
        return rows


@app.post("/api/poka-yoke/assignments/")
def py_add_assignment(body: dict, user=Depends(get_current_user)):
    with get_conn() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            INSERT INTO mes_py_assignments
                (py_no, py_name, side, model_type, model_name, model_series,
                 old_model_no, d_bit, desired_value, machine_name)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
        """, (
            body.get("pyNo"), body.get("pyName"), body.get("typeSide"),
            body.get("modelType"), body.get("modelName"), body.get("modelSeries"),
            body.get("oldModelNo"), body.get("dBit"),
            body.get("desiredValue") if body.get("desiredValue") not in ("", None) else None,
            body.get("machineFixture")
        ))
        return {"ok": True, "id": cur.fetchone()["id"]}


@app.delete("/api/poka-yoke/assignments/{aid}")
def py_delete_assignment(aid: int, user=Depends(get_current_user)):
    with get_conn() as conn:
        dict_cursor(conn).execute(
            "DELETE FROM mes_py_assignments WHERE id=%s", (aid,))
    return {"ok": True}


# ── BULK IMPORT ───────────────────────────────────────────────
@app.post("/api/poka-yoke/import/bulk")
def py_bulk_import(body: BulkImportBody, user=Depends(get_current_user)):
    sheet    = body.sheet.strip().upper()
    rows     = body.rows
    col_map  = body.col_map
    inserted = skipped = 0
    errors   = []

    with get_conn() as conn:
        cur = dict_cursor(conn)

        if "MODEL" in sheet and "MASTER" in sheet:
            for r in rows:
                mn = _col(r, "Model Name", col_map)
                if not mn:
                    skipped += 1
                    continue
                try:
                    cur.execute("""
                        INSERT INTO mes_py_model_master
                            (model_name, model_type, old_model_no, series)
                        VALUES (%s,%s,%s,%s)
                        ON CONFLICT (model_name) DO UPDATE SET
                            model_type   = EXCLUDED.model_type,
                            old_model_no = EXCLUDED.old_model_no,
                            series       = EXCLUDED.series,
                            is_active    = true
                    """, (
                        mn,
                        _col(r, "type", col_map) or None,
                        _col(r, "Old Model No", col_map) or None,
                        _col(r, "model", col_map) or None
                    ))
                    inserted += 1
                except Exception as e:
                    skipped += 1
                    errors.append(str(e))

        elif "POKA" in sheet:
            for r in rows:
                py_no = _col(r, "Poka Yoke No", col_map)
                if not py_no:
                    skipped += 1
                    continue
                try:
                    cur.execute("""
                        INSERT INTO mes_py_master
                            (py_no, description, model_type, side, bit, desired_value, machine_name)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (py_no) DO UPDATE SET
                            description  = EXCLUDED.description,
                            model_type   = EXCLUDED.model_type,
                            machine_name = EXCLUDED.machine_name,
                            is_active    = true
                    """, (
                        py_no,
                        _col(r, "Poka Yoke Name", col_map) or None,
                        _col(r, "Model Type", col_map) or None,
                        _col(r, "Side", col_map) or "ALL",
                        _col(r, "D Bit", col_map) or None,
                        _int(r, "Desired Value", col_map),
                        _col(r, "Machine/Fixture", col_map) or None
                    ))
                    inserted += 1
                except Exception as e:
                    skipped += 1
                    errors.append(str(e))

        elif "FINAL" in sheet or "SEAT" in sheet or "ASSIGN" in sheet:
            for r in rows:
                py_no = _col(r, "Poka Yoke No", col_map)
                mn    = _col(r, "Model Name", col_map)
                if not py_no or not mn:
                    skipped += 1
                    continue
                # Handle both newline and space variants of column name
                side = (
                    _col(r, "Type\nSide", col_map)
                    or _col(r, "Type Side", col_map)
                    or _col(r, "typeSide", col_map)
                    or "ALL"
                )
                if side.upper() in ("L", "LH"):   side = "LH"
                elif side.upper() in ("R", "RH"): side = "RH"
                try:
                    cur.execute(
                        "SELECT id FROM mes_py_master WHERE py_no=%s AND is_active=true LIMIT 1",
                        (py_no,))
                    py_row = cur.fetchone()
                    cur.execute(
                        "SELECT id FROM mes_py_model_master WHERE model_name=%s AND is_active=true LIMIT 1",
                        (mn,))
                    m_row = cur.fetchone()
                    # Handle both newline and space variants for desired value
                    dval = (
                        _int(r, "Desired Value\n(0/1/2)", col_map)
                        or _int(r, "Desired Value (0/1/2)", col_map)
                        or _int(r, "desiredValue", col_map)
                    )
                    cur.execute("""
                        INSERT INTO mes_py_assignments
                            (py_id, model_id, py_no, py_name, side, model_type,
                             model_name, model_series, old_model_no,
                             d_bit, desired_value, machine_name)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        py_row["id"] if py_row else None,
                        m_row["id"]  if m_row  else None,
                        py_no,
                        _col(r, "Poka Yoke Name", col_map),
                        side,
                        _col(r, "Model Type", col_map),
                        mn,
                        _col(r, "Model", col_map),
                        _col(r, "Old Model No", col_map),
                        _col(r, "D bit From PLC", col_map) or None,
                        dval,
                        _col(r, "Machine/Fixture", col_map)
                    ))
                    inserted += 1
                except Exception as e:
                    skipped += 1
                    errors.append(str(e))
        else:
            raise HTTPException(
                400, f"Unknown sheet: {body.sheet}. Use 'MODEL MASTER', 'POKA YOKE MASTER', or 'final seat'")

    return {
        "ok":       True,
        "sheet":    body.sheet,
        "inserted": inserted,
        "skipped":  skipped,
        "errors":   errors[:10]
    }



# ── Poka-Yoke Bypass Email Notification ───────────────────────────────────────
# Called by Fullscreen when a bypass is first detected.
# Uses smtplib so no extra dependency — configure SMTP in .env or environment vars.
@app.post("/api/poka-yoke/notify-bypass")
async def notify_bypass(request: Request):
    """
    Fire-and-forget email to quality department when a poka-yoke is bypassed.
    No auth required — called from frontend on bypass detection.
    Expected body: { line_id, rule_id, poka_yoke_name, bit, machine_name, alert_level, model_name }
    Configure via environment variables:
      SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, NOTIFY_EMAIL
    """
    import smtplib
    import os
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    body = await request.json()
    smtp_host  = os.getenv("SMTP_HOST",    "smtp.gmail.com")
    smtp_port  = int(os.getenv("SMTP_PORT", "587"))
    smtp_user  = os.getenv("SMTP_USER",    "")
    smtp_pass  = os.getenv("SMTP_PASS",    "")
    notify_to  = os.getenv("NOTIFY_EMAIL", "")

    if not smtp_user or not notify_to:
        # SMTP not configured — log and return silently
        print(f"[BYPASS] Email not configured. Bypass: {body.get('poka_yoke_name')} on line {body.get('line_id')}")
        return {"ok": True, "sent": False, "reason": "SMTP not configured"}

    try:
        # Get line name for context
        line_name = body.get("line_id", "Unknown Line")
        try:
            with get_conn() as conn:
                cur = dict_cursor(conn)
                cur.execute("SELECT line_name FROM mes_lines WHERE id = %s", (body.get("line_id"),))
                row = cur.fetchone()
                if row: line_name = row["line_name"]
        except Exception:
            pass

        level      = body.get("alert_level", "WARNING")
        py_name    = body.get("poka_yoke_name", "Unknown")
        bit        = body.get("bit", "—")
        machine    = body.get("machine_name", "—")
        model      = body.get("model_name", "—")
        subject    = f"[{level}] Poka-Yoke Bypassed — {line_name}"

        html = f"""
<html><body style="font-family:Arial,sans-serif;color:#0f172a;">
  <div style="border-left:5px solid {"#ef4444" if level=="CRITICAL" else "#f59e0b"};padding:16px 20px;background:#fff;">
    <h2 style="margin:0 0 8px;color:{"#ef4444" if level=="CRITICAL" else "#f59e0b"};">
      {"🚨 CRITICAL" if level=="CRITICAL" else "⚠️ WARNING"} — Poka-Yoke Bypass Detected
    </h2>
    <table style="border-collapse:collapse;width:100%;margin-top:12px;">
      <tr><td style="padding:6px 12px;font-weight:700;background:#f8fafc;width:140px;">Line</td><td style="padding:6px 12px;">{line_name}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f1f5f9;">Check Name</td><td style="padding:6px 12px;">{py_name}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f8fafc;">Bit / Register</td><td style="padding:6px 12px;font-family:monospace;">{bit}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f1f5f9;">Machine</td><td style="padding:6px 12px;">{machine}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f8fafc;">Model</td><td style="padding:6px 12px;">{model}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f1f5f9;">Alert Level</td>
          <td style="padding:6px 12px;color:{"#ef4444" if level=="CRITICAL" else "#f59e0b"};font-weight:700;">{level}</td></tr>
      <tr><td style="padding:6px 12px;font-weight:700;background:#f8fafc;">Time</td><td style="padding:6px 12px;">{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</td></tr>
    </table>
    <p style="margin-top:16px;font-size:13px;color:#64748b;">
      Please investigate and acknowledge the bypass in the Fullscreen dashboard.<br/>
      This is an automated alert from the Production Monitoring System.
    </p>
  </div>
</body></html>"""

        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = smtp_user
        msg["To"]      = notify_to
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, notify_to.split(","), msg.as_string())

        print(f"[BYPASS EMAIL] Sent: {py_name} — {line_name} → {notify_to}")
        return {"ok": True, "sent": True}

    except Exception as e:
        print(f"[BYPASS EMAIL] Failed: {e}")
        return {"ok": True, "sent": False, "reason": str(e)}

# ── Entry point ────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n" + "=" * 65)
    print("  Platform — Toyota Boshoku Device India Pvt. Ltd.")
    print("=" * 65)
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=False, workers=1)